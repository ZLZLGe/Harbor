#!/usr/bin/env python3

from __future__ import annotations

from datetime import date, time

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

WORKBOOK_PATH = "/root/clinic_staffing.xlsx"

TITLE_FILL = PatternFill("solid", fgColor="244062")
HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")

AVAILABILITY_ROWS = [
    ("S-001", "Nina Patel", date(2026, 4, 15), time(8, 0), time(12, 0), "Compliance training"),
    ("S-005", "Maya Gomez", date(2026, 4, 15), time(10, 0), time(12, 0), "Provider meeting"),
    ("S-002", "Omar Reed", date(2026, 4, 14), time(7, 0), time(8, 0), "Commute delay"),
]

ASSIGNMENT_ROWS = [
    ("CL-1001", date(2026, 4, 14), time(8, 0), time(12, 0), "Exam 1", "Nurse", "S-001", "Nina Patel"),
    ("CL-1002", date(2026, 4, 14), time(8, 0), time(12, 0), "Exam 2", "Nurse", "S-002", "Omar Reed"),
    ("CL-1003", date(2026, 4, 14), time(9, 30), time(13, 0), "Lab A", "Tech", "S-003", "Ava Lin"),
    ("CL-1004", date(2026, 4, 14), time(12, 0), time(16, 0), "Imaging", "Tech", "S-003", "Ava Lin"),
    ("CL-1005", date(2026, 4, 14), time(13, 0), time(17, 0), "Triage", "Nurse", "S-004", "Leo Chen"),
    ("CL-1006", date(2026, 4, 15), time(8, 0), time(12, 0), "Pediatrics", "RN", "S-005", "Maya Gomez"),
    ("CL-1007", date(2026, 4, 15), time(8, 0), time(12, 0), "Exam 1", "RN", "S-001", "Nina Patel"),
    ("CL-1008", date(2026, 4, 15), time(13, 0), time(17, 0), "Lab A", "Tech", "S-002", "Omar Reed"),
    ("CL-1008", date(2026, 4, 15), time(13, 0), time(17, 0), "Lab A", "Tech", "S-004", "Leo Chen"),
    ("CL-1009", date(2026, 4, 15), time(13, 0), time(17, 0), "Imaging", "Tech", "S-004", "Leo Chen"),
]

COVERAGE_ROWS = [
    ("CL-1001", date(2026, 4, 14), time(8, 0), time(12, 0), "Exam 1", 1),
    ("CL-1002", date(2026, 4, 14), time(8, 0), time(12, 0), "Exam 2", 1),
    ("CL-1003", date(2026, 4, 14), time(9, 30), time(13, 0), "Lab A", 1),
    ("CL-1004", date(2026, 4, 14), time(12, 0), time(16, 0), "Imaging", 1),
    ("CL-1005", date(2026, 4, 14), time(13, 0), time(17, 0), "Triage", 2),
    ("CL-1006", date(2026, 4, 15), time(8, 0), time(12, 0), "Pediatrics", 1),
    ("CL-1007", date(2026, 4, 15), time(8, 0), time(12, 0), "Exam 1", 1),
    ("CL-1008", date(2026, 4, 15), time(13, 0), time(17, 0), "Lab A", 2),
    ("CL-1009", date(2026, 4, 15), time(13, 0), time(17, 0), "Imaging", 1),
]


def style_title(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = TITLE_FILL


def style_headers(sheet, row_number: int, columns: int) -> None:
    for column in range(1, columns + 1):
        cell = sheet.cell(row=row_number, column=column)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL


def apply_date_time_formats(sheet, start_row: int, end_row: int, date_col: str, start_col: str, end_col: str) -> None:
    for row in range(start_row, end_row + 1):
        sheet[f"{date_col}{row}"].number_format = "yyyy-mm-dd"
        sheet[f"{start_col}{row}"].number_format = "hh:mm"
        sheet[f"{end_col}{row}"].number_format = "hh:mm"


def build_availability_sheet(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "Availability"
    sheet["A1"] = "Clinic Staff Availability"
    style_title(sheet["A1"])
    sheet["A2"] = "Rows list blackout periods when staff cannot cover a shift."
    headers = ["Staff ID", "Staff Name", "Date", "Start Time", "End Time", "Reason"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=4, column=column, value=header)
    style_headers(sheet, 4, len(headers))

    for row_number, row in enumerate(AVAILABILITY_ROWS, start=5):
        for column, value in enumerate(row, start=1):
            sheet.cell(row=row_number, column=column, value=value)

    apply_date_time_formats(sheet, 5, 4 + len(AVAILABILITY_ROWS), "C", "D", "E")
    for column, width in {"A": 12, "B": 18, "C": 14, "D": 12, "E": 12, "F": 24}.items():
        sheet.column_dimensions[column].width = width


def build_assignments_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Assignments")
    sheet["A1"] = "Assigned Clinic Shifts"
    style_title(sheet["A1"])
    sheet["A2"] = "Header row starts after the planning notes."
    sheet["A3"] = "Only listed assignments count toward room coverage."
    headers = ["Shift ID", "Date", "Start Time", "End Time", "Room", "Role", "Staff ID", "Staff Name"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=5, column=column, value=header)
    style_headers(sheet, 5, len(headers))

    for row_number, row in enumerate(ASSIGNMENT_ROWS, start=6):
        for column, value in enumerate(row, start=1):
            sheet.cell(row=row_number, column=column, value=value)

    apply_date_time_formats(sheet, 6, 5 + len(ASSIGNMENT_ROWS), "B", "C", "D")
    for column, width in {"A": 12, "B": 14, "C": 12, "D": 12, "E": 14, "F": 12, "G": 12, "H": 18}.items():
        sheet.column_dimensions[column].width = width


def build_coverage_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Room Coverage")
    sheet["A1"] = "Required Room Coverage"
    style_title(sheet["A1"])
    sheet["A2"] = "Compare this table with the assignment rows by Shift ID."
    headers = ["Shift ID", "Date", "Start Time", "End Time", "Room", "Required Staff"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=6, column=column, value=header)
    style_headers(sheet, 6, len(headers))

    for row_number, row in enumerate(COVERAGE_ROWS, start=7):
        for column, value in enumerate(row, start=1):
            sheet.cell(row=row_number, column=column, value=value)

    apply_date_time_formats(sheet, 7, 6 + len(COVERAGE_ROWS), "B", "C", "D")
    for column, width in {"A": 12, "B": 14, "C": 12, "D": 12, "E": 14, "F": 16}.items():
        sheet.column_dimensions[column].width = width


def main() -> None:
    workbook = Workbook()
    build_availability_sheet(workbook)
    build_assignments_sheet(workbook)
    build_coverage_sheet(workbook)
    workbook.save(WORKBOOK_PATH)


if __name__ == "__main__":
    main()
