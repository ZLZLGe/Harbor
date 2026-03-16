import openpyxl
import pytest
from pathlib import Path


OUTPUT_FILE = Path("warehouse_sla_completed.xlsx")
SHEET_NAMES = {
    "budget": "Throughput Plan",
    "yoy": "SLA Changes (%)",
    "shares": "Lane Shares (%)",
    "review": "Capacity Review",
}

BUDGET_VALUES = [
    ("F8", 229, "FY2023 Pick-Pack"),
    ("K5", 2889, "FY2020 Total"),
    ("B9", 935, "FY2024 Same-Day"),
    ("C12", 816, "FY2027 Next-Day"),
    ("K10", 3379, "FY2025 Total"),
    ("E10", 218, "FY2025 Returns"),
]

YOY_VALUES = [
    ("D7", -1.88, "FY2023 Standard YoY"),
    ("B9", 1.93, "FY2025 Same-Day YoY"),
    ("F9", 0.0, "FY2025 Pick-Pack YoY"),
]

SHARE_VALUES = [
    ("F5", 6.68, "FY2020 Pick-Pack Share"),
    ("B10", 28.2, "FY2025 Same-Day Share"),
]

REVIEW_VALUES = [
    ("B7", 168, "5-Year Same-Day Change"),
    ("B8", 987.2, "Average Same-Day Throughput"),
    ("E4", 1.4, "Returns CAGR"),
    ("E5", 208, "FY2023 Returns"),
]


@pytest.fixture
def workbook():
    if not OUTPUT_FILE.exists():
        pytest.skip(f"{OUTPUT_FILE.name} not found")
    return openpyxl.load_workbook(OUTPUT_FILE)


@pytest.fixture
def budget_sheet(workbook):
    return workbook[SHEET_NAMES["budget"]]


@pytest.fixture
def yoy_sheet(workbook):
    return workbook[SHEET_NAMES["yoy"]]


@pytest.fixture
def share_sheet(workbook):
    return workbook[SHEET_NAMES["shares"]]


@pytest.fixture
def review_sheet(workbook):
    return workbook[SHEET_NAMES["review"]]


def test_file_exists():
    assert OUTPUT_FILE.exists()


@pytest.mark.parametrize("cell,expected,label", BUDGET_VALUES)
def test_budget_values(budget_sheet, cell, expected, label):
    actual = budget_sheet[cell].value
    assert actual == expected, f"{label} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,label", YOY_VALUES)
def test_yoy_values(yoy_sheet, cell, expected, label):
    actual = yoy_sheet[cell].value
    assert abs(actual - expected) < 0.1, f"{label} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,label", SHARE_VALUES)
def test_share_values(share_sheet, cell, expected, label):
    actual = share_sheet[cell].value
    assert abs(actual - expected) < 0.1, f"{label} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,label", REVIEW_VALUES)
def test_review_values(review_sheet, cell, expected, label):
    actual = review_sheet[cell].value
    if isinstance(expected, int):
        assert actual == expected, f"{label} ({cell}): expected {expected}, got {actual}"
    else:
        assert abs(actual - expected) < 0.1, f"{label} ({cell}): expected {expected}, got {actual}"


def test_no_remaining_placeholders(workbook):
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                if value == "???":
                    pytest.fail(f"Found placeholder in {sheet_name}")


def test_row_sums_consistent(budget_sheet):
    row5_sum = sum(budget_sheet.cell(row=5, column=col).value for col in range(2, 11))
    assert row5_sum == budget_sheet["K5"].value

    row10_sum = sum(budget_sheet.cell(row=10, column=col).value for col in range(2, 11))
    assert row10_sum == budget_sheet["K10"].value


def test_cross_sheet_consistency(budget_sheet, review_sheet):
    start = budget_sheet["E8"].value
    end = budget_sheet["E13"].value
    expected_cagr = round(((end / start) ** 0.2 - 1) * 100, 2)
    assert abs(review_sheet["E4"].value - expected_cagr) < 0.1
    assert review_sheet["E5"].value == budget_sheet["E8"].value
