import openpyxl
import pytest
from pathlib import Path


OUTPUT_FILE = Path("greenhouse_harvest_rebuilt.xlsx")
SHEET_NAMES = {
    "budget": "Harvest Plan",
    "yoy": "Season Changes (%)",
    "shares": "Zone Shares (%)",
    "review": "Yield Review",
}

BUDGET_VALUES = [
    ("F8", 99, "FY2025 Seedling Bay"),
    ("K5", 1103, "FY2022 Total"),
    ("B9", 326, "FY2026 North House"),
    ("C12", 378, "FY2029 South House"),
    ("K10", 1332, "FY2027 Total"),
    ("E10", 89, "FY2027 West House"),
]

YOY_VALUES = [
    ("D7", -2.38, "FY2025 East House YoY"),
    ("B9", 1.84, "FY2027 North House YoY"),
    ("F9", 0.91, "FY2027 Seedling Bay YoY"),
]

SHARE_VALUES = [
    ("F5", 8.07, "FY2022 Seedling Bay Share"),
    ("B10", 24.92, "FY2027 North House Share"),
]

REVIEW_VALUES = [
    ("B7", 62, "5-Year North House Change"),
    ("B8", 344.5, "Average North House Yield"),
    ("E4", 1.37, "West House CAGR"),
    ("E5", 85, "FY2025 West House"),
]


@pytest.fixture
def workbook():
    if not OUTPUT_FILE.exists():
        pytest.skip(f"{OUTPUT_FILE.name} not found")
    return openpyxl.load_workbook(OUTPUT_FILE)


@pytest.fixture
def budget_sheet(workbook):
    return workbook[SHEET_NAMES["budget"]]


@pytest.fixture
def yoy_sheet(workbook):
    return workbook[SHEET_NAMES["yoy"]]


@pytest.fixture
def share_sheet(workbook):
    return workbook[SHEET_NAMES["shares"]]


@pytest.fixture
def review_sheet(workbook):
    return workbook[SHEET_NAMES["review"]]


def test_file_exists():
    assert OUTPUT_FILE.exists()


@pytest.mark.parametrize("cell,expected,label", BUDGET_VALUES)
def test_budget_values(budget_sheet, cell, expected, label):
    actual = budget_sheet[cell].value
    assert actual == expected, f"{label} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,label", YOY_VALUES)
def test_yoy_values(yoy_sheet, cell, expected, label):
    actual = yoy_sheet[cell].value
    assert abs(actual - expected) < 0.1, f"{label} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,label", SHARE_VALUES)
def test_share_values(share_sheet, cell, expected, label):
    actual = share_sheet[cell].value
    assert abs(actual - expected) < 0.1, f"{label} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,label", REVIEW_VALUES)
def test_review_values(review_sheet, cell, expected, label):
    actual = review_sheet[cell].value
    if isinstance(expected, int):
        assert actual == expected, f"{label} ({cell}): expected {expected}, got {actual}"
    else:
        assert abs(actual - expected) < 0.1, f"{label} ({cell}): expected {expected}, got {actual}"


def test_no_remaining_placeholders(workbook):
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                if value == "???":
                    pytest.fail(f"Found placeholder in {sheet_name}")


def test_row_sums_consistent(budget_sheet):
    row5_sum = sum(budget_sheet.cell(row=5, column=col).value for col in range(2, 11))
    assert row5_sum == budget_sheet["K5"].value

    row10_sum = sum(budget_sheet.cell(row=10, column=col).value for col in range(2, 11))
    assert row10_sum == budget_sheet["K10"].value


def test_cross_sheet_consistency(budget_sheet, review_sheet):
    start = budget_sheet["E8"].value
    end = budget_sheet["E13"].value
    expected_cagr = round(((end / start) ** 0.2 - 1) * 100, 2)
    assert abs(review_sheet["E4"].value - expected_cagr) < 0.1
    assert review_sheet["E5"].value == budget_sheet["E8"].value
