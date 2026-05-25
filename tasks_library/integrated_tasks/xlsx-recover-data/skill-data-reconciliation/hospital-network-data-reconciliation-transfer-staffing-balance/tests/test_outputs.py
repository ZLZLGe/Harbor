from pathlib import Path

import openpyxl
import pytest


@pytest.fixture
def workbook():
    path = Path("hospital_staffing_reconciled.xlsx")
    if not path.exists():
        pytest.skip("hospital_staffing_reconciled.xlsx not found")
    return openpyxl.load_workbook(path)


@pytest.fixture
def staffing_sheet(workbook):
    return workbook["Quarterly Staffing"]


@pytest.fixture
def movement_sheet(workbook):
    return workbook["Recruitment & Attrition"]


@pytest.fixture
def shares_sheet(workbook):
    return workbook["FTE Shares (%)"]


@pytest.fixture
def overview_sheet(workbook):
    return workbook["Annual Workforce Overview"]


def test_file_exists():
    assert Path("hospital_staffing_reconciled.xlsx").exists()


STAFFING_VALUES = [
    ("F5", 55, "Q1 support FTE"),
    ("G6", 432, "Q2 total FTE"),
    ("C7", 101, "Q3 surgery FTE"),
    ("B8", 133, "Q4 emergency FTE"),
]

MOVEMENT_VALUES = [
    ("C5", 24, "Q1 hires"),
    ("E6", 432, "Q2 ending FTE"),
    ("B7", 432, "Q3 opening FTE"),
    ("D7", 7, "Q3 attrition"),
    ("F8", 3.36, "Q4 net change percent"),
]

SHARE_VALUES = [
    ("F5", 13.1, "Q1 support share"),
    ("D6", 18.06, "Q2 pediatrics share"),
    ("E7", 17.0, "Q3 diagnostics share"),
]

OVERVIEW_VALUES = [
    ("G6", 9, "Surgery net change"),
    ("F7", 81.25, "Pediatrics average quarterly FTE"),
    ("B10", 420, "Total Q1 FTE"),
]


@pytest.mark.parametrize("cell,expected,desc", STAFFING_VALUES)
def test_staffing_values(staffing_sheet, cell, expected, desc):
    actual = staffing_sheet[cell].value
    assert actual == expected, f"{desc} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,desc", MOVEMENT_VALUES)
def test_movement_values(movement_sheet, cell, expected, desc):
    actual = movement_sheet[cell].value
    if isinstance(expected, int):
        assert actual == expected, f"{desc} ({cell}): expected {expected}, got {actual}"
    else:
        assert abs(actual - expected) < 0.01, f"{desc} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,desc", SHARE_VALUES)
def test_share_values(shares_sheet, cell, expected, desc):
    actual = shares_sheet[cell].value
    assert abs(actual - expected) < 0.01, f"{desc} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,desc", OVERVIEW_VALUES)
def test_overview_values(overview_sheet, cell, expected, desc):
    actual = overview_sheet[cell].value
    if isinstance(expected, int):
        assert actual == expected, f"{desc} ({cell}): expected {expected}, got {actual}"
    else:
        assert abs(actual - expected) < 0.01, f"{desc} ({cell}): expected {expected}, got {actual}"


def test_no_remaining_placeholders(workbook):
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                assert value != "???", f"Found unresolved placeholder in {sheet_name}"


def test_staffing_row_sums(staffing_sheet):
    for row in range(5, 9):
        row_sum = sum(staffing_sheet.cell(row=row, column=col).value for col in range(2, 7))
        assert row_sum == staffing_sheet.cell(row=row, column=7).value


def test_movement_continuity(movement_sheet):
    for row in range(5, 9):
        expected_ending = (
            movement_sheet.cell(row=row, column=2).value
            + movement_sheet.cell(row=row, column=3).value
            - movement_sheet.cell(row=row, column=4).value
        )
        assert expected_ending == movement_sheet.cell(row=row, column=5).value

    for row in range(6, 9):
        assert movement_sheet.cell(row=row, column=2).value == movement_sheet.cell(row=row - 1, column=5).value


def test_share_consistency(staffing_sheet, shares_sheet):
    for staffing_row, shares_row in zip(range(5, 9), range(5, 9)):
        total = staffing_sheet.cell(row=staffing_row, column=7).value
        for col in range(2, 7):
            staffing_value = staffing_sheet.cell(row=staffing_row, column=col).value
            share_value = shares_sheet.cell(row=shares_row, column=col).value
            expected_share = round(staffing_value / total * 100, 2)
            assert abs(share_value - expected_share) < 0.01


def test_overview_consistency(staffing_sheet, shares_sheet, overview_sheet):
    department_rows = {
        5: ("B5", "B6", "B7", "B8", "B8"),
        6: ("C5", "C6", "C7", "C8", "C8"),
        7: ("D5", "D6", "D7", "D8", "D8"),
        8: ("E5", "E6", "E7", "E8", "E8"),
        9: ("F5", "F6", "F7", "F8", "F8"),
    }

    for overview_row, refs in department_rows.items():
        q1_ref, q2_ref, q3_ref, q4_ref, share_ref = refs
        q1 = staffing_sheet[q1_ref].value
        q2 = staffing_sheet[q2_ref].value
        q3 = staffing_sheet[q3_ref].value
        q4 = staffing_sheet[q4_ref].value
        assert overview_sheet.cell(row=overview_row, column=2).value == q1
        assert overview_sheet.cell(row=overview_row, column=3).value == q2
        assert overview_sheet.cell(row=overview_row, column=4).value == q3
        assert overview_sheet.cell(row=overview_row, column=5).value == q4
        expected_avg = round((q1 + q2 + q3 + q4) / 4, 2)
        assert abs(overview_sheet.cell(row=overview_row, column=6).value - expected_avg) < 0.01
        assert overview_sheet.cell(row=overview_row, column=7).value == q4 - q1
        assert abs(overview_sheet.cell(row=overview_row, column=8).value - shares_sheet[share_ref].value) < 0.01

    total_qs = [staffing_sheet[f"G{row}"].value for row in range(5, 9)]
    assert overview_sheet["B10"].value == total_qs[0]
    assert overview_sheet["C10"].value == total_qs[1]
    assert overview_sheet["D10"].value == total_qs[2]
    assert overview_sheet["E10"].value == total_qs[3]
    assert abs(overview_sheet["F10"].value - round(sum(total_qs) / 4, 2)) < 0.01
    assert overview_sheet["G10"].value == total_qs[3] - total_qs[0]
    assert abs(overview_sheet["H10"].value - 100.0) < 0.01
