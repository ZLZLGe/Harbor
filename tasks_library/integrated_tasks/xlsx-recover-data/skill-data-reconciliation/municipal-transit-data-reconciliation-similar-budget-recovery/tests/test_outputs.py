from pathlib import Path

import openpyxl
import pytest


@pytest.fixture
def workbook():
    path = Path("transit_budget_recovered.xlsx")
    if not path.exists():
        pytest.skip("transit_budget_recovered.xlsx not found")
    return openpyxl.load_workbook(path)


@pytest.fixture
def budget_sheet(workbook):
    return workbook["Operating Budget by Division"]


@pytest.fixture
def yoy_sheet(workbook):
    return workbook["YoY Changes (%)"]


@pytest.fixture
def shares_sheet(workbook):
    return workbook["Division Shares (%)"]


@pytest.fixture
def growth_sheet(workbook):
    return workbook["Growth Analysis"]


def test_file_exists():
    assert Path("transit_budget_recovered.xlsx").exists()


BUDGET_VALUES = [
    ("K6", 3234, "FY2018 total"),
    ("F8", 485, "FY2020 passenger services"),
    ("B9", 989, "FY2021 rail operations"),
    ("E10", 754, "FY2022 fleet maintenance"),
    ("K10", 3885, "FY2022 total"),
    ("C12", 319, "FY2024 signal and power"),
]

YOY_VALUES = [
    ("D7", 10.24, "FY2020 station services YoY"),
    ("F8", 4.54, "FY2021 passenger services YoY"),
    ("B9", 2.93, "FY2022 rail operations YoY"),
]

SHARE_VALUES = [
    ("F5", 14.24, "FY2017 passenger services share"),
    ("B10", 26.20, "FY2022 rail operations share"),
]

GROWTH_VALUES = [
    ("E4", 5.73, "Fleet maintenance CAGR"),
    ("E5", 675, "FY2020 fleet maintenance"),
    ("H7", 91, "Infrastructure renewal 5-year change"),
    ("B8", 1037.5, "Rail operations average annual budget"),
]


@pytest.mark.parametrize("cell,expected,desc", BUDGET_VALUES)
def test_budget_values(budget_sheet, cell, expected, desc):
    actual = budget_sheet[cell].value
    assert actual == expected, f"{desc} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,desc", YOY_VALUES)
def test_yoy_values(yoy_sheet, cell, expected, desc):
    actual = yoy_sheet[cell].value
    assert abs(actual - expected) < 0.01, f"{desc} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,desc", SHARE_VALUES)
def test_share_values(shares_sheet, cell, expected, desc):
    actual = shares_sheet[cell].value
    assert abs(actual - expected) < 0.01, f"{desc} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,desc", GROWTH_VALUES)
def test_growth_values(growth_sheet, cell, expected, desc):
    actual = growth_sheet[cell].value
    if isinstance(expected, int):
        assert actual == expected, f"{desc} ({cell}): expected {expected}, got {actual}"
    else:
        assert abs(actual - expected) < 0.01, f"{desc} ({cell}): expected {expected}, got {actual}"


def test_no_remaining_placeholders(workbook):
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                assert value != "???", f"Found unresolved placeholder in {sheet_name}"


def test_budget_row_sums(budget_sheet):
    row6_sum = sum(budget_sheet.cell(row=6, column=col).value for col in range(2, 11))
    row10_sum = sum(budget_sheet.cell(row=10, column=col).value for col in range(2, 11))
    assert row6_sum == budget_sheet["K6"].value
    assert row10_sum == budget_sheet["K10"].value


def test_cross_sheet_consistency(budget_sheet, shares_sheet, yoy_sheet, growth_sheet):
    expected_share = round(budget_sheet["B10"].value / budget_sheet["K10"].value * 100, 2)
    assert abs(shares_sheet["B10"].value - expected_share) < 0.01

    expected_yoy = round((budget_sheet["F9"].value - budget_sheet["F8"].value) / budget_sheet["F8"].value * 100, 2)
    assert abs(yoy_sheet["F8"].value - expected_yoy) < 0.01

    expected_cagr = round(((budget_sheet["E13"].value / budget_sheet["E8"].value) ** 0.2 - 1) * 100, 2)
    assert abs(growth_sheet["E4"].value - expected_cagr) < 0.01
