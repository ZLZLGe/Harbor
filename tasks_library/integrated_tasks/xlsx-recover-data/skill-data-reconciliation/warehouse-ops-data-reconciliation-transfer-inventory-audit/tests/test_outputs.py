from pathlib import Path

import openpyxl
import pytest


@pytest.fixture
def workbook():
    path = Path("warehouse_inventory_reconciled.xlsx")
    if not path.exists():
        pytest.skip("warehouse_inventory_reconciled.xlsx not found")
    return openpyxl.load_workbook(path)


@pytest.fixture
def inventory_sheet(workbook):
    return workbook["Ending Inventory by Warehouse"]


@pytest.fixture
def flow_sheet(workbook):
    return workbook["Inbound-Outbound Flow"]


@pytest.fixture
def shares_sheet(workbook):
    return workbook["Warehouse Shares (%)"]


@pytest.fixture
def shrink_sheet(workbook):
    return workbook["Shrinkage Analysis"]


def test_file_exists():
    assert Path("warehouse_inventory_reconciled.xlsx").exists()


INVENTORY_VALUES = [
    ("F5", 250, "Jan overflow inventory"),
    ("G7", 4450, "Mar total ending inventory"),
    ("C8", 1080, "Apr south warehouse inventory"),
    ("G9", 4950, "May total ending inventory"),
    ("B9", 1450, "May north warehouse inventory"),
    ("E10", 900, "Jun west warehouse inventory"),
]

FLOW_VALUES = [
    ("F6", 4200, "Feb ending inventory"),
    ("B7", 4200, "Mar opening inventory"),
    ("F9", 4950, "May ending inventory"),
    ("C10", 840, "Jun inbound units"),
]

SHARE_VALUES = [
    ("F5", 6.25, "Jan overflow share"),
    ("C7", 22.92, "Mar south share"),
    ("D10", 22.5, "Jun east share"),
]

SHRINK_VALUES = [
    ("D8", 310, "East net change"),
    ("B9", 1355.0, "North average ending inventory"),
]


@pytest.mark.parametrize("cell,expected,desc", INVENTORY_VALUES)
def test_inventory_values(inventory_sheet, cell, expected, desc):
    actual = inventory_sheet[cell].value
    assert actual == expected, f"{desc} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,desc", FLOW_VALUES)
def test_flow_values(flow_sheet, cell, expected, desc):
    actual = flow_sheet[cell].value
    assert actual == expected, f"{desc} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,desc", SHARE_VALUES)
def test_share_values(shares_sheet, cell, expected, desc):
    actual = shares_sheet[cell].value
    assert abs(actual - expected) < 0.01, f"{desc} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,desc", SHRINK_VALUES)
def test_shrink_values(shrink_sheet, cell, expected, desc):
    actual = shrink_sheet[cell].value
    if isinstance(expected, int):
        assert actual == expected, f"{desc} ({cell}): expected {expected}, got {actual}"
    else:
        assert abs(actual - expected) < 0.01, f"{desc} ({cell}): expected {expected}, got {actual}"


def test_no_remaining_placeholders(workbook):
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                assert value != "???", f"Found unresolved placeholder in {sheet_name}"


def test_inventory_row_sums(inventory_sheet):
    for row in (5, 7, 8, 9, 10):
        row_sum = sum(inventory_sheet.cell(row=row, column=col).value for col in range(2, 7))
        assert row_sum == inventory_sheet.cell(row=row, column=7).value


def test_flow_continuity(flow_sheet):
    for row in range(5, 11):
        ending = flow_sheet.cell(row=row, column=2).value + flow_sheet.cell(row=row, column=3).value
        ending -= flow_sheet.cell(row=row, column=4).value + flow_sheet.cell(row=row, column=5).value
        assert ending == flow_sheet.cell(row=row, column=6).value

    for row in range(6, 11):
        assert flow_sheet.cell(row=row, column=2).value == flow_sheet.cell(row=row - 1, column=6).value


def test_cross_sheet_consistency(inventory_sheet, shares_sheet, shrink_sheet):
    expected_share = round(inventory_sheet["F5"].value / inventory_sheet["G5"].value * 100, 2)
    assert abs(shares_sheet["F5"].value - expected_share) < 0.01

    expected_jun_east_share = round(inventory_sheet["D10"].value / inventory_sheet["G10"].value * 100, 2)
    assert abs(shares_sheet["D10"].value - expected_jun_east_share) < 0.01

    expected_net_change = inventory_sheet["D10"].value - inventory_sheet["D5"].value
    assert shrink_sheet["D8"].value == expected_net_change

    expected_west_rate = round(shrink_sheet["E10"].value / shrink_sheet["E9"].value * 100, 2)
    assert abs(shrink_sheet["E5"].value - expected_west_rate) < 0.01
