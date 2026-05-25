from pathlib import Path

import openpyxl
import pytest


@pytest.fixture
def workbook():
    path = Path("solar_generation_reconciled.xlsx")
    if not path.exists():
        pytest.skip("solar_generation_reconciled.xlsx not found")
    return openpyxl.load_workbook(path)


@pytest.fixture
def generation_sheet(workbook):
    return workbook["Monthly Generation (MWh)"]


@pytest.fixture
def outage_sheet(workbook):
    return workbook["Outage Impact (%)"]


@pytest.fixture
def shares_sheet(workbook):
    return workbook["Site Contribution (%)"]


@pytest.fixture
def annual_sheet(workbook):
    return workbook["Annual Capacity Review"]


def test_file_exists():
    assert Path("solar_generation_reconciled.xlsx").exists()


GENERATION_VALUES = [
    ("C6", 1045, "Feb Desert Bloom generation"),
    ("D7", 846, "Mar Coastal Ray generation"),
    ("E8", 720, "Apr Highland Peak generation"),
    ("B10", 4953, "Sun Valley four-month total"),
    ("F10", 15339, "Fleet four-month total"),
]

OUTAGE_VALUES = [
    ("D5", 5.00, "Jan Coastal Ray outage"),
    ("B6", 7.00, "Feb Sun Valley outage"),
    ("E7", 6.00, "Mar Highland Peak outage"),
    ("F8", 4.91, "Apr fleet outage"),
]

SHARE_VALUES = [
    ("B5", 32.61, "Jan Sun Valley share"),
    ("C6", 27.55, "Feb Desert Bloom share"),
    ("D7", 21.87, "Mar Coastal Ray share"),
    ("E8", 18.70, "Apr Highland Peak share"),
]

ANNUAL_VALUES = [
    ("E5", 4953, "Sun Valley four-month actual"),
    ("F6", 242, "Desert Bloom outage loss"),
    ("I7", 22.41, "Coastal Ray fleet share"),
    ("G8", 93.00, "Highland Peak utilization"),
    ("H9", 3834.75, "Fleet average monthly generation"),
]


@pytest.mark.parametrize("cell,expected,desc", GENERATION_VALUES)
def test_generation_values(generation_sheet, cell, expected, desc):
    actual = generation_sheet[cell].value
    assert actual == expected, f"{desc} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,desc", OUTAGE_VALUES)
def test_outage_values(outage_sheet, cell, expected, desc):
    actual = outage_sheet[cell].value
    assert abs(actual - expected) < 0.01, f"{desc} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,desc", SHARE_VALUES)
def test_share_values(shares_sheet, cell, expected, desc):
    actual = shares_sheet[cell].value
    assert abs(actual - expected) < 0.01, f"{desc} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,desc", ANNUAL_VALUES)
def test_annual_values(annual_sheet, cell, expected, desc):
    actual = annual_sheet[cell].value
    if isinstance(expected, int):
        assert actual == expected, f"{desc} ({cell}): expected {expected}, got {actual}"
    else:
        assert abs(actual - expected) < 0.01, f"{desc} ({cell}): expected {expected}, got {actual}"


def test_no_remaining_placeholders(workbook):
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                assert value != "???", f"Found unresolved placeholder in {sheet_name}"


def test_generation_rollups(generation_sheet):
    for row in range(5, 9):
        row_sum = sum(generation_sheet[f"{col}{row}"].value for col in "BCDE")
        assert row_sum == generation_sheet[f"F{row}"].value

    for col in "BCDEF":
        col_sum = sum(generation_sheet[f"{col}{row}"].value for row in range(5, 9))
        assert col_sum == generation_sheet[f"{col}10"].value


def test_outage_consistency(generation_sheet, outage_sheet, annual_sheet):
    nominal_cells = {"B": "C5", "C": "C6", "D": "C7", "E": "C8"}

    for row in range(5, 9):
        month_total = generation_sheet[f"F{row}"].value
        fleet_nominal = annual_sheet["C9"].value
        expected_fleet_outage = round((1 - month_total / fleet_nominal) * 100, 2)
        assert abs(outage_sheet[f"F{row}"].value - expected_fleet_outage) < 0.01

        for col, annual_cell in nominal_cells.items():
            actual_generation = generation_sheet[f"{col}{row}"].value
            nominal_monthly = annual_sheet[annual_cell].value
            expected_outage = round((1 - actual_generation / nominal_monthly) * 100, 2)
            assert abs(outage_sheet[f"{col}{row}"].value - expected_outage) < 0.01

    for col in "BCDEF":
        values = [outage_sheet[f"{col}{row}"].value for row in range(5, 9)]
        assert abs(outage_sheet[f"{col}10"].value - round(sum(values) / 4, 2)) < 0.01


def test_share_consistency(generation_sheet, shares_sheet, annual_sheet):
    for row in range(5, 9):
        fleet_total = generation_sheet[f"F{row}"].value
        running_sum = 0.0
        for col in "BCDE":
            expected_share = round(generation_sheet[f"{col}{row}"].value / fleet_total * 100, 2)
            running_sum += shares_sheet[f"{col}{row}"].value
            assert abs(shares_sheet[f"{col}{row}"].value - expected_share) < 0.01
        assert abs(shares_sheet[f"F{row}"].value - 100.0) < 0.01
        assert abs(running_sum - 100.0) < 0.05

    fleet_actual = annual_sheet["E9"].value
    for col, annual_cell in zip("BCDE", ["E5", "E6", "E7", "E8"]):
        expected_four_month_share = round(annual_sheet[annual_cell].value / fleet_actual * 100, 2)
        assert abs(shares_sheet[f"{col}10"].value - expected_four_month_share) < 0.01
    assert abs(shares_sheet["F10"].value - 100.0) < 0.01


def test_annual_review_consistency(generation_sheet, annual_sheet):
    site_columns = {
        5: "B",
        6: "C",
        7: "D",
        8: "E",
    }

    for annual_row, generation_col in site_columns.items():
        four_month_actual = sum(generation_sheet[f"{generation_col}{row}"].value for row in range(5, 9))
        nominal_monthly = annual_sheet[f"C{annual_row}"].value
        nominal_total = nominal_monthly * 4
        assert annual_sheet[f"D{annual_row}"].value == nominal_total
        assert annual_sheet[f"E{annual_row}"].value == four_month_actual
        assert annual_sheet[f"F{annual_row}"].value == nominal_total - four_month_actual
        assert abs(annual_sheet[f"G{annual_row}"].value - round(four_month_actual / nominal_total * 100, 2)) < 0.01
        assert abs(annual_sheet[f"H{annual_row}"].value - round(four_month_actual / 4, 2)) < 0.01

    fleet_actual = sum(annual_sheet[f"E{row}"].value for row in range(5, 9))
    fleet_nominal_monthly = sum(annual_sheet[f"C{row}"].value for row in range(5, 9))
    assert annual_sheet["C9"].value == fleet_nominal_monthly
    assert annual_sheet["D9"].value == fleet_nominal_monthly * 4
    assert annual_sheet["E9"].value == fleet_actual
    assert annual_sheet["F9"].value == annual_sheet["D9"].value - fleet_actual
    assert abs(annual_sheet["G9"].value - round(fleet_actual / annual_sheet["D9"].value * 100, 2)) < 0.01
    assert abs(annual_sheet["H9"].value - round(fleet_actual / 4, 2)) < 0.01
    assert abs(annual_sheet["I9"].value - 100.0) < 0.01
