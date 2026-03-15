from copy import deepcopy

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


PDF_PATH = "/root/archived_badge_packets.pdf"
XLSX_PATH = "/root/current_badge_clearance_workbook.xlsx"

HEADERS = [
    "Employee ID",
    "Employee Name",
    "Badge ID",
    "Access Zone",
    "Clearance Level",
]

ARCHIVED_ROWS = [
    {"Employee ID": "EMP2001", "Employee Name": "Ava Bell", "Badge ID": "BDG-1001", "Access Zone": "Office East", "Clearance Level": "Level 2"},
    {"Employee ID": "EMP2002", "Employee Name": "Brandon Cole", "Badge ID": "BDG-1002", "Access Zone": "Office West", "Clearance Level": "Level 2"},
    {"Employee ID": "EMP2003", "Employee Name": "Carmen Diaz", "Badge ID": "BDG-1003", "Access Zone": "Research Lab", "Clearance Level": "Level 3"},
    {"Employee ID": "EMP2004", "Employee Name": "Derek Ellis", "Badge ID": "BDG-1004", "Access Zone": "Lobby", "Clearance Level": "Level 1"},
    {"Employee ID": "EMP2005", "Employee Name": "Emily Frost", "Badge ID": "BDG-1005", "Access Zone": "Research Lab", "Clearance Level": "Level 3"},
    {"Employee ID": "EMP2006", "Employee Name": "Felix Grant", "Badge ID": "BDG-1006", "Access Zone": "Server Room", "Clearance Level": "Level 4"},
    {"Employee ID": "EMP2007", "Employee Name": "Grace Holt", "Badge ID": "BDG-1007", "Access Zone": "Office East", "Clearance Level": "Level 2"},
    {"Employee ID": "EMP2008", "Employee Name": "Henry Ibarra", "Badge ID": "BDG-1008", "Access Zone": "Research Lab", "Clearance Level": "Level 3"},
    {"Employee ID": "EMP2009", "Employee Name": "Isla Jordan", "Badge ID": "BDG-1009", "Access Zone": "Lobby", "Clearance Level": "Level 1"},
    {"Employee ID": "EMP2010", "Employee Name": "Jonah Kerr", "Badge ID": "BDG-1010", "Access Zone": "Executive Suite", "Clearance Level": "Level 5"},
    {"Employee ID": "EMP2011", "Employee Name": "Kira Lane", "Badge ID": "BDG-1011", "Access Zone": "Office West", "Clearance Level": "Level 2"},
    {"Employee ID": "EMP2012", "Employee Name": "Leo Martin", "Badge ID": "BDG-1012", "Access Zone": "Server Room", "Clearance Level": "Level 4"},
    {"Employee ID": "EMP2013", "Employee Name": "Mina Nolan", "Badge ID": "BDG-1013", "Access Zone": "Research Lab", "Clearance Level": "Level 3"},
    {"Employee ID": "EMP2014", "Employee Name": "Noah Patel", "Badge ID": "BDG-1014", "Access Zone": "Executive Suite", "Clearance Level": "Level 5"},
    {"Employee ID": "EMP2015", "Employee Name": "Olivia Reed", "Badge ID": "BDG-1015", "Access Zone": "Lobby", "Clearance Level": "Level 1"},
]

POLICY_ROWS = [
    {"Access Zone": "Lobby", "Required Clearance": "Level 1"},
    {"Access Zone": "Office East", "Required Clearance": "Level 2"},
    {"Access Zone": "Office West", "Required Clearance": "Level 2"},
    {"Access Zone": "Research Lab", "Required Clearance": "Level 3"},
    {"Access Zone": "Server Room", "Required Clearance": "Level 4"},
    {"Access Zone": "Executive Suite", "Required Clearance": "Level 5"},
]


def build_badge_roster():
    roster = deepcopy(ARCHIVED_ROWS)
    roster = [row for row in roster if row["Employee ID"] != "EMP2012"]

    statuses = {
        "EMP2003": "Removed",
        "EMP2007": "Suspended",
    }

    badge_roster_rows = []
    for row in roster:
        badge_roster_rows.append(
            {
                "Employee ID": row["Employee ID"],
                "Employee Name": row["Employee Name"],
                "Badge ID": row["Badge ID"],
                "Badge Status": statuses.get(row["Employee ID"], "Active"),
            }
        )

    badge_roster_rows.append(
        {
            "Employee ID": "EMP2016",
            "Employee Name": "Parker Stone",
            "Badge ID": "BDG-1016",
            "Badge Status": "Active",
        }
    )
    return badge_roster_rows


def build_zone_assignments():
    zone_updates = {
        "BDG-1002": "Research Lab",
        "BDG-1005": "Server Room",
        "BDG-1009": "Office East",
        "BDG-1014": "Office West",
        "BDG-1016": "Lobby",
    }

    rows = []
    for row in ARCHIVED_ROWS:
        if row["Employee ID"] == "EMP2012":
            continue
        rows.append(
            {
                "Badge ID": row["Badge ID"],
                "Access Zone": zone_updates.get(row["Badge ID"], row["Access Zone"]),
            }
        )

    rows.append({"Badge ID": "BDG-1016", "Access Zone": "Lobby"})
    return rows


def build_clearance_registry():
    clearance_updates = {
        "EMP2002": "Level 3",
        "EMP2004": "Level 2",
        "EMP2005": "Level 3",
        "EMP2008": "Level 4",
        "EMP2009": "Level 2",
        "EMP2010": "Level 4",
        "EMP2014": "Level 5",
        "EMP2016": "Level 1",
    }

    rows = []
    for row in ARCHIVED_ROWS:
        rows.append(
            {
                "Employee ID": row["Employee ID"],
                "Clearance Level": clearance_updates.get(row["Employee ID"], row["Clearance Level"]),
            }
        )

    rows.append({"Employee ID": "EMP2016", "Clearance Level": "Level 1"})
    return rows


def write_excel():
    workbook = Workbook()

    badge_roster_sheet = workbook.active
    badge_roster_sheet.title = "Badge Roster"
    badge_roster_headers = ["Employee ID", "Employee Name", "Badge ID", "Badge Status"]
    badge_roster_sheet.append(badge_roster_headers)
    for row in build_badge_roster():
        badge_roster_sheet.append([row[column] for column in badge_roster_headers])

    zone_sheet = workbook.create_sheet("Zone Assignments")
    zone_headers = ["Badge ID", "Access Zone"]
    zone_sheet.append(zone_headers)
    for row in build_zone_assignments():
        zone_sheet.append([row[column] for column in zone_headers])

    clearance_sheet = workbook.create_sheet("Clearance Registry")
    clearance_headers = ["Employee ID", "Clearance Level"]
    clearance_sheet.append(clearance_headers)
    for row in build_clearance_registry():
        clearance_sheet.append([row[column] for column in clearance_headers])

    policy_sheet = workbook.create_sheet("Policy Matrix")
    policy_headers = ["Access Zone", "Required Clearance"]
    policy_sheet.append(policy_headers)
    for row in POLICY_ROWS:
        policy_sheet.append([row[column] for column in policy_headers])

    for sheet in workbook.worksheets:
        for column_letter in ["A", "B", "C", "D"]:
            sheet.column_dimensions[column_letter].width = 22

    workbook.save(XLSX_PATH)


def table_data(chunk):
    return [HEADERS] + [[row[column] for column in HEADERS] for row in chunk]


def write_pdf():
    doc = SimpleDocTemplate(
        PDF_PATH,
        pagesize=landscape(letter),
        leftMargin=24,
        rightMargin=24,
        topMargin=28,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    elements = []

    rows_per_page = 5
    for start in range(0, len(ARCHIVED_ROWS), rows_per_page):
        chunk = ARCHIVED_ROWS[start : start + rows_per_page]
        page_index = start // rows_per_page + 1
        elements.append(Paragraph(f"Archived Badge Assignment Packet - Page {page_index}", styles["Heading2"]))
        elements.append(Spacer(1, 10))

        table = Table(
            table_data(chunk),
            colWidths=[88, 120, 78, 120, 92],
            repeatRows=1,
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DCEAF7")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F8FB")]),
                ]
            )
        )
        elements.append(table)

        if start + rows_per_page < len(ARCHIVED_ROWS):
            elements.append(PageBreak())

    doc.build(elements)


def main():
    write_excel()
    write_pdf()


if __name__ == "__main__":
    main()
