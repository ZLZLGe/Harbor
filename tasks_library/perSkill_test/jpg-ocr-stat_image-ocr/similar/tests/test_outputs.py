import os
from typing import Any

from openpyxl import load_workbook


def _to_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _read_rows(path: str, sheet_name: str) -> list[list[str]]:
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[sheet_name]
        rows: list[list[str]] = []
        for r in range(1, (ws.max_row or 0) + 1):
            rows.append([_to_str(ws.cell(row=r, column=c).value) for c in range(1, (ws.max_column or 0) + 1)])
        return rows
    finally:
        wb.close()


def _sheetnames(path: str) -> list[str]:
    wb = load_workbook(path, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def test_outputs() -> None:
    actual_path = "/app/workspace/similar_receipt_ledger.xlsx"
    expected_path = os.path.join(os.path.dirname(__file__), "similar_oracle.xlsx")

    assert os.path.exists(actual_path), "Output file missing: /app/workspace/similar_receipt_ledger.xlsx"

    actual_sheets = _sheetnames(actual_path)
    expected_sheets = _sheetnames(expected_path)
    assert actual_sheets == expected_sheets, f"Sheet names mismatch: actual={actual_sheets}, expected={expected_sheets}"
    assert actual_sheets == ["ledger"], f"Workbook must contain exactly one sheet named 'ledger', got {actual_sheets}"

    actual_rows = _read_rows(actual_path, "ledger")
    expected_rows = _read_rows(expected_path, "ledger")

    assert actual_rows == expected_rows, (
        "Workbook content mismatch against oracle.\n"
        f"Actual rows ({len(actual_rows)}): {actual_rows}\n"
        f"Expected rows ({len(expected_rows)}): {expected_rows}"
    )
