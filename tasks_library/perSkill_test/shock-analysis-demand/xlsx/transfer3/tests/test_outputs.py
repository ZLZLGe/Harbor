import os
from pathlib import Path

import openpyxl

OUTPUT_FILE = Path(os.environ.get("OUTPUT_FILE", "/root/summit_staffing_transfer_completed.xlsx"))
REQUIRED_SHEETS = ["Sessions", "Rates", "Staffing_Model", "Scenario_Compare", "Summary"]
FORMULA_CHECKS = [
    ("Staffing_Model", "A2", "=Sessions!A2"),
    ("Staffing_Model", "B2", "=Sessions!B2"),
    ("Staffing_Model", "C2", "=ROUNDUP(B2/120*Rates!$B$2,0)"),
    ("Staffing_Model", "D3", "=ROUNDUP(B3/180*Rates!$B$3,0)"),
    ("Staffing_Model", "E4", "=ROUNDUP(B4/250*Rates!$B$4,0)"),
    ("Staffing_Model", "F5", "=ROUNDUP((C5+D5+E5)*Rates!$B$5,0)"),
    ("Staffing_Model", "G2", "=(C2+D2+E2+F2)*Sessions!C2"),
    ("Staffing_Model", "H2", "=Sessions!C2*(C2*Rates!$B$6+D2*Rates!$B$7+E2*Rates!$B$8+F2*Rates!$B$9)"),
    ("Staffing_Model", "I6", "=B6*Sessions!D6"),
    ("Scenario_Compare", "B2", "=SUM(Staffing_Model!B2:B6)*1"),
    ("Scenario_Compare", "C3", "=SUM(Staffing_Model!G2:G6)*Rates!$B$10"),
    ("Scenario_Compare", "D4", "=SUM(Staffing_Model!H2:H6)*Rates!$B$11"),
    ("Scenario_Compare", "E2", "=Rates!$B$12 + B2*Rates!$B$13"),
    ("Scenario_Compare", "F4", "=E4-D4"),
    ("Summary", "B2", "=INDEX(Staffing_Model!A2:A6,MATCH(MAX(Staffing_Model!H2:H6),Staffing_Model!H2:H6,0))"),
    ("Summary", "B3", "=Scenario_Compare!F2"),
    ("Summary", "B4", "=Scenario_Compare!F4"),
    ("Summary", "B5", "=SUM(Staffing_Model!I2:I6)"),
]
VALUE_CHECKS = [
    ("Sessions", "B2", 420),
    ("Sessions", "D5", 0.22),
    ("Rates", "B9", 35),
    ("Rates", "B13", 145),
]


def load_workbook_checked():
    assert OUTPUT_FILE.exists(), f"missing output workbook: {OUTPUT_FILE}"
    return openpyxl.load_workbook(OUTPUT_FILE, data_only=False)


def assert_required_sheets(wb):
    assert wb.sheetnames == REQUIRED_SHEETS, f"unexpected sheets: {wb.sheetnames}"


def assert_formula_checks(wb):
    for sheet_name, cell, expected in FORMULA_CHECKS:
        actual = wb[sheet_name][cell].value
        assert actual == expected, f"{sheet_name}!{cell} expected {expected!r}, got {actual!r}"


def assert_value_checks(wb):
    for sheet_name, cell, expected in VALUE_CHECKS:
        actual = wb[sheet_name][cell].value
        assert actual == expected, f"{sheet_name}!{cell} expected {expected!r}, got {actual!r}"


def assert_no_todo_strings(wb):
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str):
                    assert "TODO" not in value, f"TODO placeholder remains in {sheet.title}"


def main():
    wb = load_workbook_checked()
    assert_required_sheets(wb)
    assert_formula_checks(wb)
    assert_value_checks(wb)
    assert_no_todo_strings(wb)


if __name__ == "__main__":
    main()
