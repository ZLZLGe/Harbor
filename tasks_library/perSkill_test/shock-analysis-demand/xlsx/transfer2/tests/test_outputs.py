import os
from pathlib import Path

import openpyxl

OUTPUT_FILE = Path(os.environ.get("OUTPUT_FILE", "/root/fleet_electrification_transfer_completed.xlsx"))
REQUIRED_SHEETS = ["Fleet", "Tariffs", "Deployment", "Summary"]
FORMULA_CHECKS = [
    ("Deployment", "A2", "=Fleet!A2"),
    ("Deployment", "B2", "=Fleet!C2*365*Fleet!B2"),
    ("Deployment", "C2", "=B2*Fleet!D2/100*Tariffs!$B$2"),
    ("Deployment", "D2", "=B2*Fleet!E2*Tariffs!$B$3"),
    ("Deployment", "E4", "=C4-D4"),
    ("Deployment", "F5", "=ROUNDUP(Fleet!B5/4,0)"),
    ("Deployment", "G3", "=F3*Tariffs!$B$4"),
    ("Deployment", "H2", "=Fleet!B2*Tariffs!$B$5"),
    ("Deployment", "I2", "=((B2*Fleet!D2/100*Tariffs!$B$6)-(B2*Fleet!E2*Tariffs!$B$7))/1000"),
    ("Summary", "B2", "=SUM(Deployment!E2:E5)"),
    ("Summary", "B3", "=SUM(Deployment!F2:F5)"),
    ("Summary", "B4", "=SUM(Deployment!G2:G5)+SUM(Deployment!H2:H5)"),
    ("Summary", "B5", "=B4/B2"),
    ("Summary", "B6", "=INDEX(Deployment!A2:A5,MATCH(MAX(Deployment!E2:E5),Deployment!E2:E5,0))"),
    ("Summary", "B7", "=SUM(Deployment!I2:I5)"),
]
VALUE_CHECKS = [
    ("Fleet", "B2", 12),
    ("Fleet", "E5", 0.96),
    ("Tariffs", "B4", 18500),
    ("Tariffs", "B7", 0.43),
]


def load_workbook_checked():
    assert OUTPUT_FILE.exists(), f"missing output workbook: {OUTPUT_FILE}"
    return openpyxl.load_workbook(OUTPUT_FILE, data_only=False)


def assert_required_sheets(wb):
    assert wb.sheetnames == REQUIRED_SHEETS, f"unexpected sheets: {wb.sheetnames}"


def assert_formula_checks(wb):
    for sheet_name, cell, expected in FORMULA_CHECKS:
        actual = wb[sheet_name][cell].value
        assert actual == expected, f"{sheet_name}!{cell} expected {expected!r}, got {actual!r}"


def assert_value_checks(wb):
    for sheet_name, cell, expected in VALUE_CHECKS:
        actual = wb[sheet_name][cell].value
        assert actual == expected, f"{sheet_name}!{cell} expected {expected!r}, got {actual!r}"


def assert_no_todo_strings(wb):
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str):
                    assert "TODO" not in value, f"TODO placeholder remains in {sheet.title}"


def main():
    wb = load_workbook_checked()
    assert_required_sheets(wb)
    assert_formula_checks(wb)
    assert_value_checks(wb)
    assert_no_todo_strings(wb)


if __name__ == "__main__":
    main()
