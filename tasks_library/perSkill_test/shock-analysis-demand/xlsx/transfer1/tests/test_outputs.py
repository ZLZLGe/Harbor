import os
from pathlib import Path

import openpyxl

OUTPUT_FILE = Path(os.environ.get("OUTPUT_FILE", "/root/cold_chain_transfer_completed.xlsx"))
REQUIRED_SHEETS = ["Demand", "Sites", "Config", "Capacity_Model", "Summary"]
FORMULA_CHECKS = [
    ("Capacity_Model", "B2", "=AVERAGE(Demand!B2:B7)"),
    ("Capacity_Model", "C2", "=B2/Config!$B$2*(1+Config!$B$3)"),
    ("Capacity_Model", "D2", "=Sites!B2"),
    ("Capacity_Model", "E2", "=ROUNDUP(C2/Config!$B$4,0)"),
    ("Capacity_Model", "F3", "=MAX(E3-D3,0)"),
    ("Capacity_Model", "H4", "=G4*(1+Config!$B$6)"),
    ("Capacity_Model", "J4", "=I4*Config!$B$7"),
    ("Summary", "B2", "=SUM(Capacity_Model!G2:G4)"),
    ("Summary", "B3", "=SUM(Capacity_Model!H2:H4)"),
    ("Summary", "B4", "=SUM(Capacity_Model!J2:J4)"),
    ("Summary", "B5", "=INDEX(Capacity_Model!A2:A4,MATCH(MAX(Capacity_Model!C2:C4),Capacity_Model!C2:C4,0))"),
    ("Summary", "B6", "=SUM(Capacity_Model!F2:F4)"),
]
VALUE_CHECKS = [
    ("Demand", "B2", 680),
    ("Demand", "D7", 540),
    ("Sites", "B4", 8),
    ("Config", "B5", 1850),
    ("Config", "B7", 7.5),
]


def load_workbook_checked():
    assert OUTPUT_FILE.exists(), f"missing output workbook: {OUTPUT_FILE}"
    return openpyxl.load_workbook(OUTPUT_FILE, data_only=False)


def assert_required_sheets(wb):
    assert wb.sheetnames == REQUIRED_SHEETS, f"unexpected sheets: {wb.sheetnames}"


def assert_formula_checks(wb):
    for sheet_name, cell, expected in FORMULA_CHECKS:
        actual = wb[sheet_name][cell].value
        assert actual == expected, f"{sheet_name}!{cell} expected {expected!r}, got {actual!r}"


def assert_value_checks(wb):
    for sheet_name, cell, expected in VALUE_CHECKS:
        actual = wb[sheet_name][cell].value
        assert actual == expected, f"{sheet_name}!{cell} expected {expected!r}, got {actual!r}"


def assert_no_todo_strings(wb):
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str):
                    assert "TODO" not in value, f"TODO placeholder remains in {sheet.title}"


def main():
    wb = load_workbook_checked()
    assert_required_sheets(wb)
    assert_formula_checks(wb)
    assert_value_checks(wb)
    assert_no_todo_strings(wb)


if __name__ == "__main__":
    main()
