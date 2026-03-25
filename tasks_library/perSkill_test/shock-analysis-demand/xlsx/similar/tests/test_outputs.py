import os
from pathlib import Path

import openpyxl

OUTPUT_FILE = Path(os.environ.get("OUTPUT_FILE", "/root/macro_demand_similar_completed.xlsx"))
REQUIRED_SHEETS = [
    "Inputs",
    "Baseline",
    "Scenario_Base",
    "Scenario_HigherMultiplier",
    "Scenario_HigherImport",
    "Summary",
]
FORMULA_CHECKS = [
    ("Scenario_Base", "A2", "=Inputs!A10"),
    ("Scenario_Base", "B2", "=Inputs!B10"),
    ("Scenario_Base", "C2", "=Inputs!$B$2*Inputs!$B$3*B2"),
    ("Scenario_Base", "D2", "=C2*(1-Inputs!$B$4)*Inputs!$B$5"),
    ("Scenario_Base", "E2", "=D2/Baseline!C2"),
    ("Scenario_Base", "F2", "=E2/(Baseline!B2*1000)"),
    ("Scenario_HigherMultiplier", "D5", "=C5*(1-Inputs!$B$4)*Inputs!$B$6"),
    ("Scenario_HigherImport", "D8", "=C8*(1-Inputs!$B$7)*Inputs!$B$5"),
    ("Summary", "B2", "=SUM(Scenario_Base!C2:C9)"),
    ("Summary", "C3", "=SUM(Scenario_HigherMultiplier!D2:D9)"),
    ("Summary", "D4", "=MAX(Scenario_HigherImport!E2:E9)"),
    ("Summary", "B6", "=INDEX(Scenario_Base!A2:A9,MATCH(MAX(Scenario_Base!F2:F9),Scenario_Base!F2:F9,0))"),
]
VALUE_CHECKS = [
    ("Inputs", "B2", 7200),
    ("Inputs", "B3", 2.81),
    ("Inputs", "B7", 0.48),
    ("Baseline", "B9", 102.1),
    ("Baseline", "C6", 1.16),
]


def load_workbook_checked():
    assert OUTPUT_FILE.exists(), f"missing output workbook: {OUTPUT_FILE}"
    return openpyxl.load_workbook(OUTPUT_FILE, data_only=False)


def assert_required_sheets(wb):
    assert wb.sheetnames == REQUIRED_SHEETS, f"unexpected sheets: {wb.sheetnames}"


def assert_formula_checks(wb):
    for sheet_name, cell, expected in FORMULA_CHECKS:
        actual = wb[sheet_name][cell].value
        assert actual == expected, f"{sheet_name}!{cell} expected {expected!r}, got {actual!r}"


def assert_value_checks(wb):
    for sheet_name, cell, expected in VALUE_CHECKS:
        actual = wb[sheet_name][cell].value
        assert actual == expected, f"{sheet_name}!{cell} expected {expected!r}, got {actual!r}"


def assert_no_todo_strings(wb):
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str):
                    assert "TODO" not in value, f"TODO placeholder remains in {sheet.title}"


def main():
    wb = load_workbook_checked()
    assert_required_sheets(wb)
    assert_formula_checks(wb)
    assert_value_checks(wb)
    assert_no_todo_strings(wb)


if __name__ == "__main__":
    main()
