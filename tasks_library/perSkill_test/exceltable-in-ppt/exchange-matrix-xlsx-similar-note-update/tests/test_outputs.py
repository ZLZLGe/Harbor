import math
import os
import zipfile

from openpyxl import load_workbook

INPUT_FILE = "/root/input_exchange_matrix.xlsx"
RESULT_FILE = "/root/results_exchange_matrix.xlsx"
MATRIX_SHEET = "汇率矩阵"
NOTE_SHEET = "更新说明"
EXTRA_SHEET = "封面"


def load_pair_details(workbook_path):
    wb = load_workbook(workbook_path, data_only=False)
    note = wb[NOTE_SHEET]
    base_currency = str(note["B4"].value).strip()
    quote_currency = str(note["B5"].value).strip()
    new_rate = float(note["B6"].value)

    matrix = wb[MATRIX_SHEET]
    row_map = {}
    for row_idx in range(2, matrix.max_row + 1):
        value = matrix.cell(row=row_idx, column=1).value
        if value is None:
            continue
        row_map[str(value).strip()] = row_idx

    col_map = {}
    for col_idx in range(2, matrix.max_column + 1):
        value = matrix.cell(row=1, column=col_idx).value
        if value is None:
            continue
        col_map[str(value).strip()] = col_idx

    direct_coord = matrix.cell(row=row_map[base_currency], column=col_map[quote_currency]).coordinate
    inverse_coord = matrix.cell(row=row_map[quote_currency], column=col_map[base_currency]).coordinate
    wb.close()
    return base_currency, quote_currency, new_rate, direct_coord, inverse_coord


def is_formula(value):
    return isinstance(value, str) and value.startswith("=")


def test_output_exists():
    assert os.path.exists(RESULT_FILE), f"Missing output file: {RESULT_FILE}"


def test_output_is_valid_xlsx():
    assert zipfile.is_zipfile(RESULT_FILE), f"{RESULT_FILE} is not a valid xlsx file"


def test_sheet_names_preserved():
    input_wb = load_workbook(INPUT_FILE, data_only=False)
    output_wb = load_workbook(RESULT_FILE, data_only=False)
    assert input_wb.sheetnames == output_wb.sheetnames == [MATRIX_SHEET, NOTE_SHEET, EXTRA_SHEET]
    input_wb.close()
    output_wb.close()


def test_note_and_cover_sheets_unchanged():
    input_wb = load_workbook(INPUT_FILE, data_only=False)
    output_wb = load_workbook(RESULT_FILE, data_only=False)

    for sheet_name in [NOTE_SHEET, EXTRA_SHEET]:
        input_ws = input_wb[sheet_name]
        output_ws = output_wb[sheet_name]
        for row in range(1, max(input_ws.max_row, output_ws.max_row) + 1):
            for col in range(1, max(input_ws.max_column, output_ws.max_column) + 1):
                assert input_ws.cell(row=row, column=col).value == output_ws.cell(row=row, column=col).value

    input_wb.close()
    output_wb.close()


def test_only_editable_input_cell_changed_and_formula_preserved():
    _, _, new_rate, direct_coord, inverse_coord = load_pair_details(INPUT_FILE)
    input_wb = load_workbook(INPUT_FILE, data_only=False)
    output_wb = load_workbook(RESULT_FILE, data_only=False)
    input_ws = input_wb[MATRIX_SHEET]
    output_ws = output_wb[MATRIX_SHEET]

    direct_input = input_ws[direct_coord].value
    inverse_input = input_ws[inverse_coord].value
    direct_output = output_ws[direct_coord].value
    inverse_output = output_ws[inverse_coord].value

    assert is_formula(direct_input) != is_formula(inverse_input), "Input matrix should have one formula side and one editable side"

    if is_formula(direct_input):
        assert direct_output == direct_input, "Direct-rate formula was modified"
        assert not is_formula(inverse_output), "Editable inverse cell was turned into a formula"
        assert math.isclose(float(inverse_output), 1 / new_rate, rel_tol=0, abs_tol=1e-9)
        changed_coord = inverse_coord
        untouched_formula_coord = direct_coord
    else:
        assert math.isclose(float(direct_output), new_rate, rel_tol=0, abs_tol=1e-9), (
            "Editable direct cell was not updated to the note-sheet rate"
        )
        assert inverse_output == inverse_input, "Reverse-rate formula was modified"
        changed_coord = direct_coord
        untouched_formula_coord = inverse_coord

    for row in range(2, input_ws.max_row + 1):
        for col in range(2, input_ws.max_column + 1):
            coord = input_ws.cell(row=row, column=col).coordinate
            if coord == changed_coord:
                continue
            assert output_ws[coord].value == input_ws[coord].value, f"Unexpected change at {coord}"

    assert is_formula(output_ws[untouched_formula_coord].value), "Formula cell should remain a formula"

    input_wb.close()
    output_wb.close()


def test_displayed_rate_matches_note_after_recalc():
    _, _, new_rate, direct_coord, inverse_coord = load_pair_details(INPUT_FILE)
    formula_wb = load_workbook(INPUT_FILE, data_only=False)
    formula_ws = formula_wb[MATRIX_SHEET]
    direct_is_formula = is_formula(formula_ws[direct_coord].value)
    formula_wb.close()

    value_wb = load_workbook(RESULT_FILE, data_only=True)
    value_ws = value_wb[MATRIX_SHEET]

    if direct_is_formula:
        assert math.isclose(float(value_ws[direct_coord].value), new_rate, rel_tol=0, abs_tol=1e-6)
        assert math.isclose(float(value_ws[inverse_coord].value), 1 / new_rate, rel_tol=0, abs_tol=1e-9)
    else:
        assert math.isclose(float(value_ws[direct_coord].value), new_rate, rel_tol=0, abs_tol=1e-9)

    value_wb.close()
