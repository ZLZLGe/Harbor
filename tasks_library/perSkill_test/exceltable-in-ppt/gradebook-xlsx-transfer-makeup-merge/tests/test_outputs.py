import csv
from pathlib import Path

from openpyxl import load_workbook

TEMPLATE_FILE = Path("/root/gradebook_template.xlsx")
OUTPUT_FILE = Path("/root/makeup_gradebook.xlsx")
CSV_FILE = Path("/root/makeup_scores.csv")

EXPECTED_TOTALS = {
    "20231101": 85.9,
    "20231102": 82.1,
    "20231103": 86.7,
    "20231104": 81.0,
    "20231105": 81.0,
}

EXPECTED_GRADES = {
    "20231101": "B",
    "20231102": "B",
    "20231103": "B",
    "20231104": "B",
    "20231105": "B",
}

ALLOWED_EDIT_CELLS = {"E2", "G2", "E4", "G5"}
PROTECTED_HEADERS = ["学号", "姓名", "出勤", "小测1", "小测3", "作业", "评语"]
FORMULA_CELLS = {
    ("成绩册", "I2"),
    ("成绩册", "I3"),
    ("成绩册", "I4"),
    ("成绩册", "I5"),
    ("成绩册", "I6"),
    ("成绩册", "J2"),
    ("成绩册", "J3"),
    ("成绩册", "J4"),
    ("成绩册", "J5"),
    ("成绩册", "J6"),
    ("课程概览", "B2"),
    ("课程概览", "B3"),
    ("课程概览", "B4"),
    ("课程概览", "B5"),
}


def header_map(ws):
    return {
        str(ws.cell(row=1, column=column).value).strip(): column
        for column in range(1, ws.max_column + 1)
    }


def student_row_map(ws):
    columns = header_map(ws)
    student_col = columns["学号"]
    return {
        str(ws.cell(row=row, column=student_col).value).strip(): row
        for row in range(2, ws.max_row + 1)
    }


def load_updates():
    updates = {}
    with CSV_FILE.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            updates[row["学号"].strip()] = {
                "小测2": (row.get("小测2补考") or "").strip(),
                "小测4": (row.get("小测4补考") or "").strip(),
            }
    return updates


def test_output_exists():
    assert OUTPUT_FILE.exists(), f"缺少输出文件: {OUTPUT_FILE}"


def test_makeup_scores_are_merged_by_student_id():
    workbook = load_workbook(OUTPUT_FILE, data_only=False)
    try:
        ws = workbook["成绩册"]
        columns = header_map(ws)
        rows = student_row_map(ws)
        updates = load_updates()

        assert ws.cell(rows["20231101"], columns["小测2"]).value == 78
        assert ws.cell(rows["20231101"], columns["小测4"]).value == 85
        assert ws.cell(rows["20231103"], columns["小测2"]).value == 93
        assert ws.cell(rows["20231103"], columns["小测4"]).value == 56
        assert ws.cell(rows["20231104"], columns["小测2"]).value == 76
        assert ws.cell(rows["20231104"], columns["小测4"]).value == 88
        assert ws.cell(rows["20231105"], columns["小测2"]).value == 70
        assert ws.cell(rows["20231105"], columns["小测4"]).value == 65

        assert "20239999" not in rows
        assert updates["20239999"]["小测2"] == "77"
        assert updates["20239999"]["小测4"] == "92"
    finally:
        workbook.close()


def test_protected_columns_remain_unchanged():
    template = load_workbook(TEMPLATE_FILE, data_only=False)
    output = load_workbook(OUTPUT_FILE, data_only=False)
    try:
        template_ws = template["成绩册"]
        output_ws = output["成绩册"]
        template_columns = header_map(template_ws)
        output_columns = header_map(output_ws)

        for student_id, template_row in student_row_map(template_ws).items():
            output_row = student_row_map(output_ws)[student_id]
            for header in PROTECTED_HEADERS:
                template_value = template_ws.cell(template_row, template_columns[header]).value
                output_value = output_ws.cell(output_row, output_columns[header]).value
                assert output_value == template_value, f"{student_id} 的 {header} 被意外改动"
    finally:
        template.close()
        output.close()


def test_formula_cells_are_preserved():
    template = load_workbook(TEMPLATE_FILE, data_only=False)
    output = load_workbook(OUTPUT_FILE, data_only=False)
    try:
        for sheet_name, cell_ref in FORMULA_CELLS:
            assert output[sheet_name][cell_ref].value == template[sheet_name][cell_ref].value
    finally:
        template.close()
        output.close()


def test_only_allowed_input_cells_change_in_gradebook():
    template = load_workbook(TEMPLATE_FILE, data_only=False)
    output = load_workbook(OUTPUT_FILE, data_only=False)
    try:
        template_ws = template["成绩册"]
        output_ws = output["成绩册"]
        for row in range(1, template_ws.max_row + 1):
            for column in range(1, template_ws.max_column + 1):
                coordinate = template_ws.cell(row=row, column=column).coordinate
                if row == 1:
                    assert output_ws[coordinate].value == template_ws[coordinate].value
                    continue
                if coordinate in ALLOWED_EDIT_CELLS:
                    continue
                if ("成绩册", coordinate) in FORMULA_CELLS:
                    continue
                assert output_ws[coordinate].value == template_ws[coordinate].value, f"{coordinate} 被意外改动"
    finally:
        template.close()
        output.close()


def test_totals_grades_and_summary_are_recalculated():
    workbook = load_workbook(OUTPUT_FILE, data_only=True)
    try:
        ws = workbook["成绩册"]
        summary = workbook["课程概览"]
        columns = header_map(ws)
        rows = student_row_map(ws)

        for student_id, expected_total in EXPECTED_TOTALS.items():
            actual_total = ws.cell(rows[student_id], columns["总评"]).value
            assert abs(actual_total - expected_total) < 1e-9, f"{student_id} 总评错误: {actual_total}"
            actual_grade = ws.cell(rows[student_id], columns["等级"]).value
            assert actual_grade == EXPECTED_GRADES[student_id], f"{student_id} 等级错误: {actual_grade}"

        assert summary["B2"].value == 5
        assert abs(summary["B3"].value - 83.3) < 1e-9
        assert summary["B4"].value == 5
        assert summary["B5"].value == 0
    finally:
        workbook.close()
