from pathlib import Path

from openpyxl import load_workbook

INPUT_FILE = Path("/root/fx_matrix.xlsx")
OUTPUT_FILE = Path("/root/fx_matrix_updated.xlsx")

TARGET_FORMULA_CELL = "E2"
TARGET_INPUT_CELL = "B5"
EXPECTED_RATE = 0.0064
EXPECTED_INPUT = 156.25


def load_books(data_only):
    return load_workbook(INPUT_FILE, data_only=data_only), load_workbook(OUTPUT_FILE, data_only=data_only)


def iter_cells(ws):
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            yield ws.cell(row=row, column=col)


def test_output_exists():
    assert OUTPUT_FILE.exists(), f"缺少输出文件: {OUTPUT_FILE}"


def test_target_pair_is_updated_via_editable_input():
    input_wb, output_wb = load_books(data_only=False)
    try:
        in_matrix = input_wb["汇率矩阵"]
        out_matrix = output_wb["汇率矩阵"]

        assert out_matrix[TARGET_INPUT_CELL].value == EXPECTED_INPUT
        assert in_matrix[TARGET_FORMULA_CELL].value == "=1/B5"
        assert out_matrix[TARGET_FORMULA_CELL].value == in_matrix[TARGET_FORMULA_CELL].value
    finally:
        input_wb.close()
        output_wb.close()


def test_formula_result_is_recalculated():
    data_wb = load_workbook(OUTPUT_FILE, data_only=True)
    try:
        matrix = data_wb["汇率矩阵"]
        assert abs(matrix[TARGET_FORMULA_CELL].value - EXPECTED_RATE) < 1e-9
        assert abs(matrix[TARGET_INPUT_CELL].value - EXPECTED_INPUT) < 1e-9
    finally:
        data_wb.close()


def test_other_matrix_cells_remain_unchanged():
    input_wb, output_wb = load_books(data_only=False)
    try:
        in_matrix = input_wb["汇率矩阵"]
        out_matrix = output_wb["汇率矩阵"]

        for cell in iter_cells(in_matrix):
            if cell.coordinate == TARGET_INPUT_CELL:
                continue
            assert out_matrix[cell.coordinate].value == cell.value, f"{cell.coordinate} 被意外改动"
    finally:
        input_wb.close()
        output_wb.close()


def test_note_sheet_is_unchanged():
    input_wb, output_wb = load_books(data_only=False)
    try:
        in_notes = input_wb["更新备注"]
        out_notes = output_wb["更新备注"]

        for cell in iter_cells(in_notes):
            assert out_notes[cell.coordinate].value == cell.value, f"备注表 {cell.coordinate} 被改动"
    finally:
        input_wb.close()
        output_wb.close()
