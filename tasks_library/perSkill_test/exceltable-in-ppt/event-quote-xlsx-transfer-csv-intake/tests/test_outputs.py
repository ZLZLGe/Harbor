import csv
from pathlib import Path

from openpyxl import load_workbook

TEMPLATE_FILE = Path("/root/event_quote_template.xlsx")
OUTPUT_FILE = Path("/root/event_quote_book.xlsx")
CSV_FILE = Path("/root/new_event_request.csv")


def load_request():
    with CSV_FILE.open("r", encoding="utf-8-sig", newline="") as handle:
        row = next(csv.DictReader(handle))
    row["来宾人数"] = int(row["来宾人数"])
    return row


def load_formula_map(path, sheet_names):
    workbook = load_workbook(path, data_only=False)
    try:
        formulas = {}
        for sheet_name in sheet_names:
            ws = workbook[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas[(sheet_name, cell.coordinate)] = cell.value
        return formulas
    finally:
        workbook.close()


def find_label_row(ws, label):
    for idx in range(1, ws.max_row + 1):
        value = ws.cell(row=idx, column=1).value
        if str(value).strip() == label:
            return idx
    raise AssertionError(f"未找到字段: {label}")


def test_output_exists():
    assert OUTPUT_FILE.exists(), f"缺少输出文件: {OUTPUT_FILE}"


def test_request_values_are_written_into_input_sheet():
    expected = load_request()
    workbook = load_workbook(OUTPUT_FILE, data_only=False)
    try:
        ws = workbook["需求录入"]
        for field, expected_value in expected.items():
            row = find_label_row(ws, field)
            actual_value = ws.cell(row=row, column=2).value
            assert actual_value == expected_value, f"{field} 写入不正确: {actual_value!r}"
    finally:
        workbook.close()


def test_formula_cells_are_preserved():
    template_formulas = load_formula_map(TEMPLATE_FILE, ["报价明细", "汇总"])
    output_formulas = load_formula_map(OUTPUT_FILE, ["报价明细", "汇总"])
    assert output_formulas == template_formulas


def test_amounts_and_summary_are_recalculated():
    workbook = load_workbook(OUTPUT_FILE, data_only=True)
    try:
        detail = workbook["报价明细"]
        summary = workbook["汇总"]

        assert detail["E2"].value == 41040
        assert detail["E3"].value == 4680
        assert detail["E4"].value == 0
        assert detail["E5"].value == 1200
        assert detail["E6"].value == 46920
        assert detail["E7"].value == 4692
        assert detail["E8"].value == 51612

        assert summary["B2"].value == "云岫团队夏季答谢酒会"
        assert summary["B3"].value == 180
        assert summary["B4"].value == "臻选"
        assert summary["B5"].value == 2
        assert summary["B6"].value == 46920
        assert summary["B7"].value == 4692
        assert summary["B8"].value == 51612
    finally:
        workbook.close()


def test_non_input_labels_remain_unchanged():
    template = load_workbook(TEMPLATE_FILE, data_only=False)
    output = load_workbook(OUTPUT_FILE, data_only=False)
    try:
        template_ws = template["需求录入"]
        output_ws = output["需求录入"]
        for row in range(1, template_ws.max_row + 1):
            assert output_ws.cell(row=row, column=1).value == template_ws.cell(row=row, column=1).value
    finally:
        template.close()
        output.close()
