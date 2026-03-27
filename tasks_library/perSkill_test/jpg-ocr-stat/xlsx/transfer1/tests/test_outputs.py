import os
from openpyxl import load_workbook

OUTPUT_PATH = "/app/workspace/transfer1.xlsx"
EXPECTED_ROWS = [
    ["sku", "on_hand", "reorder_point", "order_qty", "reorder_cost"],
    ["C300", "0", "10", "10", "99.90"],
    ["A100", "12", "20", "8", "28.00"],
    ["E500", "5", "12", "7", "15.75"],
    ["B200", "40", "25", "0", "0.00"],
    ["D400", "7", "7", "0", "0.00"],
]


def norm(value):
    if value is None:
        return ""
    return str(value)


assert os.path.exists(OUTPUT_PATH), f"missing output: {OUTPUT_PATH}"
wb = load_workbook(OUTPUT_PATH, data_only=True)
try:
    assert wb.sheetnames == ["plan"], f"unexpected sheets: {wb.sheetnames}"
    ws = wb["plan"]
    rows = []
    for r in range(1, ws.max_row + 1):
        rows.append([norm(ws.cell(row=r, column=c).value) for c in range(1, 6)])
    assert rows == EXPECTED_ROWS, f"rows mismatch\nactual={rows}\nexpected={EXPECTED_ROWS}"
finally:
    wb.close()
