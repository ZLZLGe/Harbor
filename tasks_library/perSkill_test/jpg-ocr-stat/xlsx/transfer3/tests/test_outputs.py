import os
from openpyxl import load_workbook

OUTPUT_PATH = "/app/workspace/transfer3.xlsx"
EXPECTED_ROWS = [
    ["student", "final_score", "grade"],
    ["Chen", "95.0", "A"],
    ["Alice", "90.2", "A"],
    ["Evan", "80.0", "B"],
    ["Bob", "70.0", "C"],
    ["Dina", "60.0", "D"],
    ["Farah", "55.5", "F"],
]


def norm(value):
    if value is None:
        return ""
    return str(value)


assert os.path.exists(OUTPUT_PATH), f"missing output: {OUTPUT_PATH}"
wb = load_workbook(OUTPUT_PATH, data_only=True)
try:
    assert wb.sheetnames == ["report"], f"unexpected sheets: {wb.sheetnames}"
    ws = wb["report"]
    rows = []
    for r in range(1, ws.max_row + 1):
        rows.append([norm(ws.cell(row=r, column=c).value) for c in range(1, 4)])
    assert rows == EXPECTED_ROWS, f"rows mismatch\nactual={rows}\nexpected={EXPECTED_ROWS}"
finally:
    wb.close()
