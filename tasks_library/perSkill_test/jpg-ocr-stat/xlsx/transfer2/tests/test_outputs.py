import os
from openpyxl import load_workbook

OUTPUT_PATH = "/app/workspace/transfer2.xlsx"
EXPECTED_ROWS = [
    ["date", "total_patients", "weighted_wait_min"],
    ["2024-04-01", "30", "16.0"],
    ["2024-04-02", "20", "15.0"],
    ["2024-04-03", "0", ""],
]


def norm(value):
    if value is None:
        return ""
    return str(value)


assert os.path.exists(OUTPUT_PATH), f"missing output: {OUTPUT_PATH}"
wb = load_workbook(OUTPUT_PATH, data_only=True)
try:
    assert wb.sheetnames == ["daily_summary"], f"unexpected sheets: {wb.sheetnames}"
    ws = wb["daily_summary"]
    rows = []
    for r in range(1, ws.max_row + 1):
        rows.append([norm(ws.cell(row=r, column=c).value) for c in range(1, 4)])
    assert rows == EXPECTED_ROWS, f"rows mismatch\nactual={rows}\nexpected={EXPECTED_ROWS}"
finally:
    wb.close()
