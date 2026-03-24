import csv
import json
import os
from datetime import datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


WORKSPACE = Path(os.environ.get("WORKSPACE_ROOT", "/app/workspace"))
OUTPUT_FILE = WORKSPACE / "volunteer_shift_plan.xlsx"
VOLUNTEERS_FILE = WORKSPACE / "data" / "volunteers.tsv"
SHIFTS_FILE = WORKSPACE / "data" / "shift_needs.csv"
AVAILABILITY_FILE = WORKSPACE / "data" / "availability.json"


def read_volunteers(path: Path) -> list[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        rows = []
        for index, row in enumerate(reader):
            row["max_shifts"] = int(row["max_shifts"])
            row["eligible_roles"] = row["eligible_roles"].split("|")
            row["order"] = index
            rows.append(row)
        return rows


def read_shifts(path: Path) -> list[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        rows = []
        for index, row in enumerate(reader):
            row["required_count"] = int(row["required_count"])
            row["shift_date_obj"] = datetime.strptime(row["shift_date"], "%Y-%m-%d").date()
            row["start_time_obj"] = datetime.strptime(row["start_time"], "%H:%M").time()
            row["end_time_obj"] = datetime.strptime(row["end_time"], "%H:%M").time()
            row["order"] = index
            rows.append(row)
        return rows


def read_availability(path: Path) -> dict[str, set[str]]:
    with path.open(encoding="utf-8") as handle:
        payload = json.load(handle)
    return {
        entry["volunteer_id"]: set(entry["available_shifts"])
        for entry in payload["availability"]
    }


def overlaps(a: dict[str, Any], b: dict[str, Any]) -> bool:
    if a["shift_date"] != b["shift_date"]:
        return False
    return a["start_time"] < b["end_time"] and a["end_time"] > b["start_time"]


def normalize_date(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "date"):
        return value.date().isoformat()
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def normalize_time(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if hasattr(value, "time"):
        return value.time().strftime("%H:%M")
    return str(value)


def normalize_int(value: Any) -> int:
    assert value not in (None, ""), f"expected numeric value, got {value!r}"
    return int(round(float(value)))


def build_expected() -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, int], dict[str, int]]:
    volunteers = read_volunteers(VOLUNTEERS_FILE)
    shifts = read_shifts(SHIFTS_FILE)
    availability = read_availability(AVAILABILITY_FILE)

    assigned_shift_details: dict[str, list[dict[str, Any]]] = {row["volunteer_id"]: [] for row in volunteers}
    assignment_counts: dict[str, int] = {row["volunteer_id"]: 0 for row in volunteers}
    assignments: list[dict[str, Any]] = []
    assigned_counts_by_shift: dict[str, int] = {row["shift_id"]: 0 for row in shifts}

    for shift in shifts:
        normalized_shift = {
            "shift_id": shift["shift_id"],
            "shift_date": shift["shift_date_obj"],
            "start_time": shift["start_time_obj"],
            "end_time": shift["end_time_obj"],
            "site": shift["site"],
            "role": shift["role"],
            "required_count": shift["required_count"],
        }

        candidates = []
        for volunteer in volunteers:
            volunteer_id = volunteer["volunteer_id"]
            if shift["shift_id"] not in availability.get(volunteer_id, set()):
                continue
            if shift["role"] not in volunteer["eligible_roles"]:
                continue
            if assignment_counts[volunteer_id] >= volunteer["max_shifts"]:
                continue
            if any(overlaps(normalized_shift, prior_shift) for prior_shift in assigned_shift_details[volunteer_id]):
                continue
            candidates.append(volunteer)

        candidates.sort(key=lambda volunteer: (assignment_counts[volunteer["volunteer_id"]], volunteer["order"]))
        selected = candidates[: shift["required_count"]]

        for volunteer in selected:
            volunteer_id = volunteer["volunteer_id"]
            assignment_counts[volunteer_id] += 1
            assigned_counts_by_shift[shift["shift_id"]] += 1
            assigned_shift_details[volunteer_id].append(normalized_shift)
            assignments.append(
                {
                    "shift_id": shift["shift_id"],
                    "shift_date": shift["shift_date"],
                    "start_time": shift["start_time"],
                    "end_time": shift["end_time"],
                    "site": shift["site"],
                    "role": shift["role"],
                    "volunteer_id": volunteer_id,
                    "volunteer_name": volunteer["volunteer_name"],
                    "team": volunteer["team"],
                    "load_after_assignment": assignment_counts[volunteer_id],
                }
            )

    required_by_shift = {shift["shift_id"]: shift["required_count"] for shift in shifts}
    gaps_by_shift = {
        shift_id: required_by_shift[shift_id] - assigned_counts_by_shift[shift_id]
        for shift_id in required_by_shift
    }
    return assignments, volunteers, assigned_counts_by_shift, gaps_by_shift


def test_volunteer_shift_plan_workbook():
    assert OUTPUT_FILE.exists(), "missing /app/workspace/volunteer_shift_plan.xlsx"

    assignments, volunteers, assigned_counts_by_shift, gaps_by_shift = build_expected()

    formula_wb = load_workbook(OUTPUT_FILE, data_only=False)
    value_wb = load_workbook(OUTPUT_FILE, data_only=True)
    try:
        assert formula_wb.sheetnames == ["班次需求", "分配结果", "人员负载", "缺口概览"], (
            f"unexpected sheets: {formula_wb.sheetnames}"
        )

        needs_formula = formula_wb["班次需求"]
        needs_value = value_wb["班次需求"]
        expected_needs_header = [
            "shift_id",
            "shift_date",
            "start_time",
            "end_time",
            "site",
            "role",
            "required_count",
            "assigned_count",
            "gap_count",
        ]
        assert [needs_value.cell(row=1, column=col).value for col in range(1, 10)] == expected_needs_header
        assert needs_value.max_row == 7
        assert needs_value.max_column == 9

        shifts = read_shifts(SHIFTS_FILE)
        for row_index, shift in enumerate(shifts, start=2):
            assert needs_value[f"A{row_index}"].value == shift["shift_id"]
            assert normalize_date(needs_value[f"B{row_index}"].value) == shift["shift_date"]
            assert normalize_time(needs_value[f"C{row_index}"].value) == shift["start_time"]
            assert normalize_time(needs_value[f"D{row_index}"].value) == shift["end_time"]
            assert needs_value[f"E{row_index}"].value == shift["site"]
            assert needs_value[f"F{row_index}"].value == shift["role"]
            assert normalize_int(needs_value[f"G{row_index}"].value) == shift["required_count"]
            assert normalize_int(needs_value[f"H{row_index}"].value) == assigned_counts_by_shift[shift["shift_id"]]
            assert normalize_int(needs_value[f"I{row_index}"].value) == gaps_by_shift[shift["shift_id"]]
            for column in ("H", "I"):
                value = needs_formula[f"{column}{row_index}"].value
                assert isinstance(value, str) and value.startswith("="), f"missing formula in 班次需求!{column}{row_index}"

        assign_formula = formula_wb["分配结果"]
        assign_value = value_wb["分配结果"]
        expected_assign_header = [
            "shift_id",
            "shift_date",
            "start_time",
            "end_time",
            "site",
            "role",
            "volunteer_id",
            "volunteer_name",
            "team",
            "load_after_assignment",
            "overlap_flag",
        ]
        assert [assign_value.cell(row=1, column=col).value for col in range(1, 12)] == expected_assign_header
        assert assign_value.max_row == len(assignments) + 1
        assert assign_value.max_column == 11

        for row_index, expected in enumerate(assignments, start=2):
            assert assign_value[f"A{row_index}"].value == expected["shift_id"]
            assert normalize_date(assign_value[f"B{row_index}"].value) == expected["shift_date"]
            assert normalize_time(assign_value[f"C{row_index}"].value) == expected["start_time"]
            assert normalize_time(assign_value[f"D{row_index}"].value) == expected["end_time"]
            assert assign_value[f"E{row_index}"].value == expected["site"]
            assert assign_value[f"F{row_index}"].value == expected["role"]
            assert assign_value[f"G{row_index}"].value == expected["volunteer_id"]
            assert assign_value[f"H{row_index}"].value == expected["volunteer_name"]
            assert assign_value[f"I{row_index}"].value == expected["team"]
            assert normalize_int(assign_value[f"J{row_index}"].value) == expected["load_after_assignment"]
            assert (assign_value[f"K{row_index}"].value or "") == ""
            for column in ("J", "K"):
                value = assign_formula[f"{column}{row_index}"].value
                assert isinstance(value, str) and value.startswith("="), f"missing formula in 分配结果!{column}{row_index}"

        load_formula = formula_wb["人员负载"]
        load_value = value_wb["人员负载"]
        expected_load_header = [
            "volunteer_id",
            "volunteer_name",
            "team",
            "max_shifts",
            "assigned_shifts",
            "remaining_capacity",
            "conflict_flag",
        ]
        assert [load_value.cell(row=1, column=col).value for col in range(1, 8)] == expected_load_header
        assert load_value.max_row == len(volunteers) + 1
        assert load_value.max_column == 7

        assignment_totals = {}
        for assignment in assignments:
            assignment_totals.setdefault(assignment["volunteer_id"], 0)
            assignment_totals[assignment["volunteer_id"]] += 1

        for row_index, volunteer in enumerate(volunteers, start=2):
            assigned = assignment_totals.get(volunteer["volunteer_id"], 0)
            assert load_value[f"A{row_index}"].value == volunteer["volunteer_id"]
            assert load_value[f"B{row_index}"].value == volunteer["volunteer_name"]
            assert load_value[f"C{row_index}"].value == volunteer["team"]
            assert normalize_int(load_value[f"D{row_index}"].value) == volunteer["max_shifts"]
            assert normalize_int(load_value[f"E{row_index}"].value) == assigned
            assert normalize_int(load_value[f"F{row_index}"].value) == volunteer["max_shifts"] - assigned
            assert (load_value[f"G{row_index}"].value or "") == ""
            for column in ("E", "F", "G"):
                value = load_formula[f"{column}{row_index}"].value
                assert isinstance(value, str) and value.startswith("="), f"missing formula in 人员负载!{column}{row_index}"

        summary_formula = formula_wb["缺口概览"]
        summary_value = value_wb["缺口概览"]
        assert [summary_value["A1"].value, summary_value["B1"].value] == ["metric", "value"]
        assert summary_value.max_row == 8
        assert summary_value.max_column == 2

        expected_metrics = [
            ("total_required", 11),
            ("total_assigned", len(assignments)),
            ("total_gap", 2),
            ("filled_shift_count", 5),
            ("unfilled_shift_count", 1),
            ("volunteers_used", 7),
            ("conflicted_assignment_rows", 0),
        ]
        for row_index, (metric, expected_value) in enumerate(expected_metrics, start=2):
            assert summary_value[f"A{row_index}"].value == metric
            assert normalize_int(summary_value[f"B{row_index}"].value) == expected_value
            formula = summary_formula[f"B{row_index}"].value
            assert isinstance(formula, str) and formula.startswith("="), f"missing formula in 缺口概览!B{row_index}"
    finally:
        formula_wb.close()
        value_wb.close()
