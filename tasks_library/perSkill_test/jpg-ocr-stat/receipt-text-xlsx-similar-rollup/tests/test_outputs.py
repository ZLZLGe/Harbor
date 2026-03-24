import os
from typing import Any

from openpyxl import load_workbook


OUTPUT_FILE = "/app/workspace/receipt_rollup.xlsx"
EXPECTED_ROWS = [
    ["filename", "date", "total_amount"],
    ["001.txt", "2024-07-16", "47.90"],
    ["002.txt", "2024-07-17", "102.30"],
    ["010.txt", "2024-07-18", "19.99"],
    ["014.txt", "", "8.00"],
    ["020.txt", "2024-07-19", "1234.56"],
    ["021.txt", "2024-07-20", ""],
    ["099.txt", "2024-07-21", "4.50"],
]


def normalize(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def read_rows(path: str) -> list[list[str]]:
    workbook = load_workbook(path, data_only=True)
    try:
        assert workbook.sheetnames == ["results"], f"unexpected sheets: {workbook.sheetnames}"
        sheet = workbook["results"]
        rows: list[list[str]] = []
        for row in sheet.iter_rows(
            min_row=1,
            max_row=sheet.max_row,
            min_col=1,
            max_col=sheet.max_column,
            values_only=True,
        ):
            rows.append([normalize(cell) for cell in row])
        return rows
    finally:
        workbook.close()


def test_receipt_rollup_workbook():
    assert os.path.exists(OUTPUT_FILE), "missing /app/workspace/receipt_rollup.xlsx"

    actual_rows = read_rows(OUTPUT_FILE)

    assert actual_rows == EXPECTED_ROWS, (
        "workbook content mismatch\n"
        f"actual: {actual_rows}\n"
        f"expected: {EXPECTED_ROWS}"
    )

    data_rows = actual_rows[1:]
    filenames = [row[0] for row in data_rows]
    assert filenames == sorted(filenames), "rows are not sorted by filename"

    for row in data_rows:
        assert len(row) == 3, f"unexpected column count in row: {row}"
        date_value = row[1]
        total_value = row[2]
        if date_value:
            assert len(date_value) == 10 and date_value.count("-") == 2, f"date is not ISO formatted: {date_value}"
        if total_value:
            whole, frac = total_value.split(".")
            assert len(frac) == 2, f"amount must have two decimal places: {total_value}"
            assert whole.replace("-", "").isdigit(), f"invalid amount: {total_value}"
