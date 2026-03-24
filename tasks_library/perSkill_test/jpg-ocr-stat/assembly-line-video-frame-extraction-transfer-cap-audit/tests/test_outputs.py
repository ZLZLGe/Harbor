from __future__ import annotations

from pathlib import Path
from typing import Any

import cv2
from openpyxl import load_workbook


WORKSPACE = Path("/app/workspace")
OUTPUT_PATH = WORKSPACE / "cap_audit.xlsx"
FRAME_DIR = WORKSPACE / "cap_audit_frames"
EXPECTED_DETAIL_ROWS = [
    ["1", "00:00:00", "cap_audit_frames/beat_01.jpg", "2", "1", "3"],
    ["2", "00:00:03", "cap_audit_frames/beat_02.jpg", "1", "3", "4"],
    ["3", "00:00:06", "cap_audit_frames/beat_03.jpg", "4", "0", "4"],
    ["4", "00:00:09", "cap_audit_frames/beat_04.jpg", "0", "2", "2"],
    ["5", "00:00:12", "cap_audit_frames/beat_05.jpg", "3", "2", "5"],
    ["6", "00:00:15", "cap_audit_frames/beat_06.jpg", "1", "1", "2"],
]
EXPECTED_TOTAL_ROW = ["TOTAL", "", "", "11", "9", "20"]
EXPECTED_HEADER = ["beat_index", "timestamp", "frame_file", "red_caps", "blue_caps", "total_caps"]
EXPECTED_FRAME_SIZE = (560, 160)


def cell_to_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def read_rows(path: Path) -> list[list[str]]:
    workbook = load_workbook(path, data_only=True)
    try:
        assert workbook.sheetnames == ["audit"], f"workbook must contain exactly one sheet named audit, got {workbook.sheetnames}"
        sheet = workbook["audit"]
        rows: list[list[str]] = []
        for row_index in range(1, (sheet.max_row or 0) + 1):
            rows.append([
                cell_to_string(sheet.cell(row=row_index, column=column_index).value)
                for column_index in range(1, (sheet.max_column or 0) + 1)
            ])
        return rows
    finally:
        workbook.close()


def build_mask(hsv_image, color_name: str):
    if color_name == "red":
        return cv2.bitwise_or(
            cv2.inRange(hsv_image, (0, 120, 120), (10, 255, 255)),
            cv2.inRange(hsv_image, (170, 120, 120), (180, 255, 255)),
        )
    if color_name == "blue":
        return cv2.inRange(hsv_image, (90, 120, 120), (125, 255, 255))
    raise ValueError(color_name)


def count_caps(image, color_name: str) -> int:
    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
    mask = build_mask(hsv, color_name)
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel)
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)
    component_count, _, stats, _ = cv2.connectedComponentsWithStats(mask)

    hits = 0
    for label in range(1, component_count):
        area = int(stats[label, cv2.CC_STAT_AREA])
        width = int(stats[label, cv2.CC_STAT_WIDTH])
        height = int(stats[label, cv2.CC_STAT_HEIGHT])
        if area >= 250 and 12 <= width <= 48 and 12 <= height <= 48:
            hits += 1
    return hits


def main() -> None:
    assert OUTPUT_PATH.exists(), "missing /app/workspace/cap_audit.xlsx"
    assert FRAME_DIR.exists(), "missing /app/workspace/cap_audit_frames"

    rows = read_rows(OUTPUT_PATH)
    assert rows, "cap_audit.xlsx is empty"
    assert rows[0] == EXPECTED_HEADER, f"header mismatch: {rows[0]}"
    assert len(rows) == 1 + len(EXPECTED_DETAIL_ROWS) + 1, f"unexpected row count: {len(rows)}"
    assert rows[1:-1] == EXPECTED_DETAIL_ROWS, (
        "detail rows mismatch.\n"
        f"Actual: {rows[1:-1]}\n"
        f"Expected: {EXPECTED_DETAIL_ROWS}"
    )
    assert rows[-1] == EXPECTED_TOTAL_ROW, (
        "total row mismatch.\n"
        f"Actual: {rows[-1]}\n"
        f"Expected: {EXPECTED_TOTAL_ROW}"
    )

    total_red = 0
    total_blue = 0
    total_caps = 0

    for detail_row in rows[1:-1]:
        beat_index, _, frame_rel, red_caps, blue_caps, total_row_caps = detail_row
        frame_path = WORKSPACE / frame_rel
        assert frame_path.exists(), f"missing extracted frame: {frame_path}"
        assert frame_path.name == f"beat_{int(beat_index):02d}.jpg", f"unexpected frame name: {frame_path.name}"

        image = cv2.imread(str(frame_path))
        assert image is not None, f"cannot read frame: {frame_path}"
        height, width = image.shape[:2]
        assert (width, height) == EXPECTED_FRAME_SIZE, (
            f"cropped frame size mismatch for {frame_path}: got {(width, height)}, expected {EXPECTED_FRAME_SIZE}"
        )

        detected_red = count_caps(image, "red")
        detected_blue = count_caps(image, "blue")
        assert detected_red == int(red_caps), (
            f"red cap count mismatch for {frame_path}: detected {detected_red}, workbook {red_caps}"
        )
        assert detected_blue == int(blue_caps), (
            f"blue cap count mismatch for {frame_path}: detected {detected_blue}, workbook {blue_caps}"
        )
        assert int(total_row_caps) == int(red_caps) + int(blue_caps), (
            f"total_caps must equal red_caps + blue_caps for {frame_path}"
        )

        total_red += int(red_caps)
        total_blue += int(blue_caps)
        total_caps += int(total_row_caps)

    assert rows[-1][3:] == [str(total_red), str(total_blue), str(total_caps)], (
        "summary totals do not match detail rows.\n"
        f"Summary: {rows[-1][3:]}\n"
        f"Computed: {[str(total_red), str(total_blue), str(total_caps)]}"
    )


if __name__ == "__main__":
    main()
