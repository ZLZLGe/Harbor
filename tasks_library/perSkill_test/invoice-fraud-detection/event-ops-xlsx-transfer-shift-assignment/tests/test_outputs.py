from collections import defaultdict
import os

from openpyxl import load_workbook


INPUT_PATH = "/root/event_shift_template.xlsx"
OUTPUT_PATH = "/root/event_shift_plan.xlsx"
EXPECTED_SHEETS = [
    "Volunteer Roster",
    "Availability Matrix",
    "Role Qualifications",
    "Shift Demand",
    "Assignments",
    "Coverage Summary",
]
ASSIGNMENT_HEADERS = (
    "shift_id",
    "shift_date",
    "zone",
    "role",
    "slot_number",
    "volunteer_id",
    "volunteer_name",
    "team",
)
SUMMARY_HEADERS = (
    "shift_id",
    "shift_date",
    "zone",
    "role",
    "required_count",
    "assigned_count",
    "gap_count",
    "status",
)


def normalize_row(row):
    normalized = []
    for value in row:
        if isinstance(value, float):
            normalized.append(round(value, 6))
        else:
            normalized.append(value)
    return tuple(normalized)


def full_sheet_values(sheet):
    return [normalize_row(row) for row in sheet.iter_rows(values_only=True)]


def non_empty_rows(sheet, min_row, max_col):
    rows = []
    for row in sheet.iter_rows(min_row=min_row, max_col=max_col, values_only=True):
        if any(value not in (None, "") for value in row):
            rows.append(normalize_row(row))
    return rows


def build_expected_outputs(workbook):
    roster_sheet = workbook["Volunteer Roster"]
    availability_sheet = workbook["Availability Matrix"]
    qualifications_sheet = workbook["Role Qualifications"]
    demand_sheet = workbook["Shift Demand"]

    roster = {}
    for volunteer_id, volunteer_name, team, max_shifts, assignment_rank in roster_sheet.iter_rows(min_row=3, values_only=True):
        if volunteer_id in (None, ""):
            continue
        roster[volunteer_id] = {
            "volunteer_name": volunteer_name,
            "team": team,
            "max_shifts": int(max_shifts),
            "assignment_rank": int(assignment_rank),
        }

    shift_ids = [cell.value for cell in availability_sheet[2][1:]]
    availability = {}
    for row in availability_sheet.iter_rows(min_row=4, values_only=True):
        volunteer_id = row[0]
        if volunteer_id in (None, ""):
            continue
        availability[volunteer_id] = {
            shift_id
            for shift_id, marker in zip(shift_ids, row[1:])
            if isinstance(marker, str) and marker.strip().upper() == "Y"
        }

    roles = [cell.value for cell in qualifications_sheet[2][2:]]
    qualifications = {}
    for row in qualifications_sheet.iter_rows(min_row=3, values_only=True):
        volunteer_id = row[0]
        if volunteer_id in (None, ""):
            continue
        qualifications[volunteer_id] = {
            role
            for role, marker in zip(roles, row[2:])
            if isinstance(marker, str) and marker.strip().upper() == "Y"
        }

    demand_rows = []
    for shift_id, shift_date, zone, role, required_count, _shift_label in demand_sheet.iter_rows(min_row=3, values_only=True):
        if shift_id in (None, ""):
            continue
        demand_rows.append((shift_id, shift_date, zone, role, int(required_count)))

    assignment_counts = defaultdict(int)
    assigned_by_shift = defaultdict(set)
    assignment_rows = []
    summary_rows = []

    for shift_id, shift_date, zone, role, required_count in demand_rows:
        assigned_count = 0
        for slot_number in range(1, required_count + 1):
            candidates = []
            for volunteer_id, volunteer in roster.items():
                if shift_id not in availability.get(volunteer_id, set()):
                    continue
                if role not in qualifications.get(volunteer_id, set()):
                    continue
                if assignment_counts[volunteer_id] >= volunteer["max_shifts"]:
                    continue
                if volunteer_id in assigned_by_shift[shift_id]:
                    continue
                candidates.append(
                    (
                        assignment_counts[volunteer_id],
                        volunteer["assignment_rank"],
                        volunteer_id,
                    )
                )

            if not candidates:
                continue

            _, _, volunteer_id = min(candidates)
            volunteer = roster[volunteer_id]
            assignment_counts[volunteer_id] += 1
            assigned_by_shift[shift_id].add(volunteer_id)
            assigned_count += 1
            assignment_rows.append(
                (
                    shift_id,
                    shift_date,
                    zone,
                    role,
                    slot_number,
                    volunteer_id,
                    volunteer["volunteer_name"],
                    volunteer["team"],
                )
            )

        gap_count = required_count - assigned_count
        summary_rows.append(
            (
                shift_id,
                shift_date,
                zone,
                role,
                required_count,
                assigned_count,
                gap_count,
                "Covered" if gap_count == 0 else "Understaffed",
            )
        )

    return assignment_rows, summary_rows


class TestOutputs:
    def test_file_exists(self):
        assert os.path.exists(OUTPUT_PATH)

    def test_workbook_structure_and_assignments(self):
        input_wb = load_workbook(INPUT_PATH, data_only=True)
        output_wb = load_workbook(OUTPUT_PATH, data_only=True)

        assert input_wb.sheetnames == EXPECTED_SHEETS
        assert output_wb.sheetnames == EXPECTED_SHEETS

        for sheet_name in EXPECTED_SHEETS[:4]:
            assert full_sheet_values(output_wb[sheet_name]) == full_sheet_values(input_wb[sheet_name]), (
                f"Sheet '{sheet_name}' was modified but should remain unchanged."
            )

        assignments_sheet = output_wb["Assignments"]
        summary_sheet = output_wb["Coverage Summary"]

        assert assignments_sheet["A1"].value == input_wb["Assignments"]["A1"].value
        assert summary_sheet["A1"].value == input_wb["Coverage Summary"]["A1"].value
        assert tuple(cell.value for cell in assignments_sheet[2][: len(ASSIGNMENT_HEADERS)]) == ASSIGNMENT_HEADERS
        assert tuple(cell.value for cell in summary_sheet[2][: len(SUMMARY_HEADERS)]) == SUMMARY_HEADERS

        expected_assignments, expected_summary = build_expected_outputs(input_wb)
        actual_assignments = non_empty_rows(assignments_sheet, 3, len(ASSIGNMENT_HEADERS))
        actual_summary = non_empty_rows(summary_sheet, 3, len(SUMMARY_HEADERS))

        assert actual_assignments == expected_assignments
        assert actual_summary == expected_summary

    def test_assignment_constraints(self):
        output_wb = load_workbook(OUTPUT_PATH, data_only=True)
        roster_sheet = output_wb["Volunteer Roster"]
        assignments_sheet = output_wb["Assignments"]
        summary_sheet = output_wb["Coverage Summary"]

        max_shifts = {}
        for volunteer_id, _name, _team, max_shift_value, _rank in roster_sheet.iter_rows(min_row=3, values_only=True):
            if volunteer_id in (None, ""):
                continue
            max_shifts[volunteer_id] = int(max_shift_value)

        assigned_counts = defaultdict(int)
        shift_volunteers = defaultdict(set)
        for shift_id, _shift_date, _zone, _role, _slot_number, volunteer_id, _volunteer_name, _team in assignments_sheet.iter_rows(
            min_row=3, max_col=len(ASSIGNMENT_HEADERS), values_only=True
        ):
            if volunteer_id in (None, ""):
                continue
            assigned_counts[volunteer_id] += 1
            assert volunteer_id not in shift_volunteers[shift_id], f"{volunteer_id} was assigned more than once in {shift_id}"
            shift_volunteers[shift_id].add(volunteer_id)

        for volunteer_id, count in assigned_counts.items():
            assert count <= max_shifts[volunteer_id], f"{volunteer_id} exceeds max_shifts"

        for _shift_id, _shift_date, _zone, _role, required_count, assigned_count, gap_count, status in summary_sheet.iter_rows(
            min_row=3, max_col=len(SUMMARY_HEADERS), values_only=True
        ):
            if required_count in (None, ""):
                continue
            assert int(required_count) - int(assigned_count) == int(gap_count)
            assert status == ("Covered" if int(gap_count) == 0 else "Understaffed")
