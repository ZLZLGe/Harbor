from pathlib import Path

import openpyxl

OUTPUT_FILE = "shares_transfer_recovered.xlsx"
CHECKS = [('Directorate Shares (%)', 'F5', 26.08, 0.1), ('Directorate Shares (%)', 'B10', 31.35, 0.1), ('Growth Analysis', 'B7', 1534, 0.0), ('Growth Analysis', 'B8', 7444.4, 0.5), ('Growth Analysis', 'E4', 8.58, 0.1), ('Growth Analysis', 'E5', 5047, 0.0)]

def load_workbook_checked():
    path = Path(OUTPUT_FILE)
    assert path.exists(), f"missing output: {OUTPUT_FILE}"
    return openpyxl.load_workbook(path)

def assert_expected_cells(wb):
    for sheet_name, cell, expected, tolerance in CHECKS:
        actual = wb[sheet_name][cell].value
        if tolerance == 0:
            assert actual == expected, f"{sheet_name}!{cell} expected {expected}, got {actual}"
        else:
            assert abs(actual - expected) < tolerance, f"{sheet_name}!{cell} expected {expected}, got {actual}"

def assert_no_placeholders(wb):
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                assert value != "???", f"placeholder remains in {sheet_name}"

def assert_share_formulas(wb):
    budget = wb["Budget by Directorate"]
    shares = wb["Directorate Shares (%)"]
    f5 = round(budget["F5"].value / budget["K5"].value * 100, 2)
    b10 = round(budget["B10"].value / budget["K10"].value * 100, 2)
    assert abs(shares["F5"].value - f5) < 0.1
    assert abs(shares["B10"].value - b10) < 0.1

def assert_growth_cross_sheet(wb):
    budget = wb["Budget by Directorate"]
    growth = wb["Growth Analysis"]
    cagr = round(((budget["E13"].value / budget["E8"].value) ** 0.2 - 1) * 100, 2)
    assert abs(growth["E4"].value - cagr) < 0.1
    assert growth["E5"].value == budget["E8"].value

def main():
    wb = load_workbook_checked()
    assert_expected_cells(wb)
    assert_no_placeholders(wb)
    assert_share_formulas(wb)
    assert_growth_cross_sheet(wb)

if __name__ == "__main__":
    main()
