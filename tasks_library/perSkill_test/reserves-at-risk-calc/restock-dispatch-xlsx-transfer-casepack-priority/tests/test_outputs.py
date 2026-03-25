from math import ceil, isclose
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

INPUT = Path("/root/data/restock_planner.xlsx")
OUTPUT = Path("/root/output/restock_dispatch_board.xlsx")
TOL = 1e-6


def load_rows(ws):
    rows = []
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value in (None, ""):
            continue
        rows.append(row)
    return rows


def build_expected():
    wb = load_workbook(INPUT, data_only=True)
    demand_ws = wb["Demand"]
    on_hand_ws = wb["On Hand"]
    inbound_ws = wb["Inbound"]
    lead_ws = wb["Lead Times"]
    case_ws = wb["Case Packs"]

    on_hand = {}
    for row in load_rows(on_hand_ws):
        on_hand[(on_hand_ws[f"A{row}"].value, on_hand_ws[f"B{row}"].value)] = float(on_hand_ws[f"C{row}"].value)

    lead_times = {}
    for row in load_rows(lead_ws):
        lead_times[lead_ws[f"A{row}"].value] = float(lead_ws[f"B{row}"].value)

    case_packs = {}
    for row in load_rows(case_ws):
        case_packs[case_ws[f"A{row}"].value] = float(case_ws[f"B{row}"].value)

    inbound = {}
    for row in load_rows(inbound_ws):
        key = (inbound_ws[f"A{row}"].value, inbound_ws[f"B{row}"].value)
        inbound.setdefault(key, []).append(
            (float(inbound_ws[f"C{row}"].value), float(inbound_ws[f"D{row}"].value))
        )

    rows = []
    for row in load_rows(demand_ws):
        store = demand_ws[f"A{row}"].value
        sku = demand_ws[f"B{row}"].value
        avg_daily = float(demand_ws[f"C{row}"].value)
        target_cover = float(demand_ws[f"D{row}"].value)
        lead_days = lead_times[store]
        inbound_lines = inbound.get((store, sku), [])
        inbound_before = sum(qty for eta, qty in inbound_lines if eta <= lead_days)
        total_inbound = sum(qty for _, qty in inbound_lines)
        gap = max(avg_daily * lead_days - on_hand[(store, sku)] - inbound_before, 0)
        net_need = max(avg_daily * (target_cover + lead_days) - on_hand[(store, sku)] - total_inbound, 0)
        case_pack = case_packs[sku]
        suggested = 0 if net_need == 0 else ceil(net_need / case_pack) * case_pack
        post_cover = (on_hand[(store, sku)] + total_inbound + suggested) / avg_daily - lead_days
        urgency = gap * 100 + max(target_cover - post_cover, 0)
        if gap >= avg_daily * 2:
            warning = "CRITICAL"
        elif gap > 0:
            warning = "WATCH"
        else:
            warning = "OK"
        rows.append(
            {
                "store": store,
                "sku": sku,
                "avg_daily": avg_daily,
                "target_cover": target_cover,
                "on_hand": on_hand[(store, sku)],
                "lead_days": lead_days,
                "inbound_before": inbound_before,
                "total_inbound": total_inbound,
                "gap": gap,
                "net_need": net_need,
                "case_pack": case_pack,
                "suggested": suggested,
                "post_cover": post_cover,
                "urgency": urgency,
                "warning": warning,
            }
        )

    ranked = sorted(rows, key=lambda item: (-item["urgency"], -item["gap"], item["store"], item["sku"]))
    wb.close()
    return rows, ranked


def load_books():
    assert OUTPUT.exists(), f"missing output workbook: {OUTPUT}"
    return load_workbook(OUTPUT, data_only=False), load_workbook(OUTPUT, data_only=True)


def numeric(value):
    assert isinstance(value, (int, float)), f"expected numeric value, got {value!r}"
    return float(value)


def test_output_exists_and_sheet_names():
    wb_formula, wb_values = load_books()
    expected = ["Demand", "On Hand", "Inbound", "Lead Times", "Case Packs", "Restock Plan", "Dispatch Board"]
    assert wb_formula.sheetnames == expected
    assert wb_values.sheetnames == expected


def test_restock_plan_values():
    expected_rows, _ = build_expected()
    _, wb_values = load_books()
    ws = wb_values["Restock Plan"]

    for row_idx, expected in enumerate(expected_rows, start=2):
        assert ws[f"A{row_idx}"].value == expected["store"]
        assert ws[f"B{row_idx}"].value == expected["sku"]
        assert numeric(ws[f"C{row_idx}"].value) == expected["avg_daily"]
        assert numeric(ws[f"D{row_idx}"].value) == expected["target_cover"]
        assert numeric(ws[f"E{row_idx}"].value) == expected["on_hand"]
        assert numeric(ws[f"F{row_idx}"].value) == expected["lead_days"]
        assert numeric(ws[f"G{row_idx}"].value) == expected["inbound_before"]
        assert numeric(ws[f"H{row_idx}"].value) == expected["total_inbound"]
        assert numeric(ws[f"I{row_idx}"].value) == expected["gap"]
        assert numeric(ws[f"J{row_idx}"].value) == expected["net_need"]
        assert numeric(ws[f"K{row_idx}"].value) == expected["case_pack"]
        assert numeric(ws[f"L{row_idx}"].value) == expected["suggested"]
        assert isclose(numeric(ws[f"M{row_idx}"].value), expected["post_cover"], rel_tol=0, abs_tol=TOL)
        assert isclose(numeric(ws[f"N{row_idx}"].value), expected["urgency"], rel_tol=0, abs_tol=TOL)
        assert ws[f"O{row_idx}"].value == expected["warning"]


def test_dispatch_board_sorted_results():
    _, expected_ranked = build_expected()
    _, wb_values = load_books()
    ws = wb_values["Dispatch Board"]

    for rank, expected in enumerate(expected_ranked, start=1):
        row = rank + 1
        assert ws[f"A{row}"].value == rank
        assert ws[f"B{row}"].value == expected["store"]
        assert ws[f"C{row}"].value == expected["sku"]
        assert numeric(ws[f"D{row}"].value) == expected["gap"]
        assert numeric(ws[f"E{row}"].value) == expected["suggested"]
        assert isclose(numeric(ws[f"F{row}"].value), expected["post_cover"], rel_tol=0, abs_tol=TOL)
        assert ws[f"G{row}"].value == expected["warning"]
        assert isclose(numeric(ws[f"H{row}"].value), expected["urgency"], rel_tol=0, abs_tol=TOL)


def test_required_formulas_present():
    expected_rows, expected_ranked = build_expected()
    wb_formula, _ = load_books()
    plan = wb_formula["Restock Plan"]
    board = wb_formula["Dispatch Board"]

    for row_idx, _ in enumerate(expected_rows, start=2):
        for col in "ABCDEFGHIJKLMNO":
            value = plan[f"{col}{row_idx}"].value
            assert isinstance(value, str) and value.startswith("="), f"missing formula in Restock Plan!{col}{row_idx}"

    for rank, _ in enumerate(expected_ranked, start=1):
        row = rank + 1
        assert board[f"A{row}"].value == rank
        for col in "BCDEFGH":
            value = board[f"{col}{row}"].value
            assert isinstance(value, str) and value.startswith("="), f"missing formula in Dispatch Board!{col}{row}"


def test_no_macros_or_formula_errors():
    _, wb_values = load_books()

    with ZipFile(OUTPUT) as zf:
        names = zf.namelist()
    assert not any("vbaProject" in name or name.endswith(".bin") for name in names)

    for sheet in ["Restock Plan", "Dispatch Board"]:
        ws = wb_values[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    assert not cell.value.startswith("#"), f"formula error in {sheet}!{cell.coordinate}: {cell.value}"
