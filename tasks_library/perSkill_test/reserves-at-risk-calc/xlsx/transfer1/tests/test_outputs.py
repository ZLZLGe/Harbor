
import os
from pathlib import Path

import openpyxl

OUTPUT = Path(os.environ.get("OUTPUT_FILE", "/root/jet_fuel_transfer1_completed.xlsx"))
EXPECTED_SHEETS = ["Prices", "Hedges", "Summary"]
EXPECTED_FORMULAS = {
    ("Prices", "C3"): "=LN(B3/B2)*100",
    ("Prices", "D5"): "=STDEV.S(C3:C5)",
    ("Prices", "D9"): "=STDEV.S(C7:C9)",
    ("Hedges", "D2"): "=B2*C2",
    ("Hedges", "E3"): "=D3*Summary!$B$2",
    ("Hedges", "F4"): "=E4*Summary!$B$4/100*1.65",
    ("Summary", "B2"): "=AVERAGE(Prices!B2:B9)",
    ("Summary", "B4"): "=B3*SQRT(12)",
    ("Summary", "B6"): "=SUM(Hedges!E2:E4)",
    ("Summary", "B8"): "=INDEX(Hedges!A2:A4,MATCH(MAX(Hedges!E2:E4),Hedges!E2:E4,0))",
}


def main():
    assert OUTPUT.exists(), f"missing workbook: {OUTPUT}"
    wb = openpyxl.load_workbook(OUTPUT, data_only=False)
    assert wb.sheetnames == EXPECTED_SHEETS, wb.sheetnames
    for (sheet, cell), expected in EXPECTED_FORMULAS.items():
        actual = wb[sheet][cell].value
        assert actual == expected, f"{sheet}!{cell} expected {expected!r}, got {actual!r}"
    assert wb["Hedges"]["A2"].value == "Domestic"
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str):
                    assert "TODO" not in value, f"TODO remains in {ws.title}"


if __name__ == "__main__":
    main()
