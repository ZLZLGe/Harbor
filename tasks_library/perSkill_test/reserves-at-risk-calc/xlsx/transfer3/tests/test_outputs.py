
import os
from pathlib import Path

import openpyxl

OUTPUT = Path(os.environ.get("OUTPUT_FILE", "/root/cold_chain_transfer3_completed.xlsx"))
EXPECTED_SHEETS = ["Inputs", "Inventory", "Scenarios", "Scorecard"]
EXPECTED_FORMULAS = {
    ("Inventory", "E2"): "=B2*D2",
    ("Inventory", "F5"): "=B5/C5",
    ("Inventory", "G3"): "=E3*Inputs!$B$2",
    ("Inventory", "H4"): "=MAX(Inputs!$B$3-F4,0)*C4*D4*(1+Inputs!$B$4)",
    ("Scenarios", "B2"): "=SUM(Inventory!E2:E5)",
    ("Scenarios", "C4"): "=SUM(Inventory!H2:H5)",
    ("Scorecard", "B3"): '=COUNTIF(Inventory!F2:F5,"<"&Inputs!$B$3)',
    ("Scorecard", "B4"): "=INDEX(Inventory!A2:A5,MATCH(MAX(Inventory!G2:G5),Inventory!G2:G5,0))",
    ("Scorecard", "B6"): '=IF(B3=0,"GREEN",IF(B3=1,"AMBER","RED"))',
    ("Scorecard", "B7"): "=Scenarios!B3",
}


def main():
    assert OUTPUT.exists(), f"missing workbook: {OUTPUT}"
    wb = openpyxl.load_workbook(OUTPUT, data_only=False)
    assert wb.sheetnames == EXPECTED_SHEETS, wb.sheetnames
    for (sheet, cell), expected in EXPECTED_FORMULAS.items():
        actual = wb[sheet][cell].value
        assert actual == expected, f"{sheet}!{cell} expected {expected!r}, got {actual!r}"
    assert wb["Inputs"]["B2"].value == 0.18
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str):
                    assert "TODO" not in value, f"TODO remains in {ws.title}"


if __name__ == "__main__":
    main()
