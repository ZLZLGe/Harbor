
import os
from pathlib import Path

import openpyxl

OUTPUT = Path(os.environ.get("OUTPUT_FILE", "/root/reserve_stress_similar_completed.xlsx"))
EXPECTED_SHEETS = ["Answer", "Platinum price", "Value", "Volume", "Total Reserves"]
EXPECTED_FORMULAS = {
    ("Platinum price", "C3"): "=LN(B3/B2)*100",
    ("Platinum price", "C11"): "=LN(B11/B10)*100",
    ("Platinum price", "D6"): "=STDEV.S(C3:C6)",
    ("Platinum price", "D11"): "=STDEV.S(C8:C11)",
    ("Platinum price", "E11"): "=STDEV.S(C4:C11)",
    ("Answer", "C4"): "='Platinum price'!D11",
    ("Answer", "C5"): "=C4*SQRT(12)",
    ("Answer", "F12"): "=Volume!$B$2*AVERAGE('Platinum price'!$B$2:$B$7)",
    ("Answer", "C13"): "=C12*$C$3/100*$C$4",
    ("Answer", "C23"): "=XLOOKUP(C20,'Total Reserves'!$A$2:$A$4,'Total Reserves'!$B$2:$B$4)",
    ("Answer", "E24"): "=E22/E23*100",
}


def main():
    assert OUTPUT.exists(), f"missing workbook: {OUTPUT}"
    wb = openpyxl.load_workbook(OUTPUT, data_only=False)
    assert wb.sheetnames == EXPECTED_SHEETS, wb.sheetnames
    for (sheet, cell), expected in EXPECTED_FORMULAS.items():
        actual = wb[sheet][cell].value
        assert actual == expected, f"{sheet}!{cell} expected {expected!r}, got {actual!r}"
    assert wb["Answer"]["C3"].value == 1.28
    assert wb["Answer"]["C11"].value == "Asteria"
    assert wb["Answer"]["F11"].value == "Deltora"
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str):
                    assert "TODO" not in value, f"TODO remains in {ws.title}"


if __name__ == "__main__":
    main()
