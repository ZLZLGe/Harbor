
import os
from pathlib import Path

import openpyxl

OUTPUT = Path(os.environ.get("OUTPUT_FILE", "/root/cash_ladder_transfer2_completed.xlsx"))
EXPECTED_SHEETS = ["Curve", "Deposits", "Ladder", "Dashboard"]
EXPECTED_FORMULAS = {
    ("Deposits", "E2"): "=XLOOKUP(C2,Curve!$A$2:$A$5,Curve!$B$2:$B$5)",
    ("Deposits", "F5"): "=B5*E5*C5/365",
    ("Ladder", "B2"): "=SUMIF(Deposits!$D$2:$D$5,A2,Deposits!$B$2:$B$5)",
    ("Ladder", "C4"): "=SUMIF(Deposits!$D$2:$D$5,A4,Deposits!$F$2:$F$5)",
    ("Dashboard", "B2"): "=SUM(Deposits!B2:B5)",
    ("Dashboard", "B4"): "=SUMPRODUCT(Deposits!B2:B5,Deposits!E2:E5)/SUM(Deposits!B2:B5)",
    ("Dashboard", "B5"): "=INDEX(Ladder!A2:A4,MATCH(MAX(Ladder!C2:C4),Ladder!C2:C4,0))",
    ("Dashboard", "B6"): "=B3*0.6",
}


def main():
    assert OUTPUT.exists(), f"missing workbook: {OUTPUT}"
    wb = openpyxl.load_workbook(OUTPUT, data_only=False)
    assert wb.sheetnames == EXPECTED_SHEETS, wb.sheetnames
    for (sheet, cell), expected in EXPECTED_FORMULAS.items():
        actual = wb[sheet][cell].value
        assert actual == expected, f"{sheet}!{cell} expected {expected!r}, got {actual!r}"
    assert wb["Curve"]["A2"].value == 30
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str):
                    assert "TODO" not in value, f"TODO remains in {ws.title}"


if __name__ == "__main__":
    main()
