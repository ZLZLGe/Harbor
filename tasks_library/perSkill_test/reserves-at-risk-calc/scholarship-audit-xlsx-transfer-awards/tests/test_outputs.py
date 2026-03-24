import csv
import glob
from pathlib import Path

from openpyxl import load_workbook

OUTPUT_XLSX = Path("/root/output/scholarship_audit_result.xlsx")
INPUT_XLSX = Path("/root/data/scholarship-audit-template.xlsx")
CSV_PATTERN = "/root/output/review.csv.*"
TOL = 1e-6


def load_book(path, data_only):
    return load_workbook(path, data_only=data_only)


def sheet_csv_path(sheet_name):
    csv_files = sorted(glob.glob(CSV_PATTERN))
    if not csv_files:
        return None

    wb = load_book(OUTPUT_XLSX, data_only=False)
    try:
        idx = wb.sheetnames.index(sheet_name)
    finally:
        wb.close()

    candidate = Path(f"/root/output/review.csv.{idx}")
    if candidate.exists():
        return candidate
    return None


def load_csv_cells(sheet_name):
    csv_path = sheet_csv_path(sheet_name)
    data = {}
    if csv_path is None:
        return data

    with open(csv_path, encoding="utf-8", errors="ignore") as handle:
        reader = csv.reader(handle)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                col_letter = chr(ord("A") + col_idx - 1)
                cell = f"{col_letter}{row_idx}"
                if value == "":
                    data[cell] = None
                    continue
                try:
                    data[cell] = float(value)
                except ValueError:
                    data[cell] = value
    return data


def cell_value(ws, csv_cells, cell):
    value = ws[cell].value
    if value is not None and not isinstance(value, str):
        return value
    csv_value = csv_cells.get(cell)
    if csv_value is not None:
        return csv_value
    return value


def assert_close(actual, expected, label):
    assert actual is not None, f"{label}: expected {expected}, got None"
    assert abs(float(actual) - float(expected)) <= TOL, f"{label}: expected {expected}, got {actual}"


def expected_results():
    wb = load_book(INPUT_XLSX, data_only=True)
    roster_ws = wb["Student Roster"]
    grades_ws = wb["Term Grades"]
    bursar_ws = wb["Bursar Balances"]
    rules_ws = wb["Award Rules"]
    review_ws = wb["Award Review"]

    roster = {}
    for row in range(2, 10):
        student_id = int(roster_ws[f"A{row}"].value)
        roster[student_id] = {
            "name": f"{roster_ws[f'C{row}'].value} {roster_ws[f'B{row}'].value}",
            "tier": roster_ws[f"G{row}"].value,
        }

    grades = {}
    for row in range(2, grades_ws.max_row + 1):
        student_id = int(grades_ws[f"A{row}"].value)
        entry = grades.setdefault(student_id, {"attempted": 0.0, "passed": 0.0, "quality": 0.0})
        entry["attempted"] += float(grades_ws[f"C{row}"].value)
        entry["passed"] += float(grades_ws[f"D{row}"].value)
        entry["quality"] += float(grades_ws[f"E{row}"].value)

    bursar = {}
    for row in range(2, 10):
        student_id = int(bursar_ws[f"A{row}"].value)
        bursar[student_id] = {
            "balance": float(bursar_ws[f"B{row}"].value),
            "plan": bursar_ws[f"C{row}"].value,
            "hold": bursar_ws[f"D{row}"].value,
            "refund": bursar_ws[f"E{row}"].value,
        }

    rules = []
    award_amounts = {}
    for row in range(2, 5):
        tier = rules_ws[f"A{row}"].value
        min_gpa = float(rules_ws[f"B{row}"].value)
        min_passed = float(rules_ws[f"C{row}"].value)
        amount = float(rules_ws[f"D{row}"].value)
        rules.append((tier, min_gpa, min_passed, amount))
        award_amounts[tier] = amount

    student_order = [int(review_ws[f"A{row}"].value) for row in range(11, 19)]
    rows = []
    for student_id in student_order:
        grade_data = grades[student_id]
        gpa = grade_data["quality"] / grade_data["attempted"]
        earned_tier = "Ineligible"
        for tier, min_gpa, min_passed, _ in rules:
            if gpa >= min_gpa and grade_data["passed"] >= min_passed:
                earned_tier = tier
                break

        bursar_status = "Cleared"
        bursar_data = bursar[student_id]
        if not (
            (bursar_data["balance"] <= 500 or bursar_data["plan"] == "Yes")
            and bursar_data["hold"] == "None"
            and bursar_data["refund"] == "Yes"
        ):
            bursar_status = "Hold"

        if earned_tier == "Ineligible":
            review_flag = "Academic Ineligible"
        elif earned_tier == roster[student_id]["tier"]:
            review_flag = "Match"
        else:
            review_flag = "Mismatch"

        if review_flag == "Match" and bursar_status == "Cleared":
            final_status = "Approve"
            approved_amount = award_amounts[roster[student_id]["tier"]]
        elif review_flag == "Mismatch" and bursar_status == "Cleared":
            final_status = "Manual Review"
            approved_amount = 0.0
        else:
            final_status = "Hold"
            approved_amount = 0.0

        rows.append(
            {
                "student_id": student_id,
                "name": roster[student_id]["name"],
                "tier": roster[student_id]["tier"],
                "passed": grade_data["passed"],
                "gpa": gpa,
                "earned": earned_tier,
                "bursar": bursar_status,
                "flag": review_flag,
                "status": final_status,
                "amount": approved_amount,
            }
        )

    summary = {
        "C3": len(student_order),
        "C4": sum(1 for row in rows if row["earned"] != "Ineligible"),
        "C5": sum(1 for row in rows if row["flag"] == "Mismatch"),
        "C6": sum(1 for row in rows if row["bursar"] == "Hold"),
        "C7": sum(row["amount"] for row in rows),
    }

    wb.close()
    return rows, summary


def test_output_exists_and_sheets_are_preserved():
    assert OUTPUT_XLSX.exists(), f"Missing output workbook: {OUTPUT_XLSX}"
    wb = load_book(OUTPUT_XLSX, data_only=False)
    assert wb.sheetnames == [
        "Student Roster",
        "Term Grades",
        "Bursar Balances",
        "Award Rules",
        "Award Review",
    ]
    wb.close()


def test_review_table_values():
    rows, _ = expected_results()
    wb = load_book(OUTPUT_XLSX, data_only=True)
    csv_cells = load_csv_cells("Award Review")
    ws = wb["Award Review"]

    for offset, expected in enumerate(rows, start=11):
        assert int(cell_value(ws, csv_cells, f"A{offset}")) == expected["student_id"]
        assert cell_value(ws, csv_cells, f"B{offset}") == expected["name"]
        assert cell_value(ws, csv_cells, f"C{offset}") == expected["tier"]
        assert_close(cell_value(ws, csv_cells, f"D{offset}"), expected["passed"], f"D{offset}")
        assert_close(cell_value(ws, csv_cells, f"E{offset}"), expected["gpa"], f"E{offset}")
        assert cell_value(ws, csv_cells, f"F{offset}") == expected["earned"]
        assert cell_value(ws, csv_cells, f"G{offset}") == expected["bursar"]
        assert cell_value(ws, csv_cells, f"H{offset}") == expected["flag"]
        assert cell_value(ws, csv_cells, f"I{offset}") == expected["status"]
        assert_close(cell_value(ws, csv_cells, f"J{offset}"), expected["amount"], f"J{offset}")

    wb.close()


def test_summary_metrics():
    _, summary = expected_results()
    wb = load_book(OUTPUT_XLSX, data_only=True)
    csv_cells = load_csv_cells("Award Review")
    ws = wb["Award Review"]

    for cell, expected in summary.items():
        assert_close(cell_value(ws, csv_cells, cell), expected, cell)

    wb.close()


def test_formulas_are_present():
    wb = load_book(OUTPUT_XLSX, data_only=False)
    ws = wb["Award Review"]

    for row in range(11, 19):
        for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            value = ws[f"{col}{row}"].value
            assert isinstance(value, str) and value.startswith("="), f"Expected formula in {col}{row}, got {value!r}"

    for cell in ["C3", "C4", "C5", "C6", "C7"]:
        value = ws[cell].value
        assert isinstance(value, str) and value.startswith("="), f"Expected formula in {cell}, got {value!r}"

    assert "Student Roster" in ws["B11"].value
    assert "SUMIFS" in ws["D11"].value
    assert "Term Grades" in ws["E11"].value
    assert "Award Rules" in ws["F11"].value
    assert "Bursar Balances" in ws["G11"].value
    assert "Award Rules" in ws["J11"].value

    wb.close()


def test_no_excel_error_literals():
    wb = load_book(OUTPUT_XLSX, data_only=True)
    errors = []
    excel_errors = {"#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in excel_errors:
                    errors.append(f"{sheet_name}!{cell.coordinate}={cell.value}")

    wb.close()
    assert not errors, "Found Excel error literals:\n" + "\n".join(errors[:20])
