from pathlib import Path

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula


OUTPUT_FILE = Path("reservoir-drought-resilience.xlsx")
REQUIRED_SHEETS = [
    "Inflow_Plan",
    "Loss_Factors",
    "Demand_Forecast",
    "Project_Phases",
    "Storage_Rolling",
    "Restriction_Risk",
]


def load_workbook(data_only=False):
    assert OUTPUT_FILE.exists(), f"Output file not found: {OUTPUT_FILE}"
    return openpyxl.load_workbook(OUTPUT_FILE, data_only=data_only)


def get_formula_text(value):
    if value is None:
        return None
    if isinstance(value, ArrayFormula):
        return value.text
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def approx_equal(actual, expected, tol=1e-6):
    assert actual is not None, f"Expected {expected}, got None"
    assert abs(actual - expected) <= tol, f"Expected {expected}, got {actual}"


def assert_formula_range(ws, range_ref):
    for row in ws[range_ref]:
        for cell in row:
            assert (
                get_formula_text(cell.value) is not None
            ), f"{ws.title}!{cell.coordinate} must contain a formula"


def compute_expected_values(wb):
    rows = range(8, 17)

    inflow = wb["Inflow_Plan"]
    losses = wb["Loss_Factors"]
    demand = wb["Demand_Forecast"]
    phases = wb["Project_Phases"]
    storage = wb["Storage_Rolling"]

    expected = {sheet: {} for sheet in REQUIRED_SHEETS}

    phase_transfer = {}
    phase_saving = {}
    phase_evap = {}
    for row in rows:
        progress = phases[f"B{row}"].value
        phase_transfer[row] = progress * phases["B2"].value
        phase_saving[row] = progress * phases["B3"].value
        phase_evap[row] = progress * phases["B4"].value
        expected["Project_Phases"][f"C{row}"] = phase_transfer[row]
        expected["Project_Phases"][f"D{row}"] = phase_saving[row]
        expected["Project_Phases"][f"E{row}"] = phase_evap[row]

    inflow_baseline = {}
    inflow_resilient = {}
    for row in range(8, 11):
        inflow_baseline[row] = inflow[f"B{row}"].value
    for row in range(11, 17):
        inflow_baseline[row] = inflow["B10"].value * inflow[f"C{row}"].value * inflow["B2"].value
    for row in rows:
        inflow_resilient[row] = inflow_baseline[row] + phase_transfer[row]
        expected["Inflow_Plan"][f"D{row}"] = inflow_baseline[row]
        expected["Inflow_Plan"][f"E{row}"] = inflow_resilient[row]

    loss_baseline = {}
    loss_resilient = {}
    for row in rows:
        loss_baseline[row] = losses[f"B{row}"].value * losses["B2"].value / 100
        loss_resilient[row] = loss_baseline[row] * (1 - phase_evap[row])
        expected["Loss_Factors"][f"C{row}"] = loss_baseline[row]
        expected["Loss_Factors"][f"D{row}"] = phase_evap[row]
        expected["Loss_Factors"][f"E{row}"] = loss_resilient[row]

    demand_baseline = {}
    demand_resilient = {}
    for row in range(8, 11):
        demand_baseline[row] = demand[f"B{row}"].value
    for row in range(11, 17):
        demand_baseline[row] = (
            demand_baseline[row - 1] * (1 + demand["B2"].value) * demand[f"C{row}"].value
        )
    for row in rows:
        demand_resilient[row] = demand_baseline[row] * (1 - phase_saving[row])
        expected["Demand_Forecast"][f"D{row}"] = demand_baseline[row]
        expected["Demand_Forecast"][f"E{row}"] = phase_saving[row]
        expected["Demand_Forecast"][f"F{row}"] = demand_resilient[row]

    initial_storage = storage["B2"].value
    operating_release = storage["B3"].value
    dead_storage = storage["B4"].value
    stage_1 = storage["B5"].value
    stage_2 = storage["B6"].value
    stage_3 = storage["B7"].value

    baseline_end = {}
    resilient_end = {}
    for row in rows:
        baseline_start = initial_storage if row == 8 else baseline_end[row - 1]
        baseline_available = baseline_start + inflow_baseline[row] - loss_baseline[row] - operating_release
        baseline_supply = min(demand_baseline[row], max(baseline_available - dead_storage, 0))
        baseline_unmet = max(demand_baseline[row] - baseline_supply, 0)
        baseline_end[row] = max(baseline_available - baseline_supply, dead_storage)
        baseline_stage = (
            3
            if baseline_end[row] < stage_3
            else 2
            if baseline_end[row] < stage_2
            else 1
            if baseline_end[row] < stage_1
            else 0
        )

        resilient_start = initial_storage if row == 8 else resilient_end[row - 1]
        resilient_available = resilient_start + inflow_resilient[row] - loss_resilient[row] - operating_release
        resilient_supply = min(demand_resilient[row], max(resilient_available - dead_storage, 0))
        resilient_unmet = max(demand_resilient[row] - resilient_supply, 0)
        resilient_end[row] = max(resilient_available - resilient_supply, dead_storage)
        resilient_stage = (
            3
            if resilient_end[row] < stage_3
            else 2
            if resilient_end[row] < stage_2
            else 1
            if resilient_end[row] < stage_1
            else 0
        )

        expected["Storage_Rolling"][f"B{row}"] = baseline_start
        expected["Storage_Rolling"][f"C{row}"] = inflow_baseline[row]
        expected["Storage_Rolling"][f"D{row}"] = loss_baseline[row]
        expected["Storage_Rolling"][f"E{row}"] = demand_baseline[row]
        expected["Storage_Rolling"][f"F{row}"] = baseline_available
        expected["Storage_Rolling"][f"G{row}"] = baseline_supply
        expected["Storage_Rolling"][f"H{row}"] = baseline_unmet
        expected["Storage_Rolling"][f"I{row}"] = baseline_end[row]
        expected["Storage_Rolling"][f"J{row}"] = baseline_stage

        expected["Storage_Rolling"][f"K{row}"] = resilient_start
        expected["Storage_Rolling"][f"L{row}"] = inflow_resilient[row]
        expected["Storage_Rolling"][f"M{row}"] = loss_resilient[row]
        expected["Storage_Rolling"][f"N{row}"] = demand_resilient[row]
        expected["Storage_Rolling"][f"O{row}"] = resilient_available
        expected["Storage_Rolling"][f"P{row}"] = resilient_supply
        expected["Storage_Rolling"][f"Q{row}"] = resilient_unmet
        expected["Storage_Rolling"][f"R{row}"] = resilient_end[row]
        expected["Storage_Rolling"][f"S{row}"] = resilient_stage

    baseline_restriction_months = 0
    resilient_restriction_months = 0
    baseline_severe_months = 0
    resilient_severe_months = 0
    shortage_reduction_total = 0.0
    storage_gain_total = 0.0
    for row in rows:
        baseline_unmet = expected["Storage_Rolling"][f"H{row}"]
        resilient_unmet = expected["Storage_Rolling"][f"Q{row}"]
        baseline_stage = expected["Storage_Rolling"][f"J{row}"]
        resilient_stage = expected["Storage_Rolling"][f"S{row}"]
        baseline_storage = expected["Storage_Rolling"][f"I{row}"]
        resilient_storage = expected["Storage_Rolling"][f"R{row}"]

        baseline_restriction_flag = 1 if baseline_stage >= 1 else 0
        resilient_restriction_flag = 1 if resilient_stage >= 1 else 0
        baseline_severe_flag = 1 if baseline_stage >= 2 else 0
        resilient_severe_flag = 1 if resilient_stage >= 2 else 0
        monthly_improvement = baseline_unmet - resilient_unmet

        baseline_restriction_months += baseline_restriction_flag
        resilient_restriction_months += resilient_restriction_flag
        baseline_severe_months += baseline_severe_flag
        resilient_severe_months += resilient_severe_flag
        shortage_reduction_total += monthly_improvement
        storage_gain_total += resilient_storage - baseline_storage

        expected["Restriction_Risk"][f"B{row}"] = baseline_storage
        expected["Restriction_Risk"][f"C{row}"] = resilient_storage
        expected["Restriction_Risk"][f"D{row}"] = baseline_unmet
        expected["Restriction_Risk"][f"E{row}"] = resilient_unmet
        expected["Restriction_Risk"][f"F{row}"] = baseline_restriction_flag
        expected["Restriction_Risk"][f"G{row}"] = resilient_restriction_flag
        expected["Restriction_Risk"][f"H{row}"] = baseline_severe_flag
        expected["Restriction_Risk"][f"I{row}"] = resilient_severe_flag
        expected["Restriction_Risk"][f"J{row}"] = monthly_improvement

    expected["Restriction_Risk"]["B2"] = baseline_restriction_months
    expected["Restriction_Risk"]["B3"] = resilient_restriction_months
    expected["Restriction_Risk"]["B4"] = baseline_severe_months
    expected["Restriction_Risk"]["B5"] = resilient_severe_months
    expected["Restriction_Risk"]["B6"] = shortage_reduction_total
    expected["Restriction_Risk"]["B7"] = storage_gain_total / len(list(rows))

    return expected


def test_required_sheets_exist():
    wb = load_workbook()
    assert wb.sheetnames == REQUIRED_SHEETS


def test_inputs_and_month_axis_are_preserved():
    wb = load_workbook()

    inflow = wb["Inflow_Plan"]
    losses = wb["Loss_Factors"]
    demand = wb["Demand_Forecast"]
    phases = wb["Project_Phases"]
    storage = wb["Storage_Rolling"]

    assert inflow["B2"].value == 0.92
    assert losses["B2"].value == 13.8
    assert demand["B2"].value == 0.012
    assert phases["B2"].value == 12
    assert phases["B3"].value == 0.18
    assert phases["B4"].value == 0.25
    assert storage["B2"].value == 182
    assert storage["B3"].value == 8.5
    assert storage["B4"].value == 70
    assert storage["B5"].value == 125
    assert storage["B6"].value == 95
    assert storage["B7"].value == 80

    expected_months = [
        "2027-01",
        "2027-02",
        "2027-03",
        "2027-04",
        "2027-05",
        "2027-06",
        "2027-07",
        "2027-08",
        "2027-09",
    ]
    assert [inflow[f"A{row}"].value for row in range(8, 17)] == expected_months
    assert [phases[f"B{row}"].value for row in range(8, 17)] == [0, 0, 0, 0.5, 0.8, 1, 1, 1, 1]


def test_required_calculation_cells_remain_formulas():
    wb = load_workbook()
    inflow = wb["Inflow_Plan"]
    losses = wb["Loss_Factors"]
    demand = wb["Demand_Forecast"]
    phases = wb["Project_Phases"]
    storage = wb["Storage_Rolling"]
    risk = wb["Restriction_Risk"]

    assert_formula_range(inflow, "D8:E16")
    assert_formula_range(losses, "C8:E16")
    assert_formula_range(demand, "D8:F16")
    assert_formula_range(phases, "C8:E16")
    assert_formula_range(storage, "B8:S16")
    assert_formula_range(risk, "B2:B7")
    assert_formula_range(risk, "B8:J16")


def test_recalculated_values_match_expected_semantics():
    wb = load_workbook(data_only=True)
    expected = compute_expected_values(wb)

    for sheet_name, cell_map in expected.items():
        ws = wb[sheet_name]
        for cell_ref, expected_value in cell_map.items():
            approx_equal(ws[cell_ref].value, expected_value, tol=1e-9)
