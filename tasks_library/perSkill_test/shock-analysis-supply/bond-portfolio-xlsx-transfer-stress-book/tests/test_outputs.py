from datetime import datetime
from math import isclose
from pathlib import Path

import openpyxl


OUTPUT = Path("/root/bond-stress-book.xlsx")
REQUIRED_SHEETS = [
    "Control",
    "Holdings",
    "Coupon_Calendar",
    "Curves",
    "Recovery_Assumptions",
    "Cashflow_Model",
    "Scenario_Valuation",
    "Portfolio_Summary",
]


def load_workbook(data_only=False):
    assert OUTPUT.exists(), f"Missing output workbook: {OUTPUT}"
    return openpyxl.load_workbook(OUTPUT, data_only=data_only)


def formula_text(cell):
    value = cell.value
    return value if isinstance(value, str) and value.startswith("=") else None


def assert_formula_present(ws, cell_ref):
    assert formula_text(ws[cell_ref]) is not None, f"Expected formula in {ws.title}!{cell_ref}"


def assert_formula_mentions(ws, cell_ref, *tokens):
    formula = formula_text(ws[cell_ref])
    assert formula is not None, f"Expected formula in {ws.title}!{cell_ref}"
    for token in tokens:
        assert token in formula, f"Expected {ws.title}!{cell_ref} formula to mention {token!r}: {formula}"


def to_datetime(value):
    if isinstance(value, datetime):
        return value
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, str):
        return datetime.strptime(value, "%Y-%m-%d")
    raise TypeError(f"Unsupported date value: {value!r}")


def read_inputs(wb):
    control = wb["Control"]
    holdings_ws = wb["Holdings"]
    calendar_ws = wb["Coupon_Calendar"]
    curves_ws = wb["Curves"]
    recovery_ws = wb["Recovery_Assumptions"]

    valuation_date = to_datetime(control["B2"].value)
    scenarios = {
        "Parallel_Up_75": {
            "curve_shift_bps": float(control["B9"].value),
            "spread_shift_bps": float(control["C9"].value),
            "recovery_multiplier": float(control["D9"].value),
        },
        "Parallel_Down_50": {
            "curve_shift_bps": float(control["B10"].value),
            "spread_shift_bps": float(control["C10"].value),
            "recovery_multiplier": float(control["D10"].value),
        },
        "Credit_Stress": {
            "curve_shift_bps": float(control["B11"].value),
            "spread_shift_bps": float(control["C11"].value),
            "recovery_multiplier": float(control["D11"].value),
        },
    }

    holdings = {}
    for row in range(2, 6):
        holdings[holdings_ws[f"A{row}"].value] = {
            "issuer": holdings_ws[f"B{row}"].value,
            "rating": holdings_ws[f"C{row}"].value,
            "coupon_rate": float(holdings_ws[f"D{row}"].value),
            "position_face": float(holdings_ws[f"E{row}"].value),
            "coupon_frequency": float(holdings_ws[f"F{row}"].value),
            "curve_tenor": int(holdings_ws[f"G{row}"].value),
            "spread_bps": float(holdings_ws[f"H{row}"].value),
            "stress_default_flag": int(holdings_ws[f"I{row}"].value),
        }

    curves = {}
    for row in range(2, 8):
        curves[int(curves_ws[f"A{row}"].value)] = float(curves_ws[f"B{row}"].value)

    recoveries = {}
    for row in range(2, 6):
        recoveries[recovery_ws[f"A{row}"].value] = float(recovery_ws[f"B{row}"].value)

    calendar = []
    row = 2
    while calendar_ws[f"A{row}"].value:
        calendar.append(
            {
                "bond_id": calendar_ws[f"A{row}"].value,
                "payment_date": to_datetime(calendar_ws[f"B{row}"].value),
                "is_maturity": int(calendar_ws[f"C{row}"].value),
                "recovery_anchor": int(calendar_ws[f"D{row}"].value),
            }
        )
        row += 1

    return valuation_date, scenarios, holdings, curves, recoveries, calendar


def compute_expected(wb):
    valuation_date, scenarios, holdings, curves, recoveries, calendar = read_inputs(wb)

    cashflows = []
    for item in calendar:
        holding = holdings[item["bond_id"]]
        base_curve = curves[holding["curve_tenor"]]
        base_yield = base_curve + holding["spread_bps"] / 10000
        year_fraction = max((item["payment_date"] - valuation_date).days / 365, 0)
        coupon_cf = (
            holding["position_face"] * holding["coupon_rate"] / holding["coupon_frequency"]
            if year_fraction > 0
            else 0.0
        )
        principal_cf = holding["position_face"] if year_fraction > 0 and item["is_maturity"] == 1 else 0.0
        total_cf = coupon_cf + principal_cf
        base_pv = total_cf / ((1 + base_yield) ** year_fraction)

        up_yield = (
            base_yield
            + scenarios["Parallel_Up_75"]["curve_shift_bps"] / 10000
            + scenarios["Parallel_Up_75"]["spread_shift_bps"] / 10000
        )
        down_yield = (
            base_yield
            + scenarios["Parallel_Down_50"]["curve_shift_bps"] / 10000
            + scenarios["Parallel_Down_50"]["spread_shift_bps"] / 10000
        )
        credit_yield = (
            base_yield
            + scenarios["Credit_Stress"]["curve_shift_bps"] / 10000
            + scenarios["Credit_Stress"]["spread_shift_bps"] / 10000
        )

        recovery_rate = recoveries[holding["rating"]]
        if holding["stress_default_flag"] == 1:
            credit_pv = (
                holding["position_face"]
                * recovery_rate
                * scenarios["Credit_Stress"]["recovery_multiplier"]
                / ((1 + credit_yield) ** year_fraction)
                if item["recovery_anchor"] == 1 and year_fraction > 0
                else 0.0
            )
        else:
            credit_pv = total_cf / ((1 + credit_yield) ** year_fraction)

        cashflows.append(
            {
                "bond_id": item["bond_id"],
                "base_yield": base_yield,
                "year_fraction": year_fraction,
                "base_pv": base_pv,
                "pv_time": base_pv * year_fraction,
                "pv_time_sq": base_pv * year_fraction * year_fraction,
                "up_pv": total_cf / ((1 + up_yield) ** year_fraction),
                "down_pv": total_cf / ((1 + down_yield) ** year_fraction),
                "credit_pv": credit_pv,
            }
        )

    scenario_rows = {}
    for bond_id, holding in holdings.items():
        bond_rows = [row for row in cashflows if row["bond_id"] == bond_id]
        base_yield = bond_rows[0]["base_yield"]
        base_price = sum(row["base_pv"] for row in bond_rows)
        macaulay = sum(row["pv_time"] for row in bond_rows) / base_price
        modified = macaulay / (1 + base_yield)
        convexity = sum(row["pv_time_sq"] for row in bond_rows) / base_price
        price_up = sum(row["up_pv"] for row in bond_rows)
        price_down = sum(row["down_pv"] for row in bond_rows)
        price_credit = sum(row["credit_pv"] for row in bond_rows)
        scenario_rows[bond_id] = {
            "issuer": holding["issuer"],
            "rating": holding["rating"],
            "base_yield": base_yield,
            "base_price": base_price,
            "macaulay": macaulay,
            "modified": modified,
            "convexity": convexity,
            "price_up": price_up,
            "pnl_up": price_up - base_price,
            "price_down": price_down,
            "pnl_down": price_down - base_price,
            "price_credit": price_credit,
            "pnl_credit": price_credit - base_price,
        }

    total_base = sum(row["base_price"] for row in scenario_rows.values())
    weighted_modified = sum(
        row["base_price"] * row["modified"] for row in scenario_rows.values()
    ) / total_base
    weighted_convexity = sum(
        row["base_price"] * row["convexity"] for row in scenario_rows.values()
    ) / total_base

    summary = {
        "total_base": total_base,
        "weighted_modified": weighted_modified,
        "weighted_convexity": weighted_convexity,
        "scenario_totals": {
            "Parallel_Up_75": sum(row["price_up"] for row in scenario_rows.values()),
            "Parallel_Down_50": sum(row["price_down"] for row in scenario_rows.values()),
            "Credit_Stress": sum(row["price_credit"] for row in scenario_rows.values()),
        },
    }
    summary["scenario_pnl"] = {
        key: value - total_base for key, value in summary["scenario_totals"].items()
    }

    return scenario_rows, summary


def test_required_sheets_and_imported_inputs_exist():
    wb = load_workbook()
    assert wb.sheetnames == REQUIRED_SHEETS

    holdings = wb["Holdings"]
    calendar = wb["Coupon_Calendar"]
    curves = wb["Curves"]
    recovery = wb["Recovery_Assumptions"]

    assert holdings.max_row == 5
    assert calendar.max_row == 31
    assert curves.max_row == 7
    assert recovery.max_row == 5


def test_cashflow_model_formulas_are_present():
    wb = load_workbook()
    ws = wb["Cashflow_Model"]

    assert_formula_mentions(ws, "A2", "Coupon_Calendar!")
    assert_formula_mentions(ws, "E2", "Holdings!")
    assert_formula_mentions(ws, "K2", "Curves!")
    assert_formula_mentions(ws, "M2", "Control!")
    assert_formula_present(ws, "N2")
    assert_formula_present(ws, "O2")
    assert_formula_present(ws, "Q2")
    assert_formula_mentions(ws, "T2", "Control!")
    assert_formula_mentions(ws, "W2", "Recovery_Assumptions!")
    assert_formula_present(ws, "V2")


def test_scenario_and_summary_formulas_are_present():
    wb = load_workbook()
    scenario = wb["Scenario_Valuation"]
    summary = wb["Portfolio_Summary"]

    assert_formula_mentions(scenario, "A2", "Holdings!")
    assert_formula_mentions(scenario, "E2", "Cashflow_Model!")
    assert_formula_mentions(scenario, "F2", "Cashflow_Model!")
    assert_formula_present(scenario, "G2")
    assert_formula_mentions(scenario, "M2", "Cashflow_Model!")

    assert_formula_mentions(summary, "B2", "Scenario_Valuation!")
    assert_formula_mentions(summary, "B3", "Scenario_Valuation!", "B2")
    assert_formula_mentions(summary, "B4", "Scenario_Valuation!", "B2")
    assert_formula_mentions(summary, "B7", "Scenario_Valuation!")
    assert_formula_present(summary, "C9")


def test_recalculated_values_match_expected_results():
    wb_formula = load_workbook()
    wb_values = load_workbook(data_only=True)
    expected_rows, expected_summary = compute_expected(wb_formula)

    scenario = wb_values["Scenario_Valuation"]
    summary = wb_values["Portfolio_Summary"]

    row_map = {"ALPHA28": 2, "BRAVO30": 3, "COBALT29": 4, "DELTA31": 5}

    for bond_id, row_idx in row_map.items():
        expected = expected_rows[bond_id]
        assert isclose(float(scenario[f"D{row_idx}"].value), expected["base_yield"], rel_tol=0, abs_tol=1e-9)
        assert isclose(float(scenario[f"E{row_idx}"].value), expected["base_price"], rel_tol=0, abs_tol=1e-6)
        assert isclose(float(scenario[f"F{row_idx}"].value), expected["macaulay"], rel_tol=0, abs_tol=1e-9)
        assert isclose(float(scenario[f"G{row_idx}"].value), expected["modified"], rel_tol=0, abs_tol=1e-9)
        assert isclose(float(scenario[f"H{row_idx}"].value), expected["convexity"], rel_tol=0, abs_tol=1e-9)
        assert isclose(float(scenario[f"I{row_idx}"].value), expected["price_up"], rel_tol=0, abs_tol=1e-6)
        assert isclose(float(scenario[f"J{row_idx}"].value), expected["pnl_up"], rel_tol=0, abs_tol=1e-6)
        assert isclose(float(scenario[f"K{row_idx}"].value), expected["price_down"], rel_tol=0, abs_tol=1e-6)
        assert isclose(float(scenario[f"L{row_idx}"].value), expected["pnl_down"], rel_tol=0, abs_tol=1e-6)
        assert isclose(float(scenario[f"M{row_idx}"].value), expected["price_credit"], rel_tol=0, abs_tol=1e-6)
        assert isclose(float(scenario[f"N{row_idx}"].value), expected["pnl_credit"], rel_tol=0, abs_tol=1e-6)

    assert isclose(float(summary["B2"].value), expected_summary["total_base"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(summary["B3"].value), expected_summary["weighted_modified"], rel_tol=0, abs_tol=1e-9)
    assert isclose(float(summary["B4"].value), expected_summary["weighted_convexity"], rel_tol=0, abs_tol=1e-9)
    assert isclose(float(summary["B7"].value), expected_summary["scenario_totals"]["Parallel_Up_75"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(summary["C7"].value), expected_summary["scenario_pnl"]["Parallel_Up_75"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(summary["B8"].value), expected_summary["scenario_totals"]["Parallel_Down_50"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(summary["C8"].value), expected_summary["scenario_pnl"]["Parallel_Down_50"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(summary["B9"].value), expected_summary["scenario_totals"]["Credit_Stress"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(summary["C9"].value), expected_summary["scenario_pnl"]["Credit_Stress"], rel_tol=0, abs_tol=1e-6)
