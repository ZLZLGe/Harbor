from math import isclose
from pathlib import Path

import openpyxl


OUTPUT = Path("/root/reservoir-resilience-model.xlsx")
REQUIRED_SHEETS = [
    "Hydrology_Input",
    "Demand_Input",
    "Policy",
    "Monthly_Balance",
    "Scenario_Summary",
]


def load_workbook(data_only=False):
    assert OUTPUT.exists(), f"Missing output workbook: {OUTPUT}"
    return openpyxl.load_workbook(OUTPUT, data_only=data_only)


def formula_text(cell):
    value = cell.value
    return value if isinstance(value, str) and value.startswith("=") else None


def read_policy(wb):
    ws = wb["Policy"]
    values = {}
    row = 2
    while ws[f"A{row}"].value:
        values[ws[f"A{row}"].value] = float(ws[f"B{row}"].value)
        row += 1
    return values


def compute_expected(wb):
    hydro = wb["Hydrology_Input"]
    demand = wb["Demand_Input"]
    policy = read_policy(wb)

    start_storage = policy["Initial_Storage_MCM"]
    quarters = ["Q1"] * 3 + ["Q2"] * 3 + ["Q3"] * 3 + ["Q4"] * 3
    monthly = []

    for idx, row in enumerate(range(2, 14)):
        month = hydro[f"A{row}"].value
        inflow = float(hydro[f"B{row}"].value)
        evap = float(hydro[f"C{row}"].value)
        urban = float(demand[f"B{row}"].value)
        agriculture = float(demand[f"C{row}"].value)
        eco = float(demand[f"D{row}"].value)

        if start_storage < policy["Emergency_Threshold_MCM"]:
            trigger = "Emergency"
            urban_factor = policy["Urban_Emergency_Factor"]
            ag_factor = policy["Agriculture_Emergency_Factor"]
        elif start_storage < policy["Watch_Threshold_MCM"]:
            trigger = "Watch"
            urban_factor = policy["Urban_Watch_Factor"]
            ag_factor = policy["Agriculture_Watch_Factor"]
        else:
            trigger = "Normal"
            urban_factor = 1.0
            ag_factor = 1.0

        urban_target = urban * urban_factor
        ag_target = agriculture * ag_factor
        planned_release = eco + urban_target + ag_target
        base_usable = max(start_storage + inflow - evap - policy["Dead_Storage_MCM"], 0.0)
        max_extra = max(
            max(start_storage + inflow - evap - policy["Emergency_Floor_MCM"], 0.0) - base_usable,
            0.0,
        )

        if trigger == "Emergency":
            emergency_used = min(
                policy["Emergency_Buffer_Max_MCM"],
                max(0.0, planned_release - base_usable),
                max_extra,
            )
        else:
            emergency_used = 0.0

        feasible_release = min(planned_release, base_usable + emergency_used)
        eco_actual = min(eco, feasible_release)
        urban_actual = min(urban_target, max(feasible_release - eco_actual, 0.0))
        ag_actual = min(ag_target, max(feasible_release - eco_actual - urban_actual, 0.0))
        urban_shortage = urban - urban_actual
        ag_shortage = agriculture - ag_actual
        eco_shortage = eco - eco_actual
        total_shortage = urban_shortage + ag_shortage + eco_shortage
        end_storage = min(
            policy["Max_Storage_MCM"],
            start_storage + inflow - evap - eco_actual - urban_actual - ag_actual,
        )

        monthly.append(
            {
                "month": month,
                "quarter": quarters[idx],
                "inflow": inflow,
                "evap": evap,
                "urban_demand": urban,
                "ag_demand": agriculture,
                "eco_demand": eco,
                "start_storage": start_storage,
                "trigger": trigger,
                "urban_target": urban_target,
                "ag_target": ag_target,
                "emergency_used": emergency_used,
                "planned_release": planned_release,
                "base_usable": base_usable,
                "feasible_release": feasible_release,
                "eco_actual": eco_actual,
                "urban_actual": urban_actual,
                "ag_actual": ag_actual,
                "urban_shortage": urban_shortage,
                "ag_shortage": ag_shortage,
                "eco_shortage": eco_shortage,
                "total_shortage": total_shortage,
                "end_storage": end_storage,
            }
        )
        start_storage = end_storage

    quarter_shortage = {key: 0.0 for key in ("Q1", "Q2", "Q3", "Q4")}
    for item in monthly:
        quarter_shortage[item["quarter"]] += item["total_shortage"]

    summary = {
        "total_inflow": sum(item["inflow"] for item in monthly),
        "total_urban_demand": sum(item["urban_demand"] for item in monthly),
        "total_ag_demand": sum(item["ag_demand"] for item in monthly),
        "total_eco_demand": sum(item["eco_demand"] for item in monthly),
        "total_urban_actual": sum(item["urban_actual"] for item in monthly),
        "total_ag_actual": sum(item["ag_actual"] for item in monthly),
        "total_eco_actual": sum(item["eco_actual"] for item in monthly),
        "total_shortage": sum(item["total_shortage"] for item in monthly),
        "watch_months": sum(1 for item in monthly if item["trigger"] == "Watch"),
        "emergency_months": sum(1 for item in monthly if item["trigger"] == "Emergency"),
        "buffer_months": sum(1 for item in monthly if item["emergency_used"] > 0),
        "min_end_storage": min(item["end_storage"] for item in monthly),
        "end_storage": monthly[-1]["end_storage"],
    }
    summary["urban_service_ratio"] = summary["total_urban_actual"] / summary["total_urban_demand"]
    summary["ag_service_ratio"] = summary["total_ag_actual"] / summary["total_ag_demand"]
    summary["eco_compliance_ratio"] = summary["total_eco_actual"] / summary["total_eco_demand"]
    summary["end_gap"] = summary["end_storage"] - policy["Target_End_Storage_MCM"]
    summary["q3_share"] = 0.0 if summary["total_shortage"] == 0 else quarter_shortage["Q3"] / summary["total_shortage"]
    summary["resilience_score"] = (
        0.45 * summary["urban_service_ratio"]
        + 0.25 * summary["ag_service_ratio"]
        + 0.15 * summary["eco_compliance_ratio"]
        + 0.15 * (1 - summary["buffer_months"] / 12)
    )
    if summary["resilience_score"] >= 0.82:
        summary["scenario"] = "Stable"
    elif summary["resilience_score"] >= 0.60:
        summary["scenario"] = "Managed Stress"
    else:
        summary["scenario"] = "Severe Stress"

    return monthly, quarter_shortage, summary


def test_required_sheets_and_imported_inputs_exist():
    wb = load_workbook()
    assert wb.sheetnames == REQUIRED_SHEETS

    hydro = wb["Hydrology_Input"]
    demand = wb["Demand_Input"]

    assert hydro.max_row == 13
    assert demand.max_row == 13
    assert hydro["A2"].value == "Jan"
    assert hydro["B13"].value == 22
    assert demand["D8"].value == 6
    assert demand["C13"].value == 6


def test_monthly_balance_uses_recursive_cross_sheet_formulas():
    wb = load_workbook()
    ws = wb["Monthly_Balance"]

    assert "Hydrology_Input!" in formula_text(ws["B2"])
    assert "Demand_Input!" in formula_text(ws["D2"])
    assert "Demand_Input!" in formula_text(ws["F2"])
    assert "Policy!" in formula_text(ws["G2"])
    assert "V2" in formula_text(ws["G3"])
    assert "IF(" in formula_text(ws["H2"]).upper()
    assert "Policy!" in formula_text(ws["I2"])
    assert "Policy!" in formula_text(ws["J2"])
    assert "MIN(" in formula_text(ws["K2"]).upper()
    assert "MAX(" in formula_text(ws["M2"]).upper()
    assert formula_text(ws["N2"]) is not None
    assert formula_text(ws["U2"]) is not None
    assert "Policy!" in formula_text(ws["V2"])


def test_summary_formulas_are_present():
    wb = load_workbook()
    ws = wb["Scenario_Summary"]

    assert "SUM(" in formula_text(ws["B2"]).upper()
    assert "IF(" in formula_text(ws["B9"]).upper()
    assert "COUNTIF" in formula_text(ws["B13"]).upper()
    assert "COUNTIF" in formula_text(ws["B15"]).upper()
    assert "IF(" in formula_text(ws["B19"]).upper()
    assert "0.45" in formula_text(ws["B20"])
    assert "IF(" in formula_text(ws["B21"]).upper()
    assert formula_text(ws["E2"]) is not None
    assert formula_text(ws["E4"]) is not None


def test_monthly_balance_values_match_expected():
    wb_formula = load_workbook()
    wb_values = load_workbook(data_only=True)
    expected_rows, _, _ = compute_expected(wb_formula)

    ws = wb_values["Monthly_Balance"]

    for offset, expected in enumerate(expected_rows, start=2):
        assert ws[f"A{offset}"].value == expected["month"]
        assert ws[f"H{offset}"].value == expected["trigger"]

        numeric_cells = {
            "B": expected["inflow"],
            "C": expected["evap"],
            "G": expected["start_storage"],
            "I": expected["urban_target"],
            "J": expected["ag_target"],
            "K": expected["emergency_used"],
            "N": expected["feasible_release"],
            "O": expected["eco_actual"],
            "P": expected["urban_actual"],
            "Q": expected["ag_actual"],
            "R": expected["urban_shortage"],
            "S": expected["ag_shortage"],
            "T": expected["eco_shortage"],
            "U": expected["total_shortage"],
            "V": expected["end_storage"],
        }

        for col, target in numeric_cells.items():
            actual = float(ws[f"{col}{offset}"].value)
            assert isclose(actual, target, rel_tol=0, abs_tol=1e-9), (
                f"Mismatch at {col}{offset}: expected {target}, got {actual}"
            )


def test_summary_values_match_expected():
    wb_formula = load_workbook()
    wb_values = load_workbook(data_only=True)
    _, quarter_shortage, summary = compute_expected(wb_formula)

    ws = wb_values["Scenario_Summary"]

    numeric_expectations = {
        "B2": summary["total_inflow"],
        "B3": summary["total_urban_demand"],
        "B4": summary["total_ag_demand"],
        "B5": summary["total_eco_demand"],
        "B6": summary["total_urban_actual"],
        "B7": summary["total_ag_actual"],
        "B8": summary["total_eco_actual"],
        "B9": summary["urban_service_ratio"],
        "B10": summary["ag_service_ratio"],
        "B11": summary["eco_compliance_ratio"],
        "B12": summary["total_shortage"],
        "B13": summary["watch_months"],
        "B14": summary["emergency_months"],
        "B15": summary["buffer_months"],
        "B16": summary["min_end_storage"],
        "B17": summary["end_storage"],
        "B18": summary["end_gap"],
        "B19": summary["q3_share"],
        "B20": summary["resilience_score"],
        "E2": quarter_shortage["Q1"],
        "E3": quarter_shortage["Q2"],
        "E4": quarter_shortage["Q3"],
        "E5": quarter_shortage["Q4"],
    }

    for cell, target in numeric_expectations.items():
        actual = float(ws[cell].value)
        assert isclose(actual, target, rel_tol=0, abs_tol=1e-9), (
            f"Mismatch at {cell}: expected {target}, got {actual}"
        )

    assert ws["B21"].value == summary["scenario"]
