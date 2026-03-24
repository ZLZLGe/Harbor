from decimal import Decimal, ROUND_HALF_UP
import re
from pathlib import Path

import openpyxl
from openpyxl.utils.cell import get_column_letter, range_boundaries
from openpyxl.worksheet.formula import ArrayFormula


OUTPUT_FILE = Path("warehouse-buffer-stock.xlsx")
REQUIRED_SHEETS = [
    "Shipment_Forecast",
    "Capacity_Loss",
    "Replenishment_Plan",
    "Inventory_Rollforward",
    "Service_Gap",
    "Expansion_Comparison",
]
FORMULA_RANGES = {
    "Shipment_Forecast": ["C8:C16", "E8:E16"],
    "Capacity_Loss": ["C8:F16"],
    "Replenishment_Plan": ["D8:D16"],
    "Inventory_Rollforward": ["B2:B4", "B8:L16"],
    "Service_Gap": ["B2:B5", "B8:J16"],
    "Expansion_Comparison": ["B5:B6", "B8:J16"],
}


def load_workbook(data_only=False):
    assert OUTPUT_FILE.exists(), f"Output file not found: {OUTPUT_FILE}"
    return openpyxl.load_workbook(OUTPUT_FILE, data_only=data_only)


def get_formula_text(value):
    if value is None:
        return None
    if isinstance(value, ArrayFormula):
        return value.text
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def approx_equal(actual, expected, tol=1e-6):
    assert actual is not None, f"Expected {expected}, got None"
    assert abs(actual - expected) <= tol, f"Expected {expected}, got {actual}"


def excel_round(value, digits=0):
    quantum = Decimal("1").scaleb(-digits)
    return float(Decimal(str(value)).quantize(quantum, rounding=ROUND_HALF_UP))


def strip_wrapping_parentheses(text):
    while text.startswith("(") and text.endswith(")"):
        depth = 0
        wraps_entire_expression = True
        for index, char in enumerate(text):
            if char == "(":
                depth += 1
            elif char == ")":
                depth -= 1
            if depth == 0 and index != len(text) - 1:
                wraps_entire_expression = False
                break
        if not wraps_entire_expression:
            break
        text = text[1:-1]
    return text


def normalize_formula(formula_text):
    assert formula_text is not None, "Expected a formula cell"
    text = formula_text.strip()
    assert text.startswith("="), f"Expected formula text, got {formula_text}"
    text = re.sub(r"\s+", "", text[1:])
    text = text.replace("$", "").replace("'", "")
    text = strip_wrapping_parentheses(text)
    if text.startswith("+"):
        text = text[1:]
    return text.upper()


def assert_direct_reference(cell, expected_reference):
    formula_text = get_formula_text(cell.value)
    normalized = normalize_formula(formula_text)
    expected = expected_reference.replace("$", "").replace("'", "").upper()
    assert normalized == expected, f"{cell.coordinate} should directly reference {expected_reference}, got {formula_text}"


def assert_formula_range(ws, cell_range):
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            coord = f"{get_column_letter(col)}{row}"
            assert get_formula_text(ws[coord].value) is not None, f"{ws.title}!{coord} should contain a formula"


def build_expected_outputs(wb):
    forecast = wb["Shipment_Forecast"]
    capacity = wb["Capacity_Loss"]
    replenishment = wb["Replenishment_Plan"]
    comparison = wb["Expansion_Comparison"]

    growth = forecast["B2"].value
    stress_uplift = forecast["B3"].value
    base_demand = []
    stress_demand = []
    seasonality = [forecast[f"D{row}"].value or 0 for row in range(8, 17)]
    actual_shipments = [forecast[f"B{row}"].value for row in range(8, 11)]

    for index, row in enumerate(range(8, 17)):
        if row <= 10:
            base_value = actual_shipments[index]
            stress_value = base_value
        else:
            base_value = base_demand[-1] * (1 + growth)
            stress_value = base_value * (1 + stress_uplift + seasonality[index])
        base_demand.append(base_value)
        stress_demand.append(stress_value)

    monthly_loss = [capacity[f"B{row}"].value for row in range(8, 17)]
    nominal_capacity = capacity["B2"].value
    lease_capacity_add = capacity["B3"].value
    rack_upgrade_add = capacity["B4"].value
    loss_reduction = capacity["B5"].value

    baseline_capacity = [nominal_capacity * (1 - loss_rate) for loss_rate in monthly_loss]
    leased_capacity = [(nominal_capacity + lease_capacity_add) * (1 - loss_rate) for loss_rate in monthly_loss]
    upgraded_loss_rate = [max(loss_rate - loss_reduction, 0) for loss_rate in monthly_loss]
    upgraded_capacity = [(nominal_capacity + rack_upgrade_add) * (1 - loss_rate) for loss_rate in upgraded_loss_rate]

    planned_receipts = [replenishment[f"B{row}"].value for row in range(8, 17)]
    supplier_fill = [replenishment[f"C{row}"].value for row in range(8, 17)]
    lease_available = [replenishment[f"E{row}"].value for row in range(8, 17)]
    upgrade_available = [replenishment[f"F{row}"].value for row in range(8, 17)]
    net_receipts = [planned * fill_rate for planned, fill_rate in zip(planned_receipts, supplier_fill)]

    target_service = replenishment["B3"].value
    target_buffer_days = replenishment["B4"].value
    opening_inventory = []
    gross_inventory = []
    loss_units = []
    usable_inventory = []
    fulfilled_shipments = []
    ending_inventory = []
    service_level = []
    utilization = []
    cover_days = []

    for index in range(9):
        if index == 0:
            opening_value = replenishment["B2"].value
        else:
            opening_value = ending_inventory[-1]
        opening_inventory.append(opening_value)

        gross_value = opening_value + net_receipts[index]
        gross_inventory.append(gross_value)

        loss_value = excel_round(gross_value * monthly_loss[index], 0)
        loss_units.append(loss_value)

        usable_value = gross_value - loss_value
        usable_inventory.append(usable_value)

        fulfilled_value = min(base_demand[index], usable_value)
        fulfilled_shipments.append(fulfilled_value)

        ending_value = min(baseline_capacity[index], usable_value - fulfilled_value)
        ending_inventory.append(ending_value)

        service_value = fulfilled_value / base_demand[index]
        service_level.append(service_value)

        utilization_value = ending_value / baseline_capacity[index]
        utilization.append(utilization_value)

        cover_value = 0 if base_demand[index] == 0 else ending_value / (base_demand[index] / 30)
        cover_days.append(cover_value)

    stress_unmet = [max(demand - shipment, 0) for demand, shipment in zip(stress_demand, fulfilled_shipments)]
    stress_service_level = [shipment / demand for demand, shipment in zip(stress_demand, fulfilled_shipments)]
    service_gap = [max(target_service - level, 0) for level in stress_service_level]
    target_buffer_inventory = [demand / 30 * target_buffer_days for demand in stress_demand]
    buffer_gap = [max(target - actual, 0) for target, actual in zip(target_buffer_inventory, ending_inventory)]

    lease_coverage = [
        min(gap, available + lease_capacity_add * (1 - loss_rate))
        for gap, available, loss_rate in zip(buffer_gap, lease_available, monthly_loss)
    ]
    lease_remaining_gap = [max(gap - covered, 0) for gap, covered in zip(buffer_gap, lease_coverage)]
    upgrade_coverage = [
        min(gap, available + rack_upgrade_add * (1 - loss_rate))
        for gap, available, loss_rate in zip(buffer_gap, upgrade_available, upgraded_loss_rate)
    ]
    upgrade_remaining_gap = [max(gap - covered, 0) for gap, covered in zip(buffer_gap, upgrade_coverage)]
    lease_cost = [covered * comparison["B2"].value for covered in lease_coverage]
    upgrade_cost = [covered * comparison["B3"].value for covered in upgrade_coverage]
    lease_unit_cost = [0 if covered == 0 else cost / covered for covered, cost in zip(lease_coverage, lease_cost)]
    upgrade_unit_cost = [0 if covered == 0 else cost / covered for covered, cost in zip(upgrade_coverage, upgrade_cost)]

    return {
        "Shipment_Forecast": {
            "C": base_demand,
            "E": stress_demand,
        },
        "Capacity_Loss": {
            "C": baseline_capacity,
            "D": leased_capacity,
            "E": upgraded_loss_rate,
            "F": upgraded_capacity,
        },
        "Replenishment_Plan": {
            "D": net_receipts,
        },
        "Inventory_Rollforward": {
            "B2": target_service,
            "B3": sum(ending_inventory) / len(ending_inventory),
            "B4": sum(1 for value in service_level if value < target_service),
            "B": opening_inventory,
            "C": net_receipts,
            "D": gross_inventory,
            "E": loss_units,
            "F": usable_inventory,
            "G": base_demand,
            "H": fulfilled_shipments,
            "I": ending_inventory,
            "J": service_level,
            "K": utilization,
            "L": cover_days,
        },
        "Service_Gap": {
            "B2": target_buffer_days,
            "B3": sum(stress_unmet),
            "B4": max(service_gap),
            "B5": sum(buffer_gap),
            "B": stress_demand,
            "C": fulfilled_shipments,
            "D": stress_unmet,
            "E": stress_service_level,
            "F": [target_service] * 9,
            "G": service_gap,
            "H": target_buffer_inventory,
            "I": ending_inventory,
            "J": buffer_gap,
        },
        "Expansion_Comparison": {
            "B5": sum(lease_coverage),
            "B6": sum(upgrade_coverage),
            "B": buffer_gap,
            "C": lease_coverage,
            "D": lease_remaining_gap,
            "E": upgrade_coverage,
            "F": upgrade_remaining_gap,
            "G": lease_cost,
            "H": upgrade_cost,
            "I": lease_unit_cost,
            "J": upgrade_unit_cost,
        },
    }


def test_required_sheets_exist():
    wb = load_workbook()
    assert wb.sheetnames == REQUIRED_SHEETS


def test_inputs_and_month_axis_are_preserved():
    wb = load_workbook()

    forecast = wb["Shipment_Forecast"]
    capacity = wb["Capacity_Loss"]
    replenishment = wb["Replenishment_Plan"]
    comparison = wb["Expansion_Comparison"]

    assert forecast["B2"].value == 0.025
    assert forecast["B3"].value == 0.06
    assert capacity["B2"].value == 11800
    assert capacity["B3"].value == 1800
    assert capacity["B4"].value == 2600
    assert capacity["B5"].value == 0.004
    assert replenishment["B2"].value == 10800
    assert replenishment["B3"].value == 0.97
    assert replenishment["B4"].value == 18
    assert comparison["B2"].value == 22
    assert comparison["B3"].value == 46

    expected_months = [
        "2025-01",
        "2025-02",
        "2025-03",
        "2025-04",
        "2025-05",
        "2025-06",
        "2025-07",
        "2025-08",
        "2025-09",
    ]
    assert [forecast[f"A{row}"].value for row in range(8, 17)] == expected_months


def test_required_formula_regions_use_formulas():
    wb = load_workbook()
    for sheet_name, ranges in FORMULA_RANGES.items():
        ws = wb[sheet_name]
        for cell_range in ranges:
            assert_formula_range(ws, cell_range)


def test_direct_links_and_rollforward_chain_are_present():
    wb = load_workbook()

    forecast = wb["Shipment_Forecast"]
    rollforward = wb["Inventory_Rollforward"]
    gap = wb["Service_Gap"]
    comparison = wb["Expansion_Comparison"]

    for row in range(8, 11):
        assert_direct_reference(forecast[f"C{row}"], f"B{row}")
        assert_direct_reference(forecast[f"E{row}"], f"C{row}")

    assert_direct_reference(rollforward["B2"], "Replenishment_Plan!B3")
    assert_direct_reference(rollforward["B8"], "Replenishment_Plan!B2")
    for row in range(9, 17):
        assert_direct_reference(rollforward[f"B{row}"], f"I{row - 1}")
    for row in range(8, 17):
        assert_direct_reference(rollforward[f"C{row}"], f"Replenishment_Plan!D{row}")
        assert_direct_reference(rollforward[f"G{row}"], f"Shipment_Forecast!C{row}")
        assert_direct_reference(gap[f"B{row}"], f"Shipment_Forecast!E{row}")
        assert_direct_reference(gap[f"C{row}"], f"Inventory_Rollforward!H{row}")
        assert_direct_reference(gap[f"F{row}"], "Replenishment_Plan!B3")
        assert_direct_reference(gap[f"I{row}"], f"Inventory_Rollforward!I{row}")
        assert_direct_reference(comparison[f"B{row}"], f"Service_Gap!J{row}")

    assert_direct_reference(gap["B2"], "Replenishment_Plan!B4")


def test_recalculated_outputs_match_expected_model():
    wb = load_workbook(data_only=True)
    expected = build_expected_outputs(wb)

    forecast = wb["Shipment_Forecast"]
    for row, expected_value in zip(range(8, 17), expected["Shipment_Forecast"]["C"]):
        approx_equal(forecast[f"C{row}"].value, expected_value)
    for row, expected_value in zip(range(8, 17), expected["Shipment_Forecast"]["E"]):
        approx_equal(forecast[f"E{row}"].value, expected_value)

    capacity = wb["Capacity_Loss"]
    for column in ["C", "D", "E", "F"]:
        for row, expected_value in zip(range(8, 17), expected["Capacity_Loss"][column]):
            approx_equal(capacity[f"{column}{row}"].value, expected_value)

    replenishment = wb["Replenishment_Plan"]
    for row, expected_value in zip(range(8, 17), expected["Replenishment_Plan"]["D"]):
        approx_equal(replenishment[f"D{row}"].value, expected_value)

    rollforward = wb["Inventory_Rollforward"]
    approx_equal(rollforward["B2"].value, expected["Inventory_Rollforward"]["B2"])
    approx_equal(rollforward["B3"].value, expected["Inventory_Rollforward"]["B3"], tol=1e-9)
    approx_equal(rollforward["B4"].value, expected["Inventory_Rollforward"]["B4"])
    for column in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
        for row, expected_value in zip(range(8, 17), expected["Inventory_Rollforward"][column]):
            approx_equal(rollforward[f"{column}{row}"].value, expected_value, tol=1e-9)

    gap = wb["Service_Gap"]
    for summary_cell in ["B2", "B3", "B4", "B5"]:
        approx_equal(gap[summary_cell].value, expected["Service_Gap"][summary_cell], tol=1e-9)
    for column in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        for row, expected_value in zip(range(8, 17), expected["Service_Gap"][column]):
            approx_equal(gap[f"{column}{row}"].value, expected_value, tol=1e-9)

    comparison = wb["Expansion_Comparison"]
    for summary_cell in ["B5", "B6"]:
        approx_equal(comparison[summary_cell].value, expected["Expansion_Comparison"][summary_cell], tol=1e-9)
    for column in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        for row, expected_value in zip(range(8, 17), expected["Expansion_Comparison"][column]):
            approx_equal(comparison[f"{column}{row}"].value, expected_value, tol=1e-9)
