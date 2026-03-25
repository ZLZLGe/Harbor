from math import ceil, isclose
from pathlib import Path

import openpyxl


OUTPUT = Path("/root/coldchain-ops-plan.xlsx")
REQUIRED_SHEETS = [
    "Planner",
    "Store_Demand",
    "Shift_Definitions",
    "Power_Curve",
    "Tariff_Template",
    "Hourly_Load",
    "Labor_Schedule",
    "Cost_Summary",
    "Profit_Bridge",
]


def load_workbook(data_only=False):
    assert OUTPUT.exists(), f"Missing output workbook: {OUTPUT}"
    return openpyxl.load_workbook(OUTPUT, data_only=data_only)


def formula_text(cell):
    value = cell.value
    return value if isinstance(value, str) and value.startswith("=") else None


def assert_formula(cell, *required_tokens):
    formula = formula_text(cell)
    assert formula is not None, f"Expected formula in {cell.coordinate}"
    normalized = formula.upper()
    for token in required_tokens:
        assert token.upper() in normalized, f"Expected {cell.coordinate} to reference {token}"
    return formula


def read_inputs(wb):
    planner = wb["Planner"]
    demand_ws = wb["Store_Demand"]
    shifts_ws = wb["Shift_Definitions"]
    power_ws = wb["Power_Curve"]
    tariff_ws = wb["Tariff_Template"]

    assumptions = {
        "cases_per_handler": float(planner["B2"].value),
        "gross_margin_per_case": float(planner["B3"].value),
        "fixed_site_cost": float(planner["B4"].value),
        "supervisor_buffer": float(planner["B5"].value),
    }

    shares = {}
    for row in range(9, 33):
        shares[int(planner[f"A{row}"].value)] = float(planner[f"B{row}"].value)

    demands = []
    row = 2
    while demand_ws[f"A{row}"].value:
        demands.append(
            {
                "date": demand_ws[f"A{row}"].value,
                "store": demand_ws[f"B{row}"].value,
                "cases": float(demand_ws[f"C{row}"].value),
            }
        )
        row += 1

    shifts = {}
    row = 2
    while shifts_ws[f"A{row}"].value:
        shifts[shifts_ws[f"A{row}"].value] = {
            "start": float(shifts_ws[f"B{row}"].value),
            "end": float(shifts_ws[f"C{row}"].value),
            "regular": float(shifts_ws[f"D{row}"].value),
            "base": float(shifts_ws[f"E{row}"].value),
            "wage": float(shifts_ws[f"F{row}"].value),
            "ot_mult": float(shifts_ws[f"G{row}"].value),
        }
        row += 1

    power = {}
    for row in range(2, 26):
        power[int(power_ws[f"A{row}"].value)] = {
            "base_refrig": float(power_ws[f"B{row}"].value),
            "dock": float(power_ws[f"C{row}"].value),
            "handling": float(power_ws[f"D{row}"].value),
            "battery": float(power_ws[f"E{row}"].value),
        }

    tariff = {}
    for row in range(2, 26):
        tariff[int(tariff_ws[f"A{row}"].value)] = {
            "period": tariff_ws[f"B{row}"].value,
            "rate": float(tariff_ws[f"C{row}"].value),
        }

    return assumptions, shares, demands, shifts, power, tariff


def compute_expected(wb):
    assumptions, shares, demands, shifts, power, tariff = read_inputs(wb)

    total_by_date = {}
    for demand in demands:
        total_by_date.setdefault(demand["date"], 0.0)
        total_by_date[demand["date"]] += demand["cases"]

    hourly_ws = wb["Hourly_Load"]
    hourly_rows = []
    for row in range(2, 170):
        date = hourly_ws[f"A{row}"].value
        hour = int(hourly_ws[f"B{row}"].value)
        daily_cases = total_by_date[date]
        share = shares[hour]
        hourly_cases = daily_cases * share
        total_kwh = (
            power[hour]["base_refrig"]
            + power[hour]["dock"]
            + power[hour]["battery"]
            + hourly_cases * power[hour]["handling"]
        )
        electricity_cost = total_kwh * tariff[hour]["rate"]
        hourly_rows.append(
            {
                "date": date,
                "hour": hour,
                "daily_cases": daily_cases,
                "share": share,
                "hourly_cases": hourly_cases,
                "base_refrig": power[hour]["base_refrig"],
                "dock": power[hour]["dock"],
                "handling": power[hour]["handling"],
                "battery": power[hour]["battery"],
                "total_kwh": total_kwh,
                "period": tariff[hour]["period"],
                "rate": tariff[hour]["rate"],
                "electricity_cost": electricity_cost,
            }
        )

    labor_ws = wb["Labor_Schedule"]
    labor_rows = []
    for row in range(2, 23):
        date = labor_ws[f"A{row}"].value
        shift_id = labor_ws[f"B{row}"].value
        shift = shifts[shift_id]
        shift_length = shift["end"] - shift["start"]
        shift_cases = sum(
            item["hourly_cases"]
            for item in hourly_rows
            if item["date"] == date and shift["start"] <= item["hour"] < shift["end"]
        )
        required_handlers = ceil(shift_cases / shift_length / assumptions["cases_per_handler"])
        scheduled_headcount = max(shift["base"], required_handlers + assumptions["supervisor_buffer"])
        scheduled_hours = scheduled_headcount * shift_length
        overtime_hours = scheduled_headcount * max(shift_length - shift["regular"], 0)
        base_labor_cost = scheduled_hours * shift["wage"]
        overtime_premium = overtime_hours * shift["wage"] * (shift["ot_mult"] - 1)
        labor_rows.append(
            {
                "date": date,
                "shift_id": shift_id,
                "start": shift["start"],
                "end": shift["end"],
                "shift_length": shift_length,
                "base_headcount": shift["base"],
                "shift_cases": shift_cases,
                "required_handlers": required_handlers,
                "scheduled_headcount": scheduled_headcount,
                "scheduled_hours": scheduled_hours,
                "overtime_hours": overtime_hours,
                "hourly_wage": shift["wage"],
                "base_labor_cost": base_labor_cost,
                "overtime_premium": overtime_premium,
                "total_labor_cost": base_labor_cost + overtime_premium,
            }
        )

    summary = {
        "total_cases": sum(row["hourly_cases"] for row in hourly_rows),
        "total_kwh": sum(row["total_kwh"] for row in hourly_rows),
        "offpeak_cost": sum(row["electricity_cost"] for row in hourly_rows if row["period"] == "OffPeak"),
        "shoulder_cost": sum(row["electricity_cost"] for row in hourly_rows if row["period"] == "Shoulder"),
        "peak_cost": sum(row["electricity_cost"] for row in hourly_rows if row["period"] == "Peak"),
        "critical_cost": sum(row["electricity_cost"] for row in hourly_rows if row["period"] == "Critical"),
        "base_labor_cost": sum(row["base_labor_cost"] for row in labor_rows),
        "overtime_premium": sum(row["overtime_premium"] for row in labor_rows),
    }
    summary["total_labor_cost"] = summary["base_labor_cost"] + summary["overtime_premium"]
    summary["total_electricity_cost"] = sum(row["electricity_cost"] for row in hourly_rows)
    summary["total_operating_cost"] = summary["total_labor_cost"] + summary["total_electricity_cost"]

    bridge = {
        "weekly_gross_margin": summary["total_cases"] * assumptions["gross_margin_per_case"],
        "base_labor_cost": -summary["base_labor_cost"],
        "overtime_premium": -summary["overtime_premium"],
        "electricity_cost": -summary["total_electricity_cost"],
        "fixed_site_cost": -assumptions["fixed_site_cost"],
    }
    bridge["weekly_operating_profit"] = sum(bridge.values())

    return hourly_rows, labor_rows, summary, bridge


def test_required_sheets_and_imported_inputs_exist():
    wb = load_workbook()
    assert wb.sheetnames == REQUIRED_SHEETS

    demand = wb["Store_Demand"]
    shifts = wb["Shift_Definitions"]
    power = wb["Power_Curve"]
    tariff = wb["Tariff_Template"]

    assert demand.max_row == 29
    assert shifts.max_row == 4
    assert power.max_row == 25
    assert tariff.max_row == 25

    assert demand["A2"].value == "2026-07-06"
    assert demand["C29"].value == 345
    assert shifts["A2"].value == "Night_Prep"
    assert tariff["B18"].value == "Critical"


def test_hourly_load_formulas_are_present():
    wb = load_workbook()
    ws = wb["Hourly_Load"]

    for row in range(2, 170):
        assert_formula(ws[f"C{row}"], "Store_Demand!")
        assert_formula(ws[f"D{row}"], "Planner!")
        assert_formula(ws[f"E{row}"])
        assert_formula(ws[f"F{row}"], "Power_Curve!")
        assert_formula(ws[f"G{row}"], "Power_Curve!")
        assert_formula(ws[f"H{row}"], "Power_Curve!")
        assert_formula(ws[f"I{row}"], "Power_Curve!")
        assert_formula(ws[f"J{row}"])
        assert_formula(ws[f"K{row}"], "Tariff_Template!")
        assert_formula(ws[f"L{row}"], "Tariff_Template!")
        assert_formula(ws[f"M{row}"])


def test_labor_schedule_formulas_are_present():
    wb = load_workbook()
    ws = wb["Labor_Schedule"]

    for row in range(2, 23):
        assert_formula(ws[f"C{row}"], "Shift_Definitions!")
        assert_formula(ws[f"D{row}"], "Shift_Definitions!")
        assert_formula(ws[f"E{row}"])
        assert_formula(ws[f"F{row}"], "Shift_Definitions!")
        assert_formula(ws[f"G{row}"], "Hourly_Load!")
        assert_formula(ws[f"H{row}"])
        assert_formula(ws[f"I{row}"])
        assert_formula(ws[f"J{row}"])
        assert_formula(ws[f"K{row}"])
        assert_formula(ws[f"L{row}"], "Shift_Definitions!")
        assert_formula(ws[f"M{row}"])
        assert_formula(ws[f"N{row}"])
        assert_formula(ws[f"O{row}"])


def test_summary_and_bridge_formulas_are_present():
    wb = load_workbook()
    summary = wb["Cost_Summary"]
    bridge = wb["Profit_Bridge"]

    for row in range(2, 13):
        assert_formula(summary[f"B{row}"])

    assert_formula(bridge["B2"], "Planner!")
    assert_formula(bridge["B3"], "Cost_Summary!")
    assert_formula(bridge["B4"], "Cost_Summary!")
    assert_formula(bridge["B5"], "Cost_Summary!")
    assert_formula(bridge["B6"], "Planner!")
    assert_formula(bridge["B7"])


def test_recalculated_values_match_expected_results():
    wb_formula = load_workbook()
    wb_values = load_workbook(data_only=True)
    hourly_expected, labor_expected, summary_expected, bridge_expected = compute_expected(wb_formula)

    hourly_actual = wb_values["Hourly_Load"]
    for idx, expected in enumerate(hourly_expected, start=2):
        assert isclose(float(hourly_actual[f"C{idx}"].value), expected["daily_cases"], rel_tol=0, abs_tol=1e-6)
        assert isclose(float(hourly_actual[f"D{idx}"].value), expected["share"], rel_tol=0, abs_tol=1e-9)
        assert isclose(float(hourly_actual[f"E{idx}"].value), expected["hourly_cases"], rel_tol=0, abs_tol=1e-6)
        assert isclose(float(hourly_actual[f"J{idx}"].value), expected["total_kwh"], rel_tol=0, abs_tol=1e-6)
        assert hourly_actual[f"K{idx}"].value == expected["period"]
        assert isclose(float(hourly_actual[f"L{idx}"].value), expected["rate"], rel_tol=0, abs_tol=1e-9)
        assert isclose(float(hourly_actual[f"M{idx}"].value), expected["electricity_cost"], rel_tol=0, abs_tol=1e-6)

    labor_actual = wb_values["Labor_Schedule"]
    for idx, expected in enumerate(labor_expected, start=2):
        assert isclose(float(labor_actual[f"E{idx}"].value), expected["shift_length"], rel_tol=0, abs_tol=1e-6)
        assert isclose(float(labor_actual[f"G{idx}"].value), expected["shift_cases"], rel_tol=0, abs_tol=1e-6)
        assert isclose(float(labor_actual[f"H{idx}"].value), expected["required_handlers"], rel_tol=0, abs_tol=1e-6)
        assert isclose(float(labor_actual[f"I{idx}"].value), expected["scheduled_headcount"], rel_tol=0, abs_tol=1e-6)
        assert isclose(float(labor_actual[f"K{idx}"].value), expected["overtime_hours"], rel_tol=0, abs_tol=1e-6)
        assert isclose(float(labor_actual[f"M{idx}"].value), expected["base_labor_cost"], rel_tol=0, abs_tol=1e-6)
        assert isclose(float(labor_actual[f"N{idx}"].value), expected["overtime_premium"], rel_tol=0, abs_tol=1e-6)
        assert isclose(float(labor_actual[f"O{idx}"].value), expected["total_labor_cost"], rel_tol=0, abs_tol=1e-6)

    summary_actual = wb_values["Cost_Summary"]
    assert isclose(float(summary_actual["B2"].value), summary_expected["total_cases"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(summary_actual["B3"].value), summary_expected["total_kwh"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(summary_actual["B4"].value), summary_expected["offpeak_cost"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(summary_actual["B5"].value), summary_expected["shoulder_cost"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(summary_actual["B6"].value), summary_expected["peak_cost"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(summary_actual["B7"].value), summary_expected["critical_cost"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(summary_actual["B8"].value), summary_expected["base_labor_cost"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(summary_actual["B9"].value), summary_expected["overtime_premium"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(summary_actual["B10"].value), summary_expected["total_labor_cost"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(summary_actual["B11"].value), summary_expected["total_electricity_cost"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(summary_actual["B12"].value), summary_expected["total_operating_cost"], rel_tol=0, abs_tol=1e-6)

    bridge_actual = wb_values["Profit_Bridge"]
    assert isclose(float(bridge_actual["B2"].value), bridge_expected["weekly_gross_margin"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(bridge_actual["B3"].value), bridge_expected["base_labor_cost"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(bridge_actual["B4"].value), bridge_expected["overtime_premium"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(bridge_actual["B5"].value), bridge_expected["electricity_cost"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(bridge_actual["B6"].value), bridge_expected["fixed_site_cost"], rel_tol=0, abs_tol=1e-6)
    assert isclose(float(bridge_actual["B7"].value), bridge_expected["weekly_operating_profit"], rel_tol=0, abs_tol=1e-6)
