from pathlib import Path

import openpyxl

OUTPUT_FILE = Path("manufacturer_similar_completed.xlsx")
REQUIRED_SHEETS = ["Historical", "Depreciation", "Shock", "Assumptions", "Model"]


def load_workbook_checked():
    assert OUTPUT_FILE.exists(), f"missing output workbook: {OUTPUT_FILE}"
    return openpyxl.load_workbook(OUTPUT_FILE)


def assert_required_sheets(wb):
    assert wb.sheetnames == REQUIRED_SHEETS, f"unexpected sheets: {wb.sheetnames}"


def assert_no_todo(wb):
    for sheet_name in ["Depreciation", "Model"]:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            for value in row:
                assert value != "TODO", f"TODO remains in {sheet_name}"


def assert_depreciation_formulas(wb):
    dep = wb["Depreciation"]
    for row in range(2, 10):
        assert dep[f"D{row}"].value == f"=B{row}/C{row}"


def assert_historical_block(wb):
    model = wb["Model"]
    assert model["B3"].value == "=AVERAGE(Depreciation!D6:D9)"
    assert model["A7"].value == "=Historical!A2"
    assert model["B7"].value == "=Historical!B2"
    assert model["C7"].value == "=Historical!C2"
    assert model["D7"].value == "=Historical!D2"
    assert model["E7"].value == "=LN(B7)"
    assert model["F7"].value == "=LN(C7)"
    assert model["G7"].value == "=C7/((B7^Assumptions!$B$2)*(D7^(1-Assumptions!$B$2)))"
    assert model["G17"].value == "=AVERAGE(G11:G14)"


def assert_projection_block(wb):
    model = wb["Model"]
    assert model["A20"].value == "=Shock!A2"
    assert model["B20"].value == "=B14*(1-$B$3)"
    assert model["C20"].value == "=B14*(1-$B$3)+Shock!B2"
    assert model["D20"].value == "=$G$17*((B20^Assumptions!$B$2)*(Assumptions!$B$3^(1-Assumptions!$B$2)))"
    assert model["E20"].value == "=$G$17*((C20^Assumptions!$B$2)*(Assumptions!$B$3^(1-Assumptions!$B$2)))"
    assert model["B21"].value == "=B20*(1-$B$3)"
    assert model["C21"].value == "=C20*(1-$B$3)+Shock!B3"
    assert model["A27"].value == "=Shock!A9"


def main():
    wb = load_workbook_checked()
    assert_required_sheets(wb)
    assert_no_todo(wb)
    assert_depreciation_formulas(wb)
    assert_historical_block(wb)
    assert_projection_block(wb)


if __name__ == "__main__":
    main()
