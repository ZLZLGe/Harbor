from pathlib import Path

import openpyxl

OUTPUT_FILE = Path("energy_transfer_completed.xlsx")


def load_workbook_checked():
    assert OUTPUT_FILE.exists(), f"missing output workbook: {OUTPUT_FILE}"
    return openpyxl.load_workbook(OUTPUT_FILE)


def assert_no_todo(wb):
    ws = wb["Results"]
    for row in ws.iter_rows(values_only=True):
        for value in row:
            assert value != "TODO", "TODO remains in Results"


def assert_inputs(wb):
    ws = wb["Results"]
    assert ws["B3"].value == "=(Actuals!B6/Actuals!B2)^(1/4)-1"
    assert ws["B4"].value == "=(Actuals!C6/Actuals!C2)^(1/4)-1"
    assert ws["B5"].value == "=AVERAGE(Actuals!D4:D6)"


def assert_result_rows(wb):
    ws = wb["Results"]
    assert ws["A8"].value == "=Upgrade!A2"
    assert ws["B8"].value == "=Actuals!B6*(1+$B$3)"
    assert ws["C8"].value == "=B8*(1-$B$5)"
    assert ws["D8"].value == "=B8*(1-Upgrade!B2)"
    assert ws["E8"].value == "=Actuals!C6*(1+$B$4)"
    assert ws["F8"].value == "=(D8-C8)*E8"
    assert ws["G8"].value == "=F8-Upgrade!C2"
    assert ws["H8"].value == "=G8"
    assert ws["B9"].value == "=B8*(1+$B$3)"
    assert ws["H12"].value == "=H11+G12"


def main():
    wb = load_workbook_checked()
    assert_no_todo(wb)
    assert_inputs(wb)
    assert_result_rows(wb)


if __name__ == "__main__":
    main()
