from pathlib import Path

import openpyxl

OUTPUT_FILE = Path("port_transfer_completed.xlsx")


def load_workbook_checked():
    assert OUTPUT_FILE.exists(), f"missing output workbook: {OUTPUT_FILE}"
    return openpyxl.load_workbook(OUTPUT_FILE)


def assert_no_todo(wb):
    ws = wb["Model"]
    for row in ws.iter_rows(values_only=True):
        for value in row:
            assert value != "TODO", "TODO remains in Model"


def assert_growth_inputs(wb):
    ws = wb["Model"]
    assert ws["B3"].value == "=(Actuals!B6/Actuals!B2)^(1/4)-1"
    assert ws["B4"].value == "=(Actuals!C6/Actuals!C2)^(1/4)-1"


def assert_model_rows(wb):
    ws = wb["Model"]
    assert ws["A8"].value == "=Expansion!A2"
    assert ws["B8"].value == "=Actuals!B6*(1+$B$3)"
    assert ws["C8"].value == "=Actuals!D6+SUM(Expansion!B$2:B2)"
    assert ws["D8"].value == "=C8*$B$5*(1+Expansion!C2)"
    assert ws["E8"].value == "=MIN(B8,D8)"
    assert ws["F8"].value == "=Actuals!C6*(1+$B$4)"
    assert ws["G8"].value == "=E8*F8"
    assert ws["H8"].value == "=D8-B8"
    assert ws["B9"].value == "=B8*(1+$B$3)"
    assert ws["C12"].value == "=Actuals!D6+SUM(Expansion!B$2:B6)"


def main():
    wb = load_workbook_checked()
    assert_no_todo(wb)
    assert_growth_inputs(wb)
    assert_model_rows(wb)


if __name__ == "__main__":
    main()
