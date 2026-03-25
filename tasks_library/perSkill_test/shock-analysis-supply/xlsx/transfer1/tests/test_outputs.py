from pathlib import Path

import openpyxl

OUTPUT_FILE = Path("tourism_transfer_completed.xlsx")


def load_workbook_checked():
    assert OUTPUT_FILE.exists(), f"missing output workbook: {OUTPUT_FILE}"
    return openpyxl.load_workbook(OUTPUT_FILE)


def assert_no_todo(wb):
    ws = wb["Forecast"]
    for row in ws.iter_rows(values_only=True):
        for value in row:
            assert value != "TODO", "TODO remains in Forecast"


def assert_growth_inputs(wb):
    ws = wb["Forecast"]
    assert ws["B3"].value == "=(Actuals!B6/Actuals!B2)^(1/4)-1"
    assert ws["B4"].value == "=(Actuals!C6/Actuals!C2)^(1/4)-1"


def assert_forecast_rows(wb):
    ws = wb["Forecast"]
    assert ws["A8"].value == "=Projects!A2"
    assert ws["B8"].value == "=Actuals!B6*(1+$B$3)"
    assert ws["C8"].value == "=B8*(1+Projects!B2)"
    assert ws["D8"].value == "=Actuals!C6*(1+$B$4)"
    assert ws["E8"].value == "=(Actuals!D6+SUM(Projects!C$2:C2))*$B$5"
    assert ws["F8"].value == "=MIN(C8,E8)"
    assert ws["G8"].value == "=F8*D8"
    assert ws["B9"].value == "=B8*(1+$B$3)"
    assert ws["D9"].value == "=D8*(1+$B$4)"
    assert ws["E12"].value == "=(Actuals!D6+SUM(Projects!C$2:C6))*$B$5"


def main():
    wb = load_workbook_checked()
    assert_no_todo(wb)
    assert_growth_inputs(wb)
    assert_forecast_rows(wb)


if __name__ == "__main__":
    main()
