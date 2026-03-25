from math import isclose, sqrt
from pathlib import Path

import openpyxl


OUTPUT = Path("/root/bottling-expansion-model.xlsx")
REQUIRED_SHEETS = [
    "Historical_Data",
    "Expansion_Plan",
    "Assumptions",
    "Efficiency_Trend",
    "Capacity_Model",
    "Summary",
]


def load_workbook(data_only=False):
    assert OUTPUT.exists(), f"Missing output workbook: {OUTPUT}"
    return openpyxl.load_workbook(OUTPUT, data_only=data_only)


def formula_text(cell):
    value = cell.value
    return value if isinstance(value, str) and value.startswith("=") else None


def compute_expected(wb):
    hist = wb["Historical_Data"]
    plan = wb["Expansion_Plan"]
    assumptions = wb["Assumptions"]

    history = []
    for row in range(2, 9):
        history.append(
            (
                int(hist[f"A{row}"].value),
                float(hist[f"B{row}"].value),
                float(hist[f"C{row}"].value),
                float(hist[f"D{row}"].value),
            )
        )

    capex_plan = []
    for row in range(2, 7):
        capex_plan.append(
            (
                int(plan[f"A{row}"].value),
                float(plan[f"B{row}"].value),
                float(plan[f"C{row}"].value),
            )
        )

    price = float(assumptions["B2"].value)
    variable_cost = float(assumptions["B3"].value)
    maintenance_rate = float(assumptions["B4"].value)
    life = float(assumptions["B5"].value)
    growth_cap = float(assumptions["B6"].value)
    utilization = float(assumptions["B7"].value)

    smoothed = {}
    raw_efficiency = {}
    growth = {}
    raw_rows = []
    for year, cases, capital, labor in history:
        raw = sqrt((cases / labor) * (cases / capital))
        raw_efficiency[year] = raw
        raw_rows.append((year, cases, raw))

    for idx, (year, _, raw) in enumerate(raw_rows):
        if idx < 2:
            smoothed[year] = raw
        else:
            smoothed[year] = (
                0.2 * raw_rows[idx - 2][2]
                + 0.3 * raw_rows[idx - 1][2]
                + 0.5 * raw
            )

    xs = [row[0] for row in raw_rows]
    ys = [smoothed[row[0]] for row in raw_rows]
    x_bar = sum(xs) / len(xs)
    y_bar = sum(ys) / len(ys)
    slope = sum((x - x_bar) * (y - y_bar) for x, y in zip(xs, ys)) / sum(
        (x - x_bar) ** 2 for x in xs
    )
    intercept = y_bar - slope * x_bar

    for year in range(2025, 2030):
        smoothed[year] = intercept + slope * year

    for year in range(2019, 2030):
        growth[year] = smoothed[year] / smoothed[year - 1] - 1

    baseline = {2024: float(history[-1][1])}
    for year in range(2025, 2030):
        baseline[year] = baseline[year - 1] * (1 + min(growth[year], growth_cap))

    expected_rows = {}
    cumulative_capacity = 0.0
    net_capital = 0.0
    smoothed_2024 = smoothed[2024]
    for year, capex, added_capacity in capex_plan:
        depreciation = net_capital / life
        net_capital = net_capital - depreciation + capex
        cumulative_capacity += added_capacity
        with_expansion = baseline[year] + cumulative_capacity * (smoothed[year] / smoothed_2024) * utilization
        ebitda = (
            (with_expansion - baseline[year]) * (price - variable_cost)
            - net_capital * maintenance_rate
        )
        operating_profit = ebitda - depreciation
        expected_rows[year] = {
            "smoothed": smoothed[year],
            "baseline": baseline[year],
            "cum_capacity": cumulative_capacity,
            "with_expansion": with_expansion,
            "net_capital": net_capital,
            "depreciation": depreciation,
            "ebitda": ebitda,
            "operating_profit": operating_profit,
        }

    return {
        "capacity": expected_rows,
        "raw_efficiency": raw_efficiency,
        "smoothed": smoothed,
        "growth": growth,
    }


def test_required_sheets_exist():
    wb = load_workbook()
    assert wb.sheetnames == REQUIRED_SHEETS


def test_efficiency_sheet_uses_links_and_matches_expected_values():
    wb_formula = load_workbook()
    wb_values = load_workbook(data_only=True)
    expected = compute_expected(wb_formula)
    ws_formula = wb_formula["Efficiency_Trend"]
    ws_values = wb_values["Efficiency_Trend"]

    for row in range(2, 9):
        year = int(ws_values[f"A{row}"].value)
        assert "Historical_Data!" in formula_text(ws_formula[f"B{row}"])
        assert "Historical_Data!" in formula_text(ws_formula[f"C{row}"])
        assert "Historical_Data!" in formula_text(ws_formula[f"D{row}"])
        assert formula_text(ws_formula[f"E{row}"]) is not None
        assert formula_text(ws_formula[f"F{row}"]) is not None
        assert formula_text(ws_formula[f"G{row}"]) is not None
        assert isclose(
            float(ws_values[f"G{row}"].value),
            expected["raw_efficiency"][year],
            rel_tol=0,
            abs_tol=1e-6,
        )

    for row in range(2, 14):
        year = int(ws_values[f"A{row}"].value)
        formula = formula_text(ws_formula[f"H{row}"])
        assert formula is not None
        assert isclose(
            float(ws_values[f"H{row}"].value),
            expected["smoothed"][year],
            rel_tol=0,
            abs_tol=1e-6,
        )

    for row in range(3, 14):
        year = int(ws_values[f"A{row}"].value)
        assert formula_text(ws_formula[f"I{row}"]) is not None
        assert isclose(
            float(ws_values[f"I{row}"].value),
            expected["growth"][year],
            rel_tol=0,
            abs_tol=1e-6,
        )


def test_capacity_model_uses_recursive_formula_structure():
    wb = load_workbook()
    ws = wb["Capacity_Model"]

    for row in range(3, 8):
        assert formula_text(ws[f"D{row}"]) is not None
        assert "Expansion_Plan!" in formula_text(ws[f"F{row}"])
        assert "Expansion_Plan!" in formula_text(ws[f"G{row}"])
        assert formula_text(ws[f"H{row}"]) is not None
        assert "Assumptions!" in formula_text(ws[f"I{row}"])
        assert formula_text(ws[f"J{row}"]) is not None
        assert "Assumptions!" in formula_text(ws[f"K{row}"])
        assert "Assumptions!" in formula_text(ws[f"L{row}"])
        assert formula_text(ws[f"M{row}"]) is not None


def test_summary_formulas_are_present():
    wb = load_workbook()
    ws = wb["Summary"]

    assert "SUM(" in formula_text(ws["B2"]).upper()
    assert "SUM(" in formula_text(ws["B3"]).upper()
    assert "Capacity_Model!" in formula_text(ws["B4"])

    for row in range(7, 12):
        assert "Capacity_Model!" in formula_text(ws[f"E{row}"])
        assert "Capacity_Model!" in formula_text(ws[f"H{row}"])


def test_recalculated_values_match_expected_outputs():
    wb_formula = load_workbook()
    wb_values = load_workbook(data_only=True)
    expected = compute_expected(wb_formula)

    model = wb_values["Capacity_Model"]
    summary = wb_values["Summary"]

    year_to_row = {2024: 2, 2025: 3, 2026: 4, 2027: 5, 2028: 6, 2029: 7}

    for year in range(2025, 2030):
        row = year_to_row[year]
        actual_baseline = float(model[f"E{row}"].value)
        actual_with = float(model[f"I{row}"].value)
        actual_net_capital = float(model[f"J{row}"].value)
        actual_dep = float(model[f"K{row}"].value)
        actual_ebitda = float(model[f"L{row}"].value)
        actual_operating_profit = float(model[f"M{row}"].value)

        assert isclose(actual_baseline, expected["capacity"][year]["baseline"], rel_tol=0, abs_tol=1e-6)
        assert isclose(actual_with, expected["capacity"][year]["with_expansion"], rel_tol=0, abs_tol=1e-6)
        assert isclose(actual_net_capital, expected["capacity"][year]["net_capital"], rel_tol=0, abs_tol=1e-6)
        assert isclose(actual_dep, expected["capacity"][year]["depreciation"], rel_tol=0, abs_tol=1e-6)
        assert isclose(actual_ebitda, expected["capacity"][year]["ebitda"], rel_tol=0, abs_tol=1e-6)
        assert isclose(
            actual_operating_profit,
            expected["capacity"][year]["operating_profit"],
            rel_tol=0,
            abs_tol=1e-6,
        )

    total_ebitda = sum(expected["capacity"][year]["ebitda"] for year in range(2025, 2030))
    total_dep = sum(expected["capacity"][year]["depreciation"] for year in range(2025, 2030))
    with_2029 = expected["capacity"][2029]["with_expansion"]

    assert isclose(float(summary["B2"].value), total_ebitda, rel_tol=0, abs_tol=1e-6)
    assert isclose(float(summary["B3"].value), total_dep, rel_tol=0, abs_tol=1e-6)
    assert isclose(float(summary["B4"].value), with_2029, rel_tol=0, abs_tol=1e-6)

    for idx, year in enumerate(range(2025, 2030), start=7):
        assert isclose(float(summary[f"E{idx}"].value), expected["capacity"][year]["ebitda"], rel_tol=0, abs_tol=1e-6)
        assert isclose(float(summary[f"H{idx}"].value), expected["capacity"][year]["depreciation"], rel_tol=0, abs_tol=1e-6)
