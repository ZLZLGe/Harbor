from pathlib import Path
import re

import openpyxl
import pytest
from openpyxl.worksheet.formula import ArrayFormula


OUTPUT_FILE = Path("tourism-capacity-shock.xlsx")
REQUIRED_SHEETS = [
    "Tourism_Capital",
    "Visitor_Spend",
    "Expansion_Plan",
    "Demand_Forecast",
    "Revenue_Model",
]


def load_workbook():
    assert OUTPUT_FILE.exists(), f"Output file not found: {OUTPUT_FILE}"
    return openpyxl.load_workbook(OUTPUT_FILE)


def get_formula_text(value):
    if value is None:
        return None
    if isinstance(value, ArrayFormula):
        return value.text
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def normalize_formula(formula):
    if formula is None:
        return None
    return formula.replace(" ", "").replace("$", "").upper()


def is_correct_expansion_link(row, formula):
    normalized = normalize_formula(formula)
    if normalized is None:
        return False

    target_row = row - 6
    if normalized == f"=EXPANSION_PLAN!B{target_row}":
        return True

    lookup_ref_pattern = rf"(?:A{row}|TOURISM_CAPITAL!A{row})"
    patterns = [
        rf"=INDEX\(EXPANSION_PLAN!B8:B12,MATCH\({lookup_ref_pattern},EXPANSION_PLAN!A8:A12,0\)\)",
        rf"=XLOOKUP\({lookup_ref_pattern},EXPANSION_PLAN!A8:A12,EXPANSION_PLAN!B8:B12\)",
        rf"=VLOOKUP\({lookup_ref_pattern},EXPANSION_PLAN!A8:B12,2,FALSE\)",
    ]
    return any(re.fullmatch(pattern, normalized) for pattern in patterns)


def test_required_sheets_exist():
    wb = load_workbook()
    assert wb.sheetnames == REQUIRED_SHEETS


def test_input_ranges_are_preserved():
    wb = load_workbook()

    capital = wb["Tourism_Capital"]
    spend = wb["Visitor_Spend"]
    expansion = wb["Expansion_Plan"]
    demand = wb["Demand_Forecast"]

    assert capital["B2"].value == 180
    assert spend["B2"].value == 0.045
    assert spend["B3"].value == 1.18
    assert demand["B2"].value == 0.08

    expected_years = [2026, 2027, 2028, 2029, 2030]
    expected_added_rooms = [18, 24, 30, 35, 35]

    assert [expansion[f"A{row}"].value for row in range(8, 13)] == expected_years
    assert [expansion[f"B{row}"].value for row in range(8, 13)] == expected_added_rooms


def test_tourism_capital_formulas():
    wb = load_workbook()
    ws = wb["Tourism_Capital"]

    b3 = get_formula_text(ws["B3"].value)
    assert b3 and "AVERAGE" in b3.upper()

    for row in range(14, 19):
        formula = get_formula_text(ws[f"C{row}"].value)
        assert formula == "=$B$3"

    for row in range(8, 14):
        formula = get_formula_text(ws[f"D{row}"].value)
        assert formula == "=0"

    for row in range(14, 19):
        formula = get_formula_text(ws[f"D{row}"].value)
        assert formula is not None
        assert is_correct_expansion_link(row, formula), (
            f"Tourism_Capital!D{row} must link the matching year from Expansion_Plan, "
            f"got {formula!r}"
        )

    for row in range(8, 19):
        total_formula = get_formula_text(ws[f"E{row}"].value)
        capex_formula = get_formula_text(ws[f"F{row}"].value)
        assert total_formula == f"=B{row}+D{row}"
        assert capex_formula == f"=D{row}*$B$2"


def test_visitor_spend_formulas():
    wb = load_workbook()
    ws = wb["Visitor_Spend"]

    for row in range(8, 14):
        formula = get_formula_text(ws[f"D{row}"].value)
        assert formula == f"=C{row}"

    for row in range(14, 19):
        formula = get_formula_text(ws[f"D{row}"].value)
        assert formula == f"=D{row-1}*(1+$B$2)"

    for row in range(8, 19):
        formula = get_formula_text(ws[f"E{row}"].value)
        assert formula == f"=D{row}*$B$3"


def test_expansion_plan_cumulative_rooms():
    wb = load_workbook()
    ws = wb["Expansion_Plan"]

    for row in range(8, 13):
        formula = get_formula_text(ws[f"C{row}"].value)
        assert formula == f"=SUM($B$8:B{row})"


def test_demand_forecast_formulas():
    wb = load_workbook()
    ws = wb["Demand_Forecast"]

    for row in range(8, 14):
        formula = get_formula_text(ws[f"B{row}"].value)
        assert formula == f"=Visitor_Spend!B{row}"

    for row in range(14, 19):
        formula = get_formula_text(ws[f"B{row}"].value)
        assert formula == f"=TREND($B$8:$B$13,$A$8:$A$13,A{row})"

    for row in range(8, 19):
        baseline_capacity = get_formula_text(ws[f"C{row}"].value)
        baseline_served = get_formula_text(ws[f"D{row}"].value)
        expansion_capacity = get_formula_text(ws[f"F{row}"].value)
        expansion_served = get_formula_text(ws[f"G{row}"].value)

        assert baseline_capacity == f"=Tourism_Capital!B{row}*365*Tourism_Capital!C{row}"
        assert baseline_served == f"=MIN(B{row},C{row})"
        assert expansion_capacity == f"=Tourism_Capital!E{row}*365*Tourism_Capital!C{row}"
        assert expansion_served == f"=MIN(E{row},F{row})"

    for row in range(8, 14):
        formula = get_formula_text(ws[f"E{row}"].value)
        assert formula == f"=B{row}"

    for row in range(14, 19):
        formula = get_formula_text(ws[f"E{row}"].value)
        assert formula == f"=B{row}*(1+$B$2)"


def test_revenue_model_formulas():
    wb = load_workbook()
    ws = wb["Revenue_Model"]

    demand_rows = range(12, 19)
    for model_row, demand_row in zip(range(8, 15), demand_rows):
        assert get_formula_text(ws[f"B{model_row}"].value) == f"=Demand_Forecast!D{demand_row}"
        assert get_formula_text(ws[f"C{model_row}"].value) == f"=Demand_Forecast!G{demand_row}"
        assert get_formula_text(ws[f"D{model_row}"].value) == f"=Visitor_Spend!E{demand_row}"
        assert get_formula_text(ws[f"E{model_row}"].value) == f"=B{model_row}*D{model_row}"
        assert get_formula_text(ws[f"F{model_row}"].value) == f"=C{model_row}*D{model_row}"
        assert get_formula_text(ws[f"G{model_row}"].value) == f"=F{model_row}-E{model_row}"
        assert get_formula_text(ws[f"H{model_row}"].value) == f"=Tourism_Capital!F{demand_row}"
        assert get_formula_text(ws[f"I{model_row}"].value) == f'=IF(H{model_row}=0,"",G{model_row}/H{model_row})'
