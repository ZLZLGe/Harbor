import json
from openpyxl import load_workbook


INPUT_PATH = "/root/video_tracks.json"
OUTPUT_PATH = "/root/similar_count.xlsx"


def expected_rows() -> list[list[object]]:
    with open(INPUT_PATH, "r", encoding="utf-8") as f:
        payload = json.load(f)

    rows = []
    for video in payload.get("videos", []):
        filename = video["filename"]
        uniq = {
            track.get("track_id")
            for track in video.get("tracks", [])
            if track.get("label") == "pedestrian"
        }
        rows.append([filename, len(uniq)])

    rows.sort(key=lambda item: item[0])
    return rows


def read_actual_rows() -> list[list[object]]:
    wb = load_workbook(OUTPUT_PATH, data_only=True)
    try:
        assert wb.sheetnames == ["results"]
        ws = wb["results"]
        data = []
        for r in range(1, ws.max_row + 1):
            data.append([ws.cell(r, 1).value, ws.cell(r, 2).value])
        return data
    finally:
        wb.close()


def test_similar_count_workbook_exact():
    actual = read_actual_rows()
    assert actual[0] == ["filename", "number"]
    assert actual[1:] == expected_rows()
