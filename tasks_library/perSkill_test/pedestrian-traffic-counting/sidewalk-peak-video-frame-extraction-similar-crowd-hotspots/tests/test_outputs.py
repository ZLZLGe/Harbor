import os
from pathlib import Path
import re
from typing import Any

import cv2
from openpyxl import load_workbook


OUTPUT_FILE = "/app/video/pedestrian_peaks.xlsx"
REWARD_FILE = "/logs/verifier/reward.txt"
VIDEO_DIR = Path("/app/video")
VIDEO_EXTENSIONS = {".avi", ".mp4", ".mov", ".mkv", ".webm"}
EXPECTED_HEADER = ["filename", "peak_frame_id", "visible_pedestrians"]


def _to_string(value: Any) -> str:
    return "" if value is None else str(value)


def _load_rows(path: str) -> tuple[list[str], list[list[str]]]:
    workbook = load_workbook(path, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
        assert sheet_names == ["results"], f"Workbook must contain exactly one sheet named 'results', got {sheet_names}"
        sheet = workbook["results"]
        rows = []
        for row in sheet.iter_rows():
            rows.append([_to_string(cell.value) for cell in row])
        return sheet_names, rows
    finally:
        workbook.close()


def _write_reward(value: float) -> None:
    with open(REWARD_FILE, "w", encoding="utf-8") as handle:
        handle.write(f"{value:.6f}\n")


def _count_pedestrians(frame: Any) -> int:
    red_mask = (
        (frame[:, :, 2] > 180)
        & (frame[:, :, 1] < 90)
        & (frame[:, :, 0] < 90)
    )
    mask = (red_mask.astype("uint8")) * 255
    component_count, _, stats, _ = cv2.connectedComponentsWithStats(mask, connectivity=8)

    people = 0
    for label in range(1, component_count):
        area = stats[label, cv2.CC_STAT_AREA]
        if area >= 80:
            people += 1
    return people


def _analyze_video(video_path: Path) -> list[str]:
    capture = cv2.VideoCapture(str(video_path))
    if not capture.isOpened():
        raise AssertionError(f"Cannot open video: {video_path}")

    best_index = -1
    best_count = -1
    frame_index = 0
    try:
        while True:
            ok, frame = capture.read()
            if not ok:
                break
            people = _count_pedestrians(frame)
            if people > best_count:
                best_index = frame_index
                best_count = people
            frame_index += 1
    finally:
        capture.release()

    if best_index < 0:
        raise AssertionError(f"No frames found in video: {video_path}")

    return [video_path.name, f"F{best_index:04d}", str(best_count)]


def _expected_rows() -> list[list[str]]:
    videos = sorted(
        path for path in VIDEO_DIR.iterdir()
        if path.is_file() and path.suffix.lower() in VIDEO_EXTENSIONS
    )
    return [EXPECTED_HEADER, *[_analyze_video(video_path) for video_path in videos]]


def _frame_reward(actual_frame: str, expected_frame: str) -> float:
    return 1.0 if actual_frame == expected_frame else 0.0


def _count_reward(actual_count: str, expected_count: str) -> float:
    try:
        actual_number = int(actual_count)
        expected_number = int(expected_count)
    except ValueError:
        return 0.0
    return 1.0 / (1.0 + abs(actual_number - expected_number))


def _score_rows(actual_rows: list[list[str]]) -> float:
    expected_rows = _expected_rows()
    actual_map = {
        row[0]: row[1:]
        for row in actual_rows[1:]
        if len(row) >= 3 and row[0].strip()
    }

    total_reward = 0.0
    for filename, expected_frame, expected_count in expected_rows[1:]:
        actual_frame, actual_count = actual_map.get(filename, ["", ""])
        row_reward = 0.6 * _frame_reward(actual_frame, expected_frame) + 0.4 * _count_reward(actual_count, expected_count)
        total_reward += row_reward
        print(
            f"{filename} -> frame {actual_frame} vs {expected_frame}, "
            f"count {actual_count} vs {expected_count}, reward={row_reward:.4f}"
        )
    return total_reward / (len(expected_rows) - 1)


def test_outputs() -> None:
    if not os.path.exists(OUTPUT_FILE):
        _write_reward(0.0)
        raise AssertionError("pedestrian_peaks.xlsx not found at /app/video")

    try:
        expected_rows = _expected_rows()
        _, actual_rows = _load_rows(OUTPUT_FILE)

        print("\nOutput workbook rows:")
        for index, row in enumerate(actual_rows):
            print(index, row)

        assert actual_rows, "Workbook is empty"
        assert actual_rows[0] == EXPECTED_HEADER, f"Header mismatch: {actual_rows[0]}"

        non_empty_rows = [row for row in actual_rows if any(cell.strip() for cell in row)]
        assert len(non_empty_rows) == len(expected_rows), (
            f"Expected {len(expected_rows)} non-empty rows, got {len(non_empty_rows)}"
        )

        for actual in actual_rows[1:]:
            assert len(actual) == 3, f"Each data row must have exactly 3 columns, got {actual}"
            assert re.fullmatch(r"F\d{4}", actual[1]), f"Invalid peak_frame_id format: {actual[1]}"

        _write_reward(_score_rows(actual_rows))

        assert actual_rows == expected_rows, (
            "Workbook content mismatch.\n"
            f"Actual:   {actual_rows}\n"
            f"Expected: {expected_rows}"
        )
    except Exception:
        try:
            _, actual_rows = _load_rows(OUTPUT_FILE)
            _write_reward(_score_rows(actual_rows))
        except Exception:
            _write_reward(0.0)
        raise
