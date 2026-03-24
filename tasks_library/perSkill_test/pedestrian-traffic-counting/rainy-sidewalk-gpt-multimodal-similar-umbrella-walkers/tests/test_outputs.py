from pathlib import Path
from typing import Any

from openpyxl import load_workbook


OUTPUT_FILE = Path("/app/video/umbrella_counts.xlsx")
VIDEO_DIR = Path("/app/video")
REWARD_FILE = Path("/logs/verifier/reward.txt")
VIDEO_EXTENSIONS = {".mp4", ".mkv", ".avi", ".mov", ".wmv", ".flv", ".webm", ".m4v", ".mpeg", ".mpg", ".3gpp"}
EXPECTED_COUNTS = {
    "rain_cam_north.mp4": 0,
    "rain_cam_south.mp4": 0,
}


def _cell_to_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _read_sheet_as_rows(path: Path, sheet_name: str) -> list[list[str]]:
    workbook = load_workbook(path, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        rows: list[list[str]] = []
        for row_index in range(1, (worksheet.max_row or 0) + 1):
            row: list[str] = []
            for col_index in range(1, (worksheet.max_column or 0) + 1):
                row.append(_cell_to_string(worksheet.cell(row=row_index, column=col_index).value))
            rows.append(row)
        return rows
    finally:
        workbook.close()


def _assert_single_results_sheet(path: Path) -> None:
    workbook = load_workbook(path, data_only=True)
    try:
        assert workbook.sheetnames == ["results"], f"工作表必须且只能是 ['results']，实际为 {workbook.sheetnames}"
    finally:
        workbook.close()


def _collect_video_names() -> list[str]:
    return sorted(
        path.name for path in VIDEO_DIR.iterdir()
        if path.is_file() and path.suffix.lower() in VIDEO_EXTENSIONS
    )


def _write_reward(value: float) -> None:
    REWARD_FILE.parent.mkdir(parents=True, exist_ok=True)
    REWARD_FILE.write_text(f"{value:.4f}\n", encoding="utf-8")


def test_outputs() -> None:
    assert OUTPUT_FILE.exists(), f"缺少输出文件: {OUTPUT_FILE}"
    _assert_single_results_sheet(OUTPUT_FILE)

    rows = _read_sheet_as_rows(OUTPUT_FILE, "results")
    assert rows, "结果工作表为空"
    assert rows[0] == ["filename", "umbrella_walkers"], f"表头必须是 ['filename', 'umbrella_walkers']，实际为 {rows[0]}"

    data_rows = [row for row in rows[1:] if any(cell.strip() for cell in row)]
    expected_names = _collect_video_names()
    assert len(data_rows) == len(expected_names), f"数据行数量不匹配，实际 {len(data_rows)}，期望 {len(expected_names)}"
    assert expected_names == sorted(EXPECTED_COUNTS), f"测试中的固定视频集合与环境不一致: {expected_names}"

    actual_map: dict[str, str] = {}
    for row in data_rows:
        assert len(row) == 2, f"每行只能有两列，实际为 {row}"
        filename, count = row
        assert filename not in actual_map, f"文件名重复: {filename}"
        actual_map[filename] = count

    assert sorted(actual_map) == expected_names, f"filename 列必须与视频文件完全一致且按升序排列，实际 {sorted(actual_map)}，期望 {expected_names}"
    assert [row[0] for row in data_rows] == expected_names, "数据行顺序必须按文件名升序排列"

    rewards: list[float] = []
    for video_name in expected_names:
        raw_value = actual_map[video_name]
        assert raw_value.isdigit(), f"计数必须是非负整数，{video_name} 的值为 {raw_value!r}"
        actual_value = int(raw_value)
        expected_value = EXPECTED_COUNTS[video_name]
        reward = 1.0 if actual_value == expected_value else 0.0
        rewards.append(reward)

        print(f"{video_name}: actual={actual_value}, expected={expected_value}, reward={reward:.4f}")
        assert actual_value == expected_value, (
            f"{video_name} 计数错误: actual={actual_value}, expected={expected_value}"
        )

    average_reward = sum(rewards) / len(rewards)
    _write_reward(average_reward)
    print(f"average_reward={average_reward:.4f}")
