from pathlib import Path

import openpyxl

OUTPUT_FILE = "growth_transfer_recovered.xlsx"
CHECKS = [('Budget by Directorate', 'B9', 7139, 0.0), ('Budget by Directorate', 'K10', 23285, 0.0), ('Budget by Directorate', 'E10', 6555, 0.0), ('YoY Changes (%)', 'B9', 2.27, 0.1), ('Directorate Shares (%)', 'B10', 31.35, 0.1), ('Growth Analysis', 'B8', 7444.4, 0.5), ('Growth Analysis', 'E4', 8.58, 0.1)]

def load_workbook_checked():
    path = Path(OUTPUT_FILE)
    assert path.exists(), f"missing output: {OUTPUT_FILE}"
    return openpyxl.load_workbook(path)

def assert_expected_cells(wb):
    for sheet_name, cell, expected, tolerance in CHECKS:
        actual = wb[sheet_name][cell].value
        if tolerance == 0:
            assert actual == expected, f"{sheet_name}!{cell} expected {expected}, got {actual}"
        else:
            assert abs(actual - expected) < tolerance, f"{sheet_name}!{cell} expected {expected}, got {actual}"

def assert_no_placeholders(wb):
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                assert value != "???", f"placeholder remains in {sheet_name}"

def assert_row_sum_consistency(wb):
    budget = wb["Budget by Directorate"]
    row5_sum = sum(budget.cell(row=5, column=col).value for col in range(2, 11))
    row10_sum = sum(budget.cell(row=10, column=col).value for col in range(2, 11))
    assert row5_sum == budget["K5"].value
    assert row10_sum == budget["K10"].value

def assert_yoy_formulas(wb):
    budget = wb["Budget by Directorate"]
    yoy = wb["YoY Changes (%)"]
    d7 = round((budget["D8"].value - budget["D7"].value) / budget["D7"].value * 100, 2)
    f9 = round((budget["F10"].value - budget["F9"].value) / budget["F9"].value * 100, 2)
    b9 = round((budget["B10"].value - budget["B9"].value) / budget["B9"].value * 100, 2)
    assert abs(yoy["D7"].value - d7) < 0.1
    assert abs(yoy["F9"].value - f9) < 0.1
    assert abs(yoy["B9"].value - b9) < 0.1

def assert_share_formulas(wb):
    budget = wb["Budget by Directorate"]
    shares = wb["Directorate Shares (%)"]
    f5 = round(budget["F5"].value / budget["K5"].value * 100, 2)
    b10 = round(budget["B10"].value / budget["K10"].value * 100, 2)
    assert abs(shares["F5"].value - f5) < 0.1
    assert abs(shares["B10"].value - b10) < 0.1

def assert_growth_cross_sheet(wb):
    budget = wb["Budget by Directorate"]
    growth = wb["Growth Analysis"]
    cagr = round(((budget["E13"].value / budget["E8"].value) ** 0.2 - 1) * 100, 2)
    assert abs(growth["E4"].value - cagr) < 0.1
    assert growth["E5"].value == budget["E8"].value

def main():
    wb = load_workbook_checked()
    assert_expected_cells(wb)
    assert_no_placeholders(wb)
    assert_row_sum_consistency(wb)
    assert_yoy_formulas(wb)
    assert_share_formulas(wb)
    assert_growth_cross_sheet(wb)

if __name__ == "__main__":
    main()
