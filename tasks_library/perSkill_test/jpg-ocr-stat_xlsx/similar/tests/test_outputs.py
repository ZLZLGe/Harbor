import os
from openpyxl import load_workbook

OUTPUT_PATH = "/app/workspace/similar.xlsx"
EXPECTED_ROWS = [
    ["filename", "date", "total_amount"],
    ["009.jpg", "2024-03-05", "12.50"],
    ["010.jpg", "2024-03-05", "19.90"],
    ["011.jpg", "2024-03-06", "7.00"],
    ["012.jpg", "", "15.00"],
    ["013.jpg", "2024-03-07", ""],
]


def norm(value):
    if value is None:
        return ""
    return str(value)


assert os.path.exists(OUTPUT_PATH), f"missing output: {OUTPUT_PATH}"
wb = load_workbook(OUTPUT_PATH, data_only=True)
try:
    assert wb.sheetnames == ["results"], f"unexpected sheets: {wb.sheetnames}"
    ws = wb["results"]
    rows = []
    for r in range(1, ws.max_row + 1):
        rows.append([norm(ws.cell(row=r, column=c).value) for c in range(1, 4)])
    assert rows == EXPECTED_ROWS, f"rows mismatch\nactual={rows}\nexpected={EXPECTED_ROWS}"
finally:
    wb.close()
