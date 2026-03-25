from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

OUTPUT_FILE = Path("/root/close_tracker_repaired.xlsx")
TEMPLATE_FILE = Path("/root/close_tracker_template.xlsx")


def normalized_blank(value):
    return None if value in ("", None) else value


def load_pair(data_only: bool):
    return load_workbook(TEMPLATE_FILE, data_only=data_only), load_workbook(
        OUTPUT_FILE, data_only=data_only
    )


def derive_expected_results():
    workbook = load_workbook(TEMPLATE_FILE, data_only=True)
    tracker = workbook["Close Tracker"]
    owner_map = workbook["Owner Map"]

    owner_emails = {}
    for owner_id, _, email, *_ in owner_map.iter_rows(min_row=2, values_only=True):
        if owner_id not in ("", None) and email not in ("", None):
            owner_emails[owner_id] = email

    tracker_expectations = {}
    status_counts = {"Late": 0, "On Time": 0, "Open": 0}
    escalation_count = 0
    max_delay = 0
    positive_delays = []

    for row in range(4, 9):
        owner_id = tracker[f"C{row}"].value
        planned_day = tracker[f"E{row}"].value
        actual_day = tracker[f"F{row}"].value

        if actual_day in ("", None):
            status = "Open"
            delay_days = None
        else:
            delay_days = max(actual_day - planned_day, 0)
            status = "On Time" if actual_day <= planned_day else "Late"
            max_delay = max(max_delay, delay_days)
            if delay_days > 0:
                positive_delays.append(delay_days)

        escalation = "Escalate" if status == "Late" and (delay_days or 0) >= 2 else ""
        status_counts[status] += 1
        if escalation:
            escalation_count += 1

        tracker_expectations[row] = (
            status,
            delay_days,
            owner_emails[owner_id],
            escalation,
        )

    average_delay = sum(positive_delays) / len(positive_delays) if positive_delays else 0
    dashboard_expectations = {
        "B2": len(tracker_expectations),
        "B3": status_counts["Late"],
        "B4": status_counts["On Time"],
        "B5": status_counts["Open"],
        "B6": escalation_count,
        "B7": max_delay,
        "B8": average_delay,
    }

    workbook.close()
    return tracker_expectations, dashboard_expectations


def test_output_exists():
    assert OUTPUT_FILE.exists(), f"Missing output file: {OUTPUT_FILE}"


def test_sheet_names_and_order_match_template():
    template, output = load_pair(data_only=False)
    assert output.sheetnames == ["Close Tracker", "Owner Map", "Dashboard"]
    assert output.sheetnames == template.sheetnames
    template.close()
    output.close()


def test_layout_and_structure_are_preserved():
    template, output = load_pair(data_only=False)
    for sheet_name in template.sheetnames:
        template_ws = template[sheet_name]
        output_ws = output[sheet_name]

        assert output_ws.freeze_panes == template_ws.freeze_panes
        assert [str(rng) for rng in output_ws.merged_cells.ranges] == [
            str(rng) for rng in template_ws.merged_cells.ranges
        ]

        for column in "ABCDEFGHIJ" if sheet_name == "Close Tracker" else "ABCD":
            assert (
                output_ws.column_dimensions[column].width
                == template_ws.column_dimensions[column].width
            )

        for row in (1, 2, 3, 4, 8):
            assert output_ws.row_dimensions[row].height == template_ws.row_dimensions[row].height

    template.close()
    output.close()


def test_key_styles_and_owner_map_are_unchanged():
    template, output = load_pair(data_only=False)

    for cell_ref in ("A1", "A3", "B3", "C3", "J3"):
        assert output["Close Tracker"][cell_ref]._style == template["Close Tracker"][cell_ref]._style

    for cell_ref in ("A1", "A2", "A8", "B1"):
        assert output["Dashboard"][cell_ref]._style == template["Dashboard"][cell_ref]._style

    template_owner_rows = list(template["Owner Map"].iter_rows(values_only=True))
    output_owner_rows = list(output["Owner Map"].iter_rows(values_only=True))
    assert output_owner_rows == template_owner_rows

    template.close()
    output.close()


def test_manual_input_cells_are_unchanged():
    template, output = load_pair(data_only=True)
    for row in range(4, 9):
        template_values = [template["Close Tracker"][f"{column}{row}"].value for column in "ABCDEF"]
        output_values = [output["Close Tracker"][f"{column}{row}"].value for column in "ABCDEF"]
        assert output_values == template_values
    template.close()
    output.close()


def test_required_formula_cells_remain_formulas():
    workbook = load_workbook(OUTPUT_FILE, data_only=False)
    tracker = workbook["Close Tracker"]
    dashboard = workbook["Dashboard"]

    for row in range(4, 9):
        status_formula = tracker[f"G{row}"].value
        delay_formula = tracker[f"H{row}"].value
        email_formula = tracker[f"I{row}"].value
        escalation_formula = tracker[f"J{row}"].value

        assert isinstance(status_formula, str) and status_formula.startswith("=")
        assert f"E{row}" in status_formula and f"F{row}" in status_formula

        assert isinstance(delay_formula, str) and delay_formula.startswith("=")
        assert f"E{row}" in delay_formula and f"F{row}" in delay_formula

        assert isinstance(email_formula, str) and email_formula.startswith("=")
        assert "Owner Map" in email_formula and f"C{row}" in email_formula

        assert isinstance(escalation_formula, str) and escalation_formula.startswith("=")
        assert f"G{row}" in escalation_formula and f"H{row}" in escalation_formula

    for cell_ref in ("B2", "B3", "B4", "B5", "B6", "B7", "B8"):
        formula = dashboard[cell_ref].value
        assert isinstance(formula, str) and formula.startswith("=")
        assert "Close Tracker" in formula

    workbook.close()


def test_tracker_results_match_expected():
    expected_tracker, _ = derive_expected_results()
    workbook = load_workbook(OUTPUT_FILE, data_only=True)
    tracker = workbook["Close Tracker"]

    for row, expected in expected_tracker.items():
        status, delay_days, owner_email, escalation = expected
        assert tracker[f"G{row}"].value == status
        assert normalized_blank(tracker[f"H{row}"].value) == delay_days
        assert tracker[f"I{row}"].value == owner_email
        assert normalized_blank(tracker[f"J{row}"].value) == normalized_blank(escalation)

    workbook.close()


def test_dashboard_results_match_expected():
    _, expected_dashboard = derive_expected_results()
    workbook = load_workbook(OUTPUT_FILE, data_only=True)
    dashboard = workbook["Dashboard"]

    for cell_ref, expected in expected_dashboard.items():
        assert dashboard[cell_ref].value == expected

    workbook.close()


def test_no_visible_formula_errors():
    workbook = load_workbook(OUTPUT_FILE, data_only=True)
    error_prefixes = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    assert not cell.value.startswith(error_prefixes), (
                        f"Formula error at {sheet.title}!{cell.coordinate}: {cell.value}"
                    )

    workbook.close()
