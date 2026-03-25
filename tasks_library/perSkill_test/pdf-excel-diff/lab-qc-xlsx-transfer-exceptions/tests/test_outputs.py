import json
import statistics
from collections import Counter, defaultdict
from pathlib import Path

import pytest
from openpyxl import load_workbook

OUTPUT_FILE = Path("/root/lab_qc_exceptions.json")
INPUT_FILE = Path("/root/lab_qc_runbook.xlsx")


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def load_table(workbook, sheet_name, required_columns):
    ws = workbook[sheet_name]
    header_row = None
    header_map = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        normalized = [clean_text(cell) for cell in row]
        if all(column in normalized for column in required_columns):
            header_row = row_idx
            header_map = {column: normalized.index(column) for column in required_columns}
            break

    if header_row is None or header_map is None:
        raise AssertionError(f"Header row not found in {sheet_name}")

    records = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        values = []
        for column in required_columns:
            idx = header_map[column]
            values.append(row[idx] if idx < len(row) else None)
        if all(value is None or clean_text(value) == "" for value in values):
            continue
        records.append(dict(zip(required_columns, values)))
    return records


def expected_output():
    workbook = load_workbook(INPUT_FILE, data_only=True)

    batch_rows = load_table(
        workbook,
        "Batch Summary",
        ["batch_id", "panel", "duplicate_variance_limit"],
    )
    plate_rows = load_table(
        workbook,
        "Plate Map",
        ["batch_id", "well", "entry_type", "sample_id", "control_code", "duplicate_group"],
    )
    reading_rows = load_table(workbook, "Readings", ["batch_id", "well", "signal"])
    limit_rows = load_table(workbook, "QC Limits", ["control_code", "min_signal", "max_signal"])
    intake_rows = load_table(workbook, "Specimen Intake", ["intake_id", "sample_id", "patient_id"])

    reading_lookup = {
        (clean_text(row["batch_id"]), clean_text(row["well"])): float(row["signal"])
        for row in reading_rows
    }
    limit_lookup = {
        clean_text(row["control_code"]): (
            float(row["min_signal"]),
            float(row["max_signal"]),
        )
        for row in limit_rows
    }
    variance_limit_lookup = {
        clean_text(row["batch_id"]): float(row["duplicate_variance_limit"])
        for row in batch_rows
    }

    failed_control_wells = []
    duplicate_groups = defaultdict(list)

    for row in plate_rows:
        batch_id = clean_text(row["batch_id"])
        well = clean_text(row["well"])
        entry_type = clean_text(row["entry_type"])

        if entry_type == "CONTROL":
            control_code = clean_text(row["control_code"])
            signal = reading_lookup[(batch_id, well)]
            min_signal, max_signal = limit_lookup[control_code]
            if not (min_signal <= signal <= max_signal):
                failed_control_wells.append(
                    {
                        "batch_id": batch_id,
                        "well": well,
                        "control_code": control_code,
                        "signal": signal,
                    }
                )

        if entry_type == "DUPLICATE":
            group_name = clean_text(row["duplicate_group"])
            if group_name:
                duplicate_groups[(batch_id, group_name)].append(
                    {
                        "well": well,
                        "sample_id": clean_text(row["sample_id"]),
                    }
                )

    intake_counts = Counter()
    for row in intake_rows:
        sample_id = clean_text(row["sample_id"])
        if sample_id:
            intake_counts[sample_id] += 1

    high_variance_batches = []
    for (batch_id, group_name), members in duplicate_groups.items():
        signals = [reading_lookup[(batch_id, member["well"])] for member in members]
        variance = round(statistics.pvariance(signals), 2)
        variance_limit = variance_limit_lookup[batch_id]
        if variance > variance_limit:
            high_variance_batches.append(
                {
                    "batch_id": batch_id,
                    "duplicate_group": group_name,
                    "sample_id": members[0]["sample_id"],
                    "variance": variance,
                    "variance_limit": variance_limit,
                }
            )

    return {
        "failed_control_wells": sorted(
            failed_control_wells, key=lambda item: (item["batch_id"], item["well"])
        ),
        "duplicate_sample_ids": sorted(
            sample_id for sample_id, count in intake_counts.items() if count > 1
        ),
        "high_variance_batches": sorted(
            high_variance_batches,
            key=lambda item: (item["batch_id"], item["duplicate_group"]),
        ),
    }


@pytest.fixture(scope="module")
def output_data():
    assert OUTPUT_FILE.exists(), f"Missing output file: {OUTPUT_FILE}"
    with OUTPUT_FILE.open(encoding="utf-8") as handle:
        return json.load(handle)


def test_output_is_valid_json(output_data):
    assert isinstance(output_data, dict)


def test_output_has_exact_top_level_keys(output_data):
    assert set(output_data.keys()) == {
        "failed_control_wells",
        "duplicate_sample_ids",
        "high_variance_batches",
    }


def test_failed_control_wells(output_data):
    expected = expected_output()["failed_control_wells"]
    assert output_data["failed_control_wells"] == expected


def test_duplicate_sample_ids(output_data):
    expected = expected_output()["duplicate_sample_ids"]
    assert output_data["duplicate_sample_ids"] == expected


def test_high_variance_batches(output_data):
    expected = expected_output()["high_variance_batches"]
    assert output_data["high_variance_batches"] == expected
