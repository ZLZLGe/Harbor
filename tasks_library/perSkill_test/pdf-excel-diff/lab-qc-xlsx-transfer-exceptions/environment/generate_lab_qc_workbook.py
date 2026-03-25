from openpyxl import Workbook
from openpyxl.styles import Font

OUTPUT_PATH = "/root/lab_qc_runbook.xlsx"


def write_sheet(ws, title, header_row, headers, rows, notes):
    ws.title = title
    ws["A1"] = notes[0]
    ws["A1"].font = Font(bold=True)
    if len(notes) > 1:
        for idx, note in enumerate(notes[1:], start=2):
            ws.cell(row=idx, column=1, value=note)

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header).font = Font(bold=True)

    for row_offset, values in enumerate(rows, start=1):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=header_row + row_offset, column=col_idx, value=value)


def main():
    wb = Workbook()

    batch_summary = wb.active
    write_sheet(
        batch_summary,
        "Batch Summary",
        3,
        ["batch_id", "panel", "duplicate_variance_limit"],
        [
            ["B-100", "Chemistry", 4.0],
            ["B-101", "Immuno", 2.5],
            ["B-102", "Serology", 1.5],
        ],
        [
            "March analytical QC batch settings",
            "Use duplicate_variance_limit when reviewing duplicate groups.",
        ],
    )

    plate_map = wb.create_sheet()
    write_sheet(
        plate_map,
        "Plate Map",
        2,
        ["batch_id", "well", "entry_type", "sample_id", "control_code", "duplicate_group"],
        [
            ["B-100", "A01", "CONTROL", None, "NEG_CTRL", None],
            ["B-100", "A02", "CONTROL", None, "POS_CTRL", None],
            ["B-100", "B01", "SAMPLE", "S-1001", None, None],
            ["B-100", "B02", "DUPLICATE", "S-1002", None, "DG-100-A"],
            ["B-100", "B03", "DUPLICATE", "S-1002", None, "DG-100-A"],
            ["B-100", "B04", "SAMPLE", "S-1003", None, None],
            ["B-101", "A01", "CONTROL", None, "NEG_CTRL", None],
            ["B-101", "A02", "CONTROL", None, "POS_CTRL", None],
            ["B-101", "B01", "SAMPLE", "S-1004", None, None],
            ["B-101", "B02", "DUPLICATE", "S-1005", None, "DG-101-A"],
            ["B-101", "B03", "DUPLICATE", "S-1005", None, "DG-101-A"],
            ["B-101", "B04", "DUPLICATE", "S-1006", None, "DG-101-B"],
            ["B-101", "B05", "DUPLICATE", "S-1006", None, "DG-101-B"],
            ["B-102", "A01", "CONTROL", None, "NEG_CTRL", None],
            ["B-102", "A02", "CONTROL", None, "POS_CTRL", None],
            ["B-102", "B01", "SAMPLE", "S-1007", None, None],
            ["B-102", "B02", "DUPLICATE", "S-1008", None, "DG-102-A"],
            ["B-102", "B03", "DUPLICATE", "S-1008", None, "DG-102-A"],
        ],
        [
            "Plate layout for the current validation run",
        ],
    )

    readings = wb.create_sheet()
    write_sheet(
        readings,
        "Readings",
        4,
        ["batch_id", "well", "signal"],
        [
            ["B-100", "A01", 0.08],
            ["B-100", "A02", 1.32],
            ["B-100", "B01", 0.57],
            ["B-100", "B02", 4.0],
            ["B-100", "B03", 9.0],
            ["B-100", "B04", 0.61],
            ["B-101", "A01", 0.22],
            ["B-101", "A02", 1.18],
            ["B-101", "B01", 0.49],
            ["B-101", "B02", 7.0],
            ["B-101", "B03", 9.0],
            ["B-101", "B04", 2.0],
            ["B-101", "B05", 5.4],
            ["B-102", "A01", 0.11],
            ["B-102", "A02", 0.92],
            ["B-102", "B01", 0.71],
            ["B-102", "B02", 5.0],
            ["B-102", "B03", 6.0],
        ],
        [
            "Instrument output exported after plate read",
            "Signals are already baseline corrected.",
            "Match rows using batch_id and well.",
        ],
    )

    qc_limits = wb.create_sheet()
    write_sheet(
        qc_limits,
        "QC Limits",
        2,
        ["control_code", "min_signal", "max_signal"],
        [
            ["NEG_CTRL", 0.0, 0.15],
            ["POS_CTRL", 1.1, 1.5],
        ],
        [
            "Approved control ranges for this assay family",
        ],
    )

    specimen_intake = wb.create_sheet()
    write_sheet(
        specimen_intake,
        "Specimen Intake",
        3,
        ["intake_id", "sample_id", "patient_id"],
        [
            ["I-001", "S-1001", "P-01"],
            ["I-002", " S-2001 ", "P-02"],
            ["I-003", "S-2002", "P-03"],
            ["I-004", "S-2001", "P-04"],
            ["I-005", "S-3001", "P-05"],
            ["I-006", "S-4001", "P-06"],
            ["I-007", " S-4001 ", "P-07"],
            ["I-008", None, "P-08"],
            ["I-009", "S-5001", "P-09"],
        ],
        [
            "Daily accession log used before plating",
            "Blank sample IDs should be ignored in duplicate checks.",
        ],
    )

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for column in ("A", "B", "C", "D", "E", "F"):
            ws.column_dimensions[column].width = 20

    wb.save(OUTPUT_PATH)


if __name__ == "__main__":
    main()
