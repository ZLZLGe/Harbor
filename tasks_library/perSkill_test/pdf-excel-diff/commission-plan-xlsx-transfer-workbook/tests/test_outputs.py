from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

OUTPUT_FILE = Path("/root/commission_payouts.xlsx")
EXPECTED_SHEETS = ["Quota Plan", "Commission Detail", "Team Overview"]

EXPECTED_QUOTA = [
    ("R-101", "Alice Wong", "North", 50000, 0.04, 0.055),
    ("R-102", "Ben Ibarra", "North", 45000, 0.035, 0.05),
    ("R-201", "Cara Singh", "South", 60000, 0.045, 0.06),
    ("R-301", "Diego Mora", "West", 40000, 0.03, 0.045),
]

EXPECTED_DETAIL = {
    "S-001": ("2026-01", "R-101", "Alice Wong", "North", "Pine Retail", 18000, 50000, 40000, 0.8, 0.04, 720.0),
    "S-002": ("2026-01", "R-101", "Alice Wong", "North", "Harbor Labs", 22000, 50000, 40000, 0.8, 0.04, 880.0),
    "S-003": ("2026-01", "R-102", "Ben Ibarra", "North", "Metro Foods", 26000, 45000, 56000, 1.2444444444444445, 0.05, 1300.0),
    "S-004": ("2026-01", "R-102", "Ben Ibarra", "North", "Solstice Cafe", 30000, 45000, 56000, 1.2444444444444445, 0.05, 1500.0),
    "S-005": ("2026-01", "R-201", "Cara Singh", "South", "Cloud Frame", 25000, 60000, 43000, 0.7166666666666667, 0.045, 1125.0),
    "S-006": ("2026-01", "R-201", "Cara Singh", "South", "Apex Health", 18000, 60000, 43000, 0.7166666666666667, 0.045, 810.0),
    "S-007": ("2026-01", "R-301", "Diego Mora", "West", "Northwind Schools", 12000, 40000, 26000, 0.65, 0.03, 360.0),
    "S-008": ("2026-01", "R-301", "Diego Mora", "West", "Juniper Hotels", 14000, 40000, 26000, 0.65, 0.03, 420.0),
}

EXPECTED_SUMMARY = {
    "North": (95000, 96000, 4400.0, 1.0105263157894737),
    "South": (60000, 43000, 1935.0, 0.7166666666666667),
    "West": (40000, 26000, 780.0, 0.65),
}

DETAIL_HEADERS = [
    "sale_id",
    "sale_month",
    "rep_id",
    "rep_name",
    "team",
    "client",
    "net_revenue",
    "monthly_quota",
    "monthly_revenue",
    "attainment",
    "applied_rate",
    "commission",
]
SUMMARY_HEADERS = ["team", "quota_total", "total_revenue", "total_commission", "attainment"]
QUOTA_HEADERS = ["rep_id", "rep_name", "team", "monthly_quota", "standard_rate", "stretch_rate"]


def approx_equal(actual, expected, tol=1e-9):
    return abs(float(actual) - float(expected)) <= tol


def load_books():
    formulas = load_workbook(OUTPUT_FILE, data_only=False)
    values = load_workbook(OUTPUT_FILE, data_only=True)
    return formulas, values


def test_output_exists():
    assert OUTPUT_FILE.exists(), f"Missing output file: {OUTPUT_FILE}"


def test_sheet_names_and_order():
    workbook = load_workbook(OUTPUT_FILE, data_only=False)
    assert workbook.sheetnames == EXPECTED_SHEETS
    workbook.close()


def test_headers_and_quota_copy():
    workbook = load_workbook(OUTPUT_FILE, data_only=True)
    quota_sheet = workbook["Quota Plan"]
    assert [cell.value for cell in quota_sheet[1]] == QUOTA_HEADERS

    rows = list(quota_sheet.iter_rows(min_row=2, max_row=5, values_only=True))
    assert rows == EXPECTED_QUOTA
    workbook.close()


def test_detail_headers_and_sort_order():
    workbook = load_workbook(OUTPUT_FILE, data_only=True)
    detail_sheet = workbook["Commission Detail"]
    assert [cell.value for cell in detail_sheet[1]] == DETAIL_HEADERS
    sale_ids = [row[0] for row in detail_sheet.iter_rows(min_row=2, max_row=9, values_only=True)]
    assert sale_ids == sorted(EXPECTED_DETAIL)
    workbook.close()


def test_required_detail_formula_cells():
    workbook = load_workbook(OUTPUT_FILE, data_only=False)
    detail_sheet = workbook["Commission Detail"]
    for row in range(2, 10):
        for column in ("D", "E", "H", "I", "J", "K", "L"):
            value = detail_sheet[f"{column}{row}"].value
            assert isinstance(value, str) and value.startswith("=")
        assert "Quota Plan" in detail_sheet[f"D{row}"].value
        assert "Quota Plan" in detail_sheet[f"E{row}"].value
        assert "Quota Plan" in detail_sheet[f"H{row}"].value
        assert "Quota Plan" in detail_sheet[f"K{row}"].value
    workbook.close()


def test_required_summary_formula_cells():
    workbook = load_workbook(OUTPUT_FILE, data_only=False)
    summary_sheet = workbook["Team Overview"]
    assert [cell.value for cell in summary_sheet[1]] == SUMMARY_HEADERS
    for row in range(2, 5):
        for column in ("B", "C", "D", "E"):
            value = summary_sheet[f"{column}{row}"].value
            assert isinstance(value, str) and value.startswith("=")
        assert "Quota Plan" in summary_sheet[f"B{row}"].value
        assert "Commission Detail" in summary_sheet[f"C{row}"].value
        assert "Commission Detail" in summary_sheet[f"D{row}"].value
    workbook.close()


def test_detail_values_after_recalc():
    workbook = load_workbook(OUTPUT_FILE, data_only=True)
    detail_sheet = workbook["Commission Detail"]
    for row in detail_sheet.iter_rows(min_row=2, max_row=9, values_only=True):
        sale_id = row[0]
        expected = EXPECTED_DETAIL[sale_id]
        assert row[1] == expected[0]
        assert row[2] == expected[1]
        assert row[3] == expected[2]
        assert row[4] == expected[3]
        assert row[5] == expected[4]
        assert row[6] == expected[5]
        assert row[7] == expected[6]
        assert row[8] == expected[7]
        assert approx_equal(row[9], expected[8], 1e-12)
        assert approx_equal(row[10], expected[9], 1e-12)
        assert approx_equal(row[11], expected[10], 1e-9)
    workbook.close()


def test_summary_values_after_recalc():
    workbook = load_workbook(OUTPUT_FILE, data_only=True)
    summary_sheet = workbook["Team Overview"]
    rows = list(summary_sheet.iter_rows(min_row=2, max_row=4, values_only=True))
    assert [row[0] for row in rows] == sorted(EXPECTED_SUMMARY)
    for team, quota_total, total_revenue, total_commission, attainment in rows:
        expected = EXPECTED_SUMMARY[team]
        assert quota_total == expected[0]
        assert total_revenue == expected[1]
        assert approx_equal(total_commission, expected[2], 1e-9)
        assert approx_equal(attainment, expected[3], 1e-12)
    workbook.close()


def test_no_formula_errors_in_visible_values():
    workbook = load_workbook(OUTPUT_FILE, data_only=True)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    assert not cell.value.startswith("#"), f"Formula error at {sheet.title}!{cell.coordinate}: {cell.value}"
    workbook.close()
