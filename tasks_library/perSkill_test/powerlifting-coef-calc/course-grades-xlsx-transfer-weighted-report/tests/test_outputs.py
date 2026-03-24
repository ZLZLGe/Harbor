from decimal import Decimal, ROUND_HALF_UP

import openpyxl
import pytest

OUTPUT_FILE = "/root/data/course_weighted_report.xlsx"

EXPECTED_HEADERS = [
    "StudentID",
    "Section",
    "LastName",
    "FirstName",
    "Quiz1",
    "Quiz2",
    "Quiz3",
    "Quiz4",
    "DroppedQuiz",
    "QuizAverage",
    "Lab1",
    "Lab2",
    "LabAverage",
    "Midterm",
    "FinalExam",
    "QuizWeighted",
    "LabWeighted",
    "MidtermWeighted",
    "FinalWeighted",
    "TotalScore",
    "LetterGrade",
]

EXPECTED_ROSTER = [
    ("S1001", "01", "Alvarez", "Mia"),
    ("S1002", "01", "Brooks", "Evan"),
    ("S1003", "02", "Chen", "Lila"),
    ("S1004", "02", "Davis", "Omar"),
    ("S1005", "01", "Patel", "Sana"),
    ("S1006", "03", "Young", "Noah"),
]

EXPECTED_SCORES = [
    ("S1003", 88, 92, 84, 90, 93, 95, 87, 91),
    ("S1001", 95, 78, 88, 92, 100, 94, 90, 93),
    ("S1006", 42, 55, 58, 50, 60, 57, 56, 54),
    ("S1004", 72, 70, 68, 74, 75, 78, 73, 71),
    ("S1002", 83, 85, 81, 79, 88, 84, 80, 82),
    ("S1005", 50, 65, 70, 68, 72, 74, 69, 66),
]

EXPECTED_WEIGHTS = [
    ("Quiz", 0.25),
    ("Lab", 0.20),
    ("Midterm", 0.25),
    ("FinalExam", 0.30),
]


def rounded(value: float) -> float:
    return float(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def read_rows(ws):
    return [
        tuple(ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1))
        for row_idx in range(1, ws.max_row + 1)
    ]


def expected_report_rows():
    weights = {category: weight for category, weight in EXPECTED_WEIGHTS}
    scores_map = {
        row[0]: {
            "quiz": [row[1], row[2], row[3], row[4]],
            "labs": [row[5], row[6]],
            "midterm": row[7],
            "final": row[8],
        }
        for row in EXPECTED_SCORES
    }

    rows = []
    for student_id, section, last_name, first_name in EXPECTED_ROSTER:
        record = scores_map[student_id]
        dropped_quiz = min(record["quiz"])
        quiz_average = rounded((sum(record["quiz"]) - dropped_quiz) / 3)
        lab_average = rounded(sum(record["labs"]) / 2)
        quiz_weighted = rounded(quiz_average * weights["Quiz"])
        lab_weighted = rounded(lab_average * weights["Lab"])
        midterm_weighted = rounded(record["midterm"] * weights["Midterm"])
        final_weighted = rounded(record["final"] * weights["FinalExam"])
        total_score = rounded(quiz_weighted + lab_weighted + midterm_weighted + final_weighted)

        if total_score >= 90:
            letter_grade = "A"
        elif total_score >= 80:
            letter_grade = "B"
        elif total_score >= 70:
            letter_grade = "C"
        elif total_score >= 60:
            letter_grade = "D"
        else:
            letter_grade = "F"

        rows.append(
            (
                student_id,
                section,
                last_name,
                first_name,
                *record["quiz"],
                dropped_quiz,
                quiz_average,
                *record["labs"],
                lab_average,
                record["midterm"],
                record["final"],
                quiz_weighted,
                lab_weighted,
                midterm_weighted,
                final_weighted,
                total_score,
                letter_grade,
            )
        )

    return rows


@pytest.fixture(scope="module")
def workbook_formula():
    return openpyxl.load_workbook(OUTPUT_FILE, data_only=False)


@pytest.fixture(scope="module")
def workbook_values():
    return openpyxl.load_workbook(OUTPUT_FILE, data_only=True)


def test_required_sheets_exist(workbook_formula):
    assert workbook_formula.sheetnames == ["Roster", "Scores", "Weights", "Report"]


def test_source_sheets_unchanged(workbook_values):
    roster_ws = workbook_values["Roster"]
    scores_ws = workbook_values["Scores"]
    weights_ws = workbook_values["Weights"]

    assert read_rows(roster_ws) == [
        ("StudentID", "Section", "LastName", "FirstName"),
        *EXPECTED_ROSTER,
    ]
    assert read_rows(scores_ws) == [
        ("StudentID", "Quiz1", "Quiz2", "Quiz3", "Quiz4", "Lab1", "Lab2", "Midterm", "FinalExam"),
        *EXPECTED_SCORES,
    ]
    assert read_rows(weights_ws) == [
        ("Category", "Weight"),
        *EXPECTED_WEIGHTS,
    ]


def test_report_headers(workbook_values):
    ws = workbook_values["Report"]
    headers = [ws.cell(row=1, column=idx).value for idx in range(1, 22)]
    assert headers == EXPECTED_HEADERS


def test_report_row_count_matches_roster(workbook_values):
    roster_ws = workbook_values["Roster"]
    report_ws = workbook_values["Report"]
    assert report_ws.max_row == roster_ws.max_row


def test_report_uses_formulas(workbook_formula):
    ws = workbook_formula["Report"]

    assert ws["A2"].value == "=Roster!A2"
    assert isinstance(ws["E2"].value, str) and "INDEX(" in ws["E2"].value and "MATCH(" in ws["E2"].value
    assert ws["I2"].value == "=MIN(E2:H2)"
    assert isinstance(ws["J2"].value, str) and "SUM(E2:H2)" in ws["J2"].value and "COUNT(E2:H2)-1" in ws["J2"].value
    assert isinstance(ws["M2"].value, str) and ws["M2"].value == "=ROUND(AVERAGE(K2:L2),2)"
    assert isinstance(ws["P2"].value, str) and 'MATCH("Quiz"' in ws["P2"].value and "Weights!" in ws["P2"].value
    assert isinstance(ws["Q2"].value, str) and 'MATCH("Lab"' in ws["Q2"].value and "Weights!" in ws["Q2"].value
    assert isinstance(ws["R2"].value, str) and 'MATCH("Midterm"' in ws["R2"].value
    assert isinstance(ws["S2"].value, str) and 'MATCH("FinalExam"' in ws["S2"].value
    assert ws["T2"].value == "=ROUND(SUM(P2:S2),2)"
    assert isinstance(ws["U2"].value, str) and 'IF(T2>=90,"A"' in ws["U2"].value


def test_report_values_match_expected(workbook_values):
    ws = workbook_values["Report"]
    actual_rows = []
    for row_idx in range(2, ws.max_row + 1):
        actual_rows.append(
            tuple(ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 22))
        )

    assert actual_rows == expected_report_rows()


def test_grade_distribution(workbook_values):
    ws = workbook_values["Report"]
    grades = [ws[f"U{row_idx}"].value for row_idx in range(2, ws.max_row + 1)]
    assert grades == ["A", "B", "A", "C", "D", "F"]


def test_lowest_quiz_is_dropped(workbook_values):
    ws = workbook_values["Report"]
    for row_idx in range(2, ws.max_row + 1):
        quizzes = [ws[f"{column}{row_idx}"].value for column in ("E", "F", "G", "H")]
        assert ws[f"I{row_idx}"].value == min(quizzes)
