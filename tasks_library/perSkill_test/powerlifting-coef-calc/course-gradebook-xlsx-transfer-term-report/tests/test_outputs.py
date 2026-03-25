from pathlib import Path
from shutil import copyfile
import subprocess

from openpyxl import load_workbook

OUTPUT_FILE = Path("/root/data/term_grade_report.xlsx")
EXPECTED_HEADERS = [
    "StudentID",
    "StudentName",
    "Homework",
    "Midterm",
    "Project",
    "FinalExam",
    "WeightedTotal",
    "MissingFlag",
    "LetterGrade",
    "PassStatus",
]


def normalize_blank(value):
    if value in ("", None):
        return None
    return value


def choose_band(total, bands):
    matched = bands[0]
    for band in bands:
        if total >= band["MinScore"]:
            matched = band
        else:
            break
    return matched


def build_expected_rows(wb_values):
    scores_ws = wb_values["Scores"]
    weights_ws = wb_values["Weights"]
    bands_ws = wb_values["GradeBands"]

    weights = [float(weights_ws.cell(row=2, column=col).value) for col in range(1, 5)]
    grade_bands = []
    for row_idx in range(2, bands_ws.max_row + 1):
        grade_bands.append(
            {
                "MinScore": float(bands_ws[f"A{row_idx}"].value),
                "LetterGrade": bands_ws[f"B{row_idx}"].value,
                "PassStatus": bands_ws[f"C{row_idx}"].value,
            }
        )

    rows = []
    for row_idx in range(2, scores_ws.max_row + 1):
        scores = [normalize_blank(scores_ws.cell(row=row_idx, column=col).value) for col in range(3, 7)]
        weighted_total = round(
            sum((float(score) if score is not None else 0.0) * weight for score, weight in zip(scores, weights)),
            2,
        )
        band = choose_band(weighted_total, grade_bands)
        rows.append(
            {
                "StudentID": scores_ws[f"A{row_idx}"].value,
                "StudentName": scores_ws[f"B{row_idx}"].value,
                "Homework": scores[0],
                "Midterm": scores[1],
                "Project": scores[2],
                "FinalExam": scores[3],
                "WeightedTotal": weighted_total,
                "MissingFlag": "MISSING" if any(score is None for score in scores) else "OK",
                "LetterGrade": band["LetterGrade"],
                "PassStatus": band["PassStatus"],
            }
        )
    return rows


def assert_report_matches_expected(workbook_path):
    wb_values = load_workbook(workbook_path, data_only=True)
    report_ws = wb_values["Report"]
    expected_rows = build_expected_rows(wb_values)

    for row_idx, expected in enumerate(expected_rows, start=2):
        assert report_ws[f"A{row_idx}"].value == expected["StudentID"]
        assert report_ws[f"B{row_idx}"].value == expected["StudentName"]
        assert normalize_blank(report_ws[f"C{row_idx}"].value) == expected["Homework"]
        assert normalize_blank(report_ws[f"D{row_idx}"].value) == expected["Midterm"]
        assert normalize_blank(report_ws[f"E{row_idx}"].value) == expected["Project"]
        assert normalize_blank(report_ws[f"F{row_idx}"].value) == expected["FinalExam"]
        assert round(float(report_ws[f"G{row_idx}"].value), 2) == expected["WeightedTotal"]
        assert report_ws[f"H{row_idx}"].value == expected["MissingFlag"]
        assert report_ws[f"I{row_idx}"].value == expected["LetterGrade"]
        assert report_ws[f"J{row_idx}"].value == expected["PassStatus"]


def test_output_file_exists():
    assert OUTPUT_FILE.exists(), f"Missing output file: {OUTPUT_FILE}"


def test_workbook_structure_and_headers():
    wb = load_workbook(OUTPUT_FILE, data_only=False)
    assert wb.sheetnames == ["Scores", "Weights", "GradeBands", "Report"]

    report_ws = wb["Report"]
    headers = [report_ws.cell(row=1, column=col).value for col in range(1, len(EXPECTED_HEADERS) + 1)]
    assert headers == EXPECTED_HEADERS

    scores_ws = wb["Scores"]
    assert report_ws.max_row == scores_ws.max_row


def test_report_uses_formulas_for_all_rows():
    wb = load_workbook(OUTPUT_FILE, data_only=False)
    report_ws = wb["Report"]

    for row_idx in range(2, report_ws.max_row + 1):
        for col in "ABCDEFGHIJ":
            value = report_ws[f"{col}{row_idx}"].value
            assert isinstance(value, str) and value.startswith("="), f"{col}{row_idx} should contain a formula"

    assert "Weights!" in report_ws["G2"].value
    assert "GradeBands!" in report_ws["I2"].value
    assert "GradeBands!" in report_ws["J2"].value


def test_cached_values_match_current_rules():
    assert_report_matches_expected(OUTPUT_FILE)


def test_report_reacts_to_rule_changes(tmp_path):
    temp_file = tmp_path / "term_grade_report_modified.xlsx"
    copyfile(OUTPUT_FILE, temp_file)

    wb = load_workbook(temp_file)
    weights_ws = wb["Weights"]
    bands_ws = wb["GradeBands"]

    for col_idx, value in enumerate([0.10, 0.20, 0.30, 0.40], start=1):
        weights_ws.cell(row=2, column=col_idx, value=value)

    bands_ws["A6"] = 92
    bands_ws["C3"] = "Fail"
    wb.save(temp_file)

    subprocess.run(
        ["python3", "/root/.codex/skills/xlsx/recalc.py", str(temp_file), "60"],
        check=True,
        capture_output=True,
        text=True,
    )

    assert_report_matches_expected(temp_file)
