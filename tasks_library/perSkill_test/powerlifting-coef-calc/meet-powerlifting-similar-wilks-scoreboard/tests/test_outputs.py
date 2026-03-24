from pathlib import Path

import openpyxl

INPUT_FILE = Path("/root/data/meet_results.xlsx")
OUTPUT_FILE = Path("/root/data/wilks_scoreboard.xlsx")

REQUIRED_HEADERS = [
    "Name",
    "Sex",
    "BodyweightKg",
    "Best3SquatKg",
    "Best3BenchKg",
    "Best3DeadliftKg",
]
EXPECTED_HEADERS = REQUIRED_HEADERS + ["TotalKg", "Wilks"]


def calculate_wilks(sex: str, bodyweight: float, total: float) -> float:
    if sex == "M":
        x = max(40.0, min(201.9, bodyweight))
        a, b, c, d, e, f = (
            -216.0475144,
            16.2606339,
            -0.002388645,
            -0.00113732,
            7.01863e-06,
            -1.291e-08,
        )
    else:
        x = max(26.51, min(154.53, bodyweight))
        a, b, c, d, e, f = (
            594.31747775582,
            -27.23842536447,
            0.82112226871,
            -0.00930733913,
            4.731582e-05,
            -9.054e-08,
        )

    coefficient = 500 / (a + b * x + c * x**2 + d * x**3 + e * x**4 + f * x**5)
    return round(total * coefficient, 3)


def normalize(value):
    if isinstance(value, float):
        return round(value, 3)
    return value


def read_sheet_rows(sheet):
    return [
        [cell.value for cell in row]
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column)
    ]


def load_input_records():
    workbook = openpyxl.load_workbook(INPUT_FILE, data_only=False)
    sheet = workbook["Meet Results"]
    headers = [cell.value for cell in sheet[1]]
    index = {header: idx for idx, header in enumerate(headers)}

    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {header: row[idx] for header, idx in index.items()}
        total = round(
            float(record["Best3SquatKg"])
            + float(record["Best3BenchKg"])
            + float(record["Best3DeadliftKg"]),
            3,
        )
        score = calculate_wilks(record["Sex"], float(record["BodyweightKg"]), total)
        record["TotalKg"] = total
        record["WilksScore"] = score
        records.append(record)

    return sorted(records, key=lambda item: item["WilksScore"], reverse=True)


class TestWorkbook:
    def test_output_exists(self):
        assert OUTPUT_FILE.exists(), "缺少输出文件 /root/data/wilks_scoreboard.xlsx"

    def test_required_sheets_exist(self):
        workbook = openpyxl.load_workbook(OUTPUT_FILE, data_only=False)
        assert workbook.sheetnames == ["Meet Results", "Wilks"]

    def test_meet_results_sheet_is_preserved(self):
        input_wb = openpyxl.load_workbook(INPUT_FILE, data_only=False)
        output_wb = openpyxl.load_workbook(OUTPUT_FILE, data_only=False)

        assert read_sheet_rows(output_wb["Meet Results"]) == read_sheet_rows(input_wb["Meet Results"])


class TestWilksSheet:
    def test_headers_and_column_count(self):
        workbook = openpyxl.load_workbook(OUTPUT_FILE, data_only=False)
        sheet = workbook["Wilks"]
        headers = [sheet.cell(row=1, column=col).value for col in range(1, 9)]

        assert headers == EXPECTED_HEADERS
        assert sheet.max_column == 8

    def test_row_count_matches_source_data(self):
        workbook = openpyxl.load_workbook(OUTPUT_FILE, data_only=False)
        sheet = workbook["Wilks"]
        expected_rows = len(load_input_records())
        assert sheet.max_row == expected_rows + 1

    def test_rows_are_sorted_by_wilks_and_copy_expected_columns(self):
        workbook = openpyxl.load_workbook(OUTPUT_FILE, data_only=False)
        sheet = workbook["Wilks"]
        expected_records = load_input_records()

        for row_idx, expected in enumerate(expected_records, start=2):
            actual_values = [sheet.cell(row=row_idx, column=col).value for col in range(1, 7)]
            expected_values = [expected[header] for header in REQUIRED_HEADERS]
            assert [normalize(value) for value in actual_values] == [
                normalize(value) for value in expected_values
            ]

    def test_totalkg_column_uses_excel_formulas(self):
        workbook = openpyxl.load_workbook(OUTPUT_FILE, data_only=False)
        sheet = workbook["Wilks"]

        for row_idx in range(2, sheet.max_row + 1):
            formula = sheet[f"G{row_idx}"].value
            assert isinstance(formula, str) and formula.startswith("=")
            assert "ROUND(" in formula.upper()
            assert f"D{row_idx}" in formula
            assert f"E{row_idx}" in formula
            assert f"F{row_idx}" in formula

    def test_wilks_column_uses_expected_formula_structure(self):
        workbook = openpyxl.load_workbook(OUTPUT_FILE, data_only=False)
        sheet = workbook["Wilks"]

        for row_idx in range(2, sheet.max_row + 1):
            formula = sheet[f"H{row_idx}"].value
            assert isinstance(formula, str) and formula.startswith("=")
            upper_formula = formula.upper()
            assert "ROUND(" in upper_formula
            assert "IF(" in upper_formula
            assert "POWER(" in upper_formula
            assert "MAX(" in upper_formula
            assert "MIN(" in upper_formula
            assert f"B{row_idx}" in formula
            assert f"C{row_idx}" in formula
            assert f"G{row_idx}" in formula
