import openpyxl
import pytest

OUTPUT_FILE = "/root/data/regatta_team_leaderboard.xlsx"

EXPECTED_HEADERS = [
    "Athlete",
    "Team",
    "Event",
    "EventType",
    "Place",
    "Points",
    "TeamTotal",
    "TeamOrder",
]

EXPECTED_ROWS = [
    ("Ava Chen", "Harbor BC", "Women's Single 500m", "Solo", 1, 10, 38, 3),
    ("Nina Lopez", "River Club", "Women's Single 500m", "Solo", 2, 7, 40, 4),
    ("Tessa Wong", "Summit RC", "Women's Single 500m", "Solo", 3, 5, 26, 4),
    ("Theo Gray", "Harbor BC", "Men's Single 500m", "Solo", 3, 5, 38, 4),
    ("Noah Diaz", "River Club", "Men's Single 500m", "Solo", 1, 10, 40, 2),
    ("Kai Lin", "Summit RC", "Men's Single 500m", "Solo", 2, 7, 26, 2),
    ("Ben Carter", "Harbor BC", "Men's Pair Heat", "Pair", 1, 12, 38, 1),
    ("Omar Khan", "River Club", "Men's Pair Heat", "Pair", 2, 8, 40, 3),
    ("Liam Reed", "Summit RC", "Men's Pair Heat", "Pair", 3, 6, 26, 3),
    ("Mia Park", "Harbor BC", "Women's Quad Final", "Crew", 2, 11, 38, 2),
    ("Zoe Patel", "River Club", "Women's Quad Final", "Crew", 1, 15, 40, 1),
    ("Erin Scott", "Summit RC", "Women's Quad Final", "Crew", 3, 8, 26, 1),
]


@pytest.fixture(scope="module")
def workbook_formula():
    return openpyxl.load_workbook(OUTPUT_FILE, data_only=False)


@pytest.fixture(scope="module")
def workbook_values():
    return openpyxl.load_workbook(OUTPUT_FILE, data_only=True)


def test_required_sheets_exist(workbook_formula):
    assert workbook_formula.sheetnames == ["Results", "ScoringRules", "Leaderboard"]


def test_leaderboard_headers(workbook_values):
    ws = workbook_values["Leaderboard"]
    headers = [ws.cell(row=1, column=idx).value for idx in range(1, 9)]
    assert headers == EXPECTED_HEADERS


def test_row_count_matches_results(workbook_values):
    results_ws = workbook_values["Results"]
    leaderboard_ws = workbook_values["Leaderboard"]
    assert leaderboard_ws.max_row == results_ws.max_row


def test_leaderboard_uses_formulas(workbook_formula):
    ws = workbook_formula["Leaderboard"]
    assert isinstance(ws["A2"].value, str) and ws["A2"].value == "=Results!A2"
    assert isinstance(ws["F2"].value, str) and ws["F2"].value.startswith("=")
    assert "ScoringRules!" in ws["F2"].value
    assert isinstance(ws["G2"].value, str) and "SUMIF" in ws["G2"].value
    assert isinstance(ws["H2"].value, str) and "SUMPRODUCT" in ws["H2"].value


def test_leaderboard_values(workbook_values):
    ws = workbook_values["Leaderboard"]
    actual_rows = []
    for row_idx in range(2, ws.max_row + 1):
        actual_rows.append(
            tuple(ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 9))
        )
    assert actual_rows == EXPECTED_ROWS


def test_team_totals_are_consistent(workbook_values):
    ws = workbook_values["Leaderboard"]
    team_totals = {}
    for row_idx in range(2, ws.max_row + 1):
        team = ws[f"B{row_idx}"].value
        total = ws[f"G{row_idx}"].value
        team_totals.setdefault(team, set()).add(total)

    assert team_totals == {
        "Harbor BC": {38},
        "River Club": {40},
        "Summit RC": {26},
    }


def test_team_order_is_contiguous_within_each_team(workbook_values):
    ws = workbook_values["Leaderboard"]
    team_orders = {}
    for row_idx in range(2, ws.max_row + 1):
        team = ws[f"B{row_idx}"].value
        order = ws[f"H{row_idx}"].value
        team_orders.setdefault(team, []).append(order)

    for orders in team_orders.values():
        assert sorted(orders) == [1, 2, 3, 4]
