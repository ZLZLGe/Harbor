from math import isclose, sqrt
from pathlib import Path
from statistics import mean, stdev

from openpyxl import load_workbook

INPUT_FILE = "/root/data/training_block_experiment.xlsx"
OUTPUT_FILE = "/root/results/training_block_readout.xlsx"
FLOAT_TOL = 1e-6

ATHLETE_HEADERS = [
    "athlete_id",
    "athlete_name",
    "sex",
    "training_group",
    "cohort",
    "baseline_bodyweight_kg",
    "baseline_total_kg",
    "end_bodyweight_kg",
    "end_total_kg",
    "delta_total_kg",
    "delta_total_per_bw",
]

SUMMARY_HEADERS = [
    "metric",
    "treatment_mean",
    "control_mean",
    "mean_diff",
    "ci95_lower",
    "ci95_upper",
    "cohens_d",
    "n_treatment",
    "n_control",
    "sample_size_balanced",
]


def load_sheet_rows(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    header = rows[0]
    return [dict(zip(header, row)) for row in rows[1:]]


def total_kg(row):
    return row["squat_kg"] + row["bench_kg"] + row["deadlift_kg"]


def expected_athlete_rows():
    wb = load_workbook(INPUT_FILE, data_only=True)
    roster_rows = load_sheet_rows(wb["Roster"])
    baseline_rows = {
        row["athlete_id"]: row for row in load_sheet_rows(wb["BaselineTest"])
    }
    end_rows = {
        row["athlete_id"]: row for row in load_sheet_rows(wb["BlockEnd"])
    }

    rows = []
    for roster in sorted(roster_rows, key=lambda item: item["athlete_id"]):
        athlete_id = roster["athlete_id"]
        baseline = baseline_rows[athlete_id]
        endline = end_rows[athlete_id]
        baseline_total = total_kg(baseline)
        end_total = total_kg(endline)
        delta_total = end_total - baseline_total

        rows.append(
            {
                "athlete_id": athlete_id,
                "athlete_name": roster["athlete_name"],
                "sex": roster["sex"],
                "training_group": roster["training_group"],
                "cohort": roster["cohort"],
                "baseline_bodyweight_kg": baseline["bodyweight_kg"],
                "baseline_total_kg": baseline_total,
                "end_bodyweight_kg": endline["bodyweight_kg"],
                "end_total_kg": end_total,
                "delta_total_kg": delta_total,
                "delta_total_per_bw": delta_total / baseline["bodyweight_kg"],
            }
        )
    return rows


def expected_summary_rows(athlete_rows):
    summary_rows = []
    for metric in ["delta_total_kg", "delta_total_per_bw"]:
        treatment = [row[metric] for row in athlete_rows if row["training_group"] == "treatment"]
        control = [row[metric] for row in athlete_rows if row["training_group"] == "control"]

        treatment_mean = mean(treatment)
        control_mean = mean(control)
        mean_diff = treatment_mean - control_mean
        treatment_sd = stdev(treatment)
        control_sd = stdev(control)
        n_treatment = len(treatment)
        n_control = len(control)

        se = sqrt((treatment_sd ** 2) / n_treatment + (control_sd ** 2) / n_control)
        pooled_sd = sqrt(
            (
                (n_treatment - 1) * (treatment_sd ** 2)
                + (n_control - 1) * (control_sd ** 2)
            )
            / (n_treatment + n_control - 2)
        )

        summary_rows.append(
            {
                "metric": metric,
                "treatment_mean": treatment_mean,
                "control_mean": control_mean,
                "mean_diff": mean_diff,
                "ci95_lower": mean_diff - 1.96 * se,
                "ci95_upper": mean_diff + 1.96 * se,
                "cohens_d": mean_diff / pooled_sd,
                "n_treatment": n_treatment,
                "n_control": n_control,
                "sample_size_balanced": abs(n_treatment - n_control) <= 1,
            }
        )
    return summary_rows


def actual_rows(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    header = list(rows[0])
    return header, rows[1:]


def boolish(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in {"true", "1", "yes"}
    return False


def test_output_file_exists():
    assert Path(OUTPUT_FILE).exists(), "missing output workbook"


def test_output_sheet_names():
    wb = load_workbook(OUTPUT_FILE, data_only=True)
    assert wb.sheetnames == ["AthleteReadout", "Summary"]


def test_athlete_readout_schema_and_values():
    wb = load_workbook(OUTPUT_FILE, data_only=True)
    header, rows = actual_rows(wb["AthleteReadout"])
    assert header == ATHLETE_HEADERS

    expected_rows = expected_athlete_rows()
    assert len(rows) == len(expected_rows)

    for actual, expected in zip(rows, expected_rows):
        actual_map = dict(zip(header, actual))
        for key, expected_value in expected.items():
            actual_value = actual_map[key]
            if isinstance(expected_value, float):
                assert isclose(actual_value, expected_value, rel_tol=0.0, abs_tol=FLOAT_TOL), (
                    key,
                    actual_value,
                    expected_value,
                )
            else:
                assert actual_value == expected_value


def test_summary_schema_row_order_and_values():
    wb = load_workbook(OUTPUT_FILE, data_only=True)
    header, rows = actual_rows(wb["Summary"])
    assert header == SUMMARY_HEADERS

    expected_rows = expected_summary_rows(expected_athlete_rows())
    assert len(rows) == 2

    for actual, expected in zip(rows, expected_rows):
        actual_map = dict(zip(header, actual))
        assert actual_map["metric"] == expected["metric"]
        for key in [
            "treatment_mean",
            "control_mean",
            "mean_diff",
            "ci95_lower",
            "ci95_upper",
            "cohens_d",
        ]:
            assert isclose(actual_map[key], expected[key], rel_tol=0.0, abs_tol=FLOAT_TOL), (
                key,
                actual_map[key],
                expected[key],
            )
        assert actual_map["n_treatment"] == expected["n_treatment"]
        assert actual_map["n_control"] == expected["n_control"]
        assert boolish(actual_map["sample_size_balanced"]) == expected["sample_size_balanced"]
