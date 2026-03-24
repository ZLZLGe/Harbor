from collections import defaultdict
from pathlib import Path

import openpyxl

INPUT_FILE = Path("/root/data/club_cup_entries.xlsx")
OUTPUT_FILE = Path("/root/data/team_dots_summary.xlsx")

DETAIL_HEADERS = [
    "Club",
    "LifterName",
    "Sex",
    "BodyweightKg",
    "Best3SquatKg",
    "Best3BenchKg",
    "Best3DeadliftKg",
    "TotalKg",
    "Dots",
]
TEAM_HEADERS = ["Rank", "Club", "ScoringLifters", "TeamDots"]

MALE_COEFFICIENTS = (-0.0000010930, 0.0007391293, -0.1918759221, 24.0900756, -307.75076)
FEMALE_COEFFICIENTS = (-0.0000010706, 0.0005158568, -0.1126655495, 13.6175032, -57.96288)


def normalize(value):
    if isinstance(value, float):
        return round(value, 3)
    return value


def read_sheet_rows(sheet):
    return [
        [cell.value for cell in row]
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column)
    ]


def calculate_dots(sex: str, bodyweight: float, total: float) -> float:
    if sex == "M":
        adjusted = max(40.0, min(210.0, bodyweight))
        a, b, c, d, e = MALE_COEFFICIENTS
    else:
        adjusted = max(40.0, min(150.0, bodyweight))
        a, b, c, d, e = FEMALE_COEFFICIENTS

    denominator = a * adjusted**4 + b * adjusted**3 + c * adjusted**2 + d * adjusted + e
    return round(total * (500 / denominator), 3)


def load_expected_records():
    workbook = openpyxl.load_workbook(INPUT_FILE, data_only=False)
    sheet = workbook["Club Entries"]
    headers = [cell.value for cell in sheet[1]]
    index = {header: idx for idx, header in enumerate(headers)}

    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {header: row[idx] for header, idx in index.items()}
        total = round(
            float(record["Best3SquatKg"])
            + float(record["Best3BenchKg"])
            + float(record["Best3DeadliftKg"]),
            3,
        )
        dots = calculate_dots(record["Sex"], float(record["BodyweightKg"]), total)
        records.append(
            {
                "Club": record["Club"],
                "LifterName": record["LifterName"],
                "Sex": record["Sex"],
                "BodyweightKg": float(record["BodyweightKg"]),
                "Best3SquatKg": float(record["Best3SquatKg"]),
                "Best3BenchKg": float(record["Best3BenchKg"]),
                "Best3DeadliftKg": float(record["Best3DeadliftKg"]),
                "TotalKg": total,
                "Dots": dots,
            }
        )

    records.sort(key=lambda item: item["Dots"], reverse=True)
    return records


def load_expected_team_rows():
    club_records = defaultdict(list)
    for record in load_expected_records():
        club_records[record["Club"]].append(record)

    team_rows = []
    for club, records in club_records.items():
        top_three = sorted(records, key=lambda item: item["Dots"], reverse=True)[:3]
        team_rows.append(
            {
                "Club": club,
                "ScoringLifters": ", ".join(item["LifterName"] for item in top_three),
                "TeamDots": round(sum(item["Dots"] for item in top_three), 3),
            }
        )

    team_rows.sort(key=lambda item: item["TeamDots"], reverse=True)
    for rank, row in enumerate(team_rows, start=1):
        row["Rank"] = rank
    return team_rows


class TestWorkbook:
    def test_output_exists(self):
        assert OUTPUT_FILE.exists(), "缺少输出文件 /root/data/team_dots_summary.xlsx"

    def test_required_sheets_exist(self):
        workbook = openpyxl.load_workbook(OUTPUT_FILE, data_only=False)
        assert workbook.sheetnames == ["Club Entries", "Athlete Dots", "Team Podium"]

    def test_club_entries_sheet_is_preserved(self):
        input_wb = openpyxl.load_workbook(INPUT_FILE, data_only=False)
        output_wb = openpyxl.load_workbook(OUTPUT_FILE, data_only=False)

        assert read_sheet_rows(output_wb["Club Entries"]) == read_sheet_rows(input_wb["Club Entries"])


class TestAthleteDots:
    def test_headers_and_column_count(self):
        workbook = openpyxl.load_workbook(OUTPUT_FILE, data_only=False)
        sheet = workbook["Athlete Dots"]
        headers = [sheet.cell(row=1, column=col).value for col in range(1, 10)]

        assert headers == DETAIL_HEADERS
        assert sheet.max_column == 9

    def test_row_count_matches_source_data(self):
        workbook = openpyxl.load_workbook(OUTPUT_FILE, data_only=False)
        sheet = workbook["Athlete Dots"]
        assert sheet.max_row == len(load_expected_records()) + 1

    def test_rows_are_sorted_by_dots_and_copy_expected_columns(self):
        workbook = openpyxl.load_workbook(OUTPUT_FILE, data_only=False)
        sheet = workbook["Athlete Dots"]
        expected_records = load_expected_records()

        for row_idx, expected in enumerate(expected_records, start=2):
            actual_values = [sheet.cell(row=row_idx, column=col).value for col in range(1, 8)]
            expected_values = [expected[header] for header in DETAIL_HEADERS[:7]]
            assert [normalize(value) for value in actual_values] == [
                normalize(value) for value in expected_values
            ]

    def test_totalkg_column_uses_excel_formulas(self):
        workbook = openpyxl.load_workbook(OUTPUT_FILE, data_only=False)
        sheet = workbook["Athlete Dots"]

        for row_idx in range(2, sheet.max_row + 1):
            formula = sheet[f"H{row_idx}"].value
            assert isinstance(formula, str) and formula.startswith("=")
            assert "ROUND(" in formula.upper()
            assert f"E{row_idx}" in formula
            assert f"F{row_idx}" in formula
            assert f"G{row_idx}" in formula

    def test_dots_column_uses_expected_formula_structure(self):
        workbook = openpyxl.load_workbook(OUTPUT_FILE, data_only=False)
        sheet = workbook["Athlete Dots"]

        for row_idx in range(2, sheet.max_row + 1):
            formula = sheet[f"I{row_idx}"].value
            assert isinstance(formula, str) and formula.startswith("=")
            upper_formula = formula.upper()
            assert "ROUND(" in upper_formula
            assert "IF(" in upper_formula
            assert "POWER(" in upper_formula
            assert "MAX(" in upper_formula
            assert "MIN(" in upper_formula
            assert f"C{row_idx}" in formula
            assert f"D{row_idx}" in formula
            assert f"H{row_idx}" in formula


class TestTeamPodium:
    def test_headers_and_column_count(self):
        workbook = openpyxl.load_workbook(OUTPUT_FILE, data_only=False)
        sheet = workbook["Team Podium"]
        headers = [sheet.cell(row=1, column=col).value for col in range(1, 5)]

        assert headers == TEAM_HEADERS
        assert sheet.max_column == 4

    def test_team_rows_match_expected_summary(self):
        workbook = openpyxl.load_workbook(OUTPUT_FILE, data_only=False)
        sheet = workbook["Team Podium"]
        expected_rows = load_expected_team_rows()

        assert sheet.max_row == len(expected_rows) + 1

        for row_idx, expected in enumerate(expected_rows, start=2):
            actual = {
                "Rank": sheet.cell(row=row_idx, column=1).value,
                "Club": sheet.cell(row=row_idx, column=2).value,
                "ScoringLifters": sheet.cell(row=row_idx, column=3).value,
                "TeamDots": sheet.cell(row=row_idx, column=4).value,
            }
            assert actual["Rank"] == expected["Rank"]
            assert actual["Club"] == expected["Club"]
            assert actual["ScoringLifters"] == expected["ScoringLifters"]
            assert normalize(actual["TeamDots"]) == expected["TeamDots"]

    def test_teamdots_is_sorted_descending(self):
        workbook = openpyxl.load_workbook(OUTPUT_FILE, data_only=False)
        sheet = workbook["Team Podium"]
        values = [sheet.cell(row=row_idx, column=4).value for row_idx in range(2, sheet.max_row + 1)]
        normalized = [normalize(value) for value in values]
        assert normalized == sorted(normalized, reverse=True)
