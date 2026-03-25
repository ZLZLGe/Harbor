from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

OUTPUT_FILE = Path("/root/data/wilks_division_scoreboard.xlsx")
EXPECTED_HEADERS = [
    "Name",
    "Division",
    "Sex",
    "BodyweightKg",
    "Best3SquatKg",
    "Best3BenchKg",
    "Best3DeadliftKg",
    "TotalKg",
    "Wilks",
    "DivisionRank",
]

MALE_COEFFICIENTS = (-216.0475144, 16.2606339, -0.002388645, -0.00113732, 0.00000701863, -0.00000001291)
FEMALE_COEFFICIENTS = (594.31747775582, -27.23842536447, 0.82112226871, -0.00930733913, 0.00004731582, -0.00000009054)


def wilks_score(sex, bodyweight, total):
    a, b, c, d, e, f = MALE_COEFFICIENTS if sex == "M" else FEMALE_COEFFICIENTS
    denominator = a + b * bodyweight + c * bodyweight**2 + d * bodyweight**3 + e * bodyweight**4 + f * bodyweight**5
    return round(total * (500 / denominator), 3)


def build_expected_rows(results_ws):
    rows = []
    for row_idx in range(2, results_ws.max_row + 1):
        name = results_ws[f"A{row_idx}"].value
        division = results_ws[f"B{row_idx}"].value
        sex = results_ws[f"C{row_idx}"].value
        bodyweight = float(results_ws[f"D{row_idx}"].value)
        squat = float(results_ws[f"E{row_idx}"].value)
        bench = float(results_ws[f"F{row_idx}"].value)
        deadlift = float(results_ws[f"G{row_idx}"].value)
        total = squat + bench + deadlift
        wilks = wilks_score(sex, bodyweight, total)
        rows.append(
            {
                "Name": name,
                "Division": division,
                "Sex": sex,
                "BodyweightKg": bodyweight,
                "Best3SquatKg": squat,
                "Best3BenchKg": bench,
                "Best3DeadliftKg": deadlift,
                "TotalKg": total,
                "Wilks": wilks,
            }
        )

    division_to_scores = defaultdict(list)
    for row in rows:
        division_to_scores[row["Division"]].append(row["Wilks"])

    for row in rows:
        row["DivisionRank"] = 1 + sum(score > row["Wilks"] for score in division_to_scores[row["Division"]])

    return rows


def normalize_formula(value):
    return value.replace(" ", "").replace("$", "").replace("'", "").upper()


def test_output_file_exists():
    assert OUTPUT_FILE.exists(), f"Missing output file: {OUTPUT_FILE}"


def test_workbook_structure_and_headers():
    wb = load_workbook(OUTPUT_FILE, data_only=False)
    assert wb.sheetnames == ["MeetResults", "Scoreboard"]

    score_ws = wb["Scoreboard"]
    headers = [score_ws.cell(row=1, column=col).value for col in range(1, len(EXPECTED_HEADERS) + 1)]
    assert headers == EXPECTED_HEADERS

    results_ws = wb["MeetResults"]
    assert score_ws.max_row == results_ws.max_row


def test_scoreboard_uses_formulas_for_all_required_columns():
    wb = load_workbook(OUTPUT_FILE, data_only=False)
    score_ws = wb["Scoreboard"]

    for row_idx in range(2, score_ws.max_row + 1):
        for col in "ABCDEFGHIJ":
            value = score_ws[f"{col}{row_idx}"].value
            assert isinstance(value, str) and value.startswith("="), f"{col}{row_idx} should contain a formula"


def test_source_columns_reference_meetresults_cells():
    wb = load_workbook(OUTPUT_FILE, data_only=False)
    score_ws = wb["Scoreboard"]

    for row_idx in range(2, score_ws.max_row + 1):
        for col in "ABCDEFG":
            formula = normalize_formula(score_ws[f"{col}{row_idx}"].value)
            assert formula == f"=MEETRESULTS!{col}{row_idx}"


def test_wilks_formula_preserves_semantics_without_fixed_branching_syntax():
    wb = load_workbook(OUTPUT_FILE, data_only=False)
    score_ws = wb["Scoreboard"]

    for row_idx in range(2, score_ws.max_row + 1):
        formula = normalize_formula(score_ws[f"I{row_idx}"].value)
        assert "ROUND(" in formula, f"I{row_idx} should round the score to 3 decimals"
        assert ",3)" in formula, f"I{row_idx} should keep 3 decimal places"
        assert f"D{row_idx}" in formula, f"I{row_idx} should depend on BodyweightKg"
        assert (
            f"H{row_idx}" in formula
            or all(f"{col}{row_idx}" in formula for col in "EFG")
        ), f"I{row_idx} should depend on the lifter total"


def test_cached_values_match_expected_semantics():
    wb_values = load_workbook(OUTPUT_FILE, data_only=True)
    results_ws = wb_values["MeetResults"]
    score_ws = wb_values["Scoreboard"]

    expected_rows = build_expected_rows(results_ws)

    for row_idx, expected in enumerate(expected_rows, start=2):
        assert score_ws[f"A{row_idx}"].value == expected["Name"]
        assert score_ws[f"B{row_idx}"].value == expected["Division"]
        assert score_ws[f"C{row_idx}"].value == expected["Sex"]
        assert float(score_ws[f"D{row_idx}"].value) == expected["BodyweightKg"]
        assert float(score_ws[f"E{row_idx}"].value) == expected["Best3SquatKg"]
        assert float(score_ws[f"F{row_idx}"].value) == expected["Best3BenchKg"]
        assert float(score_ws[f"G{row_idx}"].value) == expected["Best3DeadliftKg"]
        assert round(float(score_ws[f"H{row_idx}"].value), 3) == round(expected["TotalKg"], 3)
        assert round(float(score_ws[f"I{row_idx}"].value), 3) == expected["Wilks"]
        assert int(score_ws[f"J{row_idx}"].value) == expected["DivisionRank"]
