from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

OUTPUT_FILE = Path("/root/data/commission_payout_audit.xlsx")
EXPECTED_HEADERS = [
    "OrderID",
    "OrderDate",
    "Quarter",
    "RepID",
    "RepName",
    "Segment",
    "Amount",
    "BaseRate",
    "QuarterRepSales",
    "AcceleratorRate",
    "BaseCommission",
    "AcceleratorBonus",
    "FinalPayout",
    "AuditFlag",
]


def normalize_blank(value):
    if value in ("", None):
        return None
    return value


def get_accelerator_rate(quarter_sales, tier):
    if tier is None:
        return None
    if quarter_sales >= tier["Tier2Min"]:
        return tier["Tier2Rate"]
    if quarter_sales >= tier["Tier1Min"]:
        return tier["Tier1Rate"]
    return 0.0


def build_expected_rows(wb_values):
    orders_ws = wb_values["Orders"]
    rules_ws = wb_values["RepRules"]
    tiers_ws = wb_values["AcceleratorTiers"]

    rep_rules = {}
    for row_idx in range(2, rules_ws.max_row + 1):
        rep_id = rules_ws[f"A{row_idx}"].value
        rep_rules[rep_id] = {
            "RepName": rules_ws[f"B{row_idx}"].value,
            "BaseRate": float(rules_ws[f"D{row_idx}"].value),
        }

    tiers = {}
    for row_idx in range(2, tiers_ws.max_row + 1):
        quarter = tiers_ws[f"A{row_idx}"].value
        tiers[quarter] = {
            "Tier1Min": float(tiers_ws[f"B{row_idx}"].value),
            "Tier1Rate": float(tiers_ws[f"C{row_idx}"].value),
            "Tier2Min": float(tiers_ws[f"D{row_idx}"].value),
            "Tier2Rate": float(tiers_ws[f"E{row_idx}"].value),
        }

    quarter_totals = defaultdict(float)
    orders = []
    for row_idx in range(2, orders_ws.max_row + 1):
        row = {
            "OrderID": orders_ws[f"A{row_idx}"].value,
            "OrderDate": orders_ws[f"B{row_idx}"].value,
            "Quarter": orders_ws[f"C{row_idx}"].value,
            "RepID": orders_ws[f"D{row_idx}"].value,
            "Segment": orders_ws[f"E{row_idx}"].value,
            "Amount": float(orders_ws[f"F{row_idx}"].value),
        }
        quarter_totals[(row["Quarter"], row["RepID"])] += row["Amount"]
        orders.append(row)

    expected_rows = []
    for order in orders:
        quarter = order["Quarter"]
        rep_id = order["RepID"]
        amount = order["Amount"]
        quarter_sales = round(quarter_totals[(quarter, rep_id)], 2)
        tier = tiers.get(quarter)
        accelerator_rate = get_accelerator_rate(quarter_sales, tier)

        rep_rule = rep_rules.get(rep_id)
        if rep_rule is None:
            expected_rows.append(
                {
                    **order,
                    "RepName": None,
                    "BaseRate": None,
                    "QuarterRepSales": quarter_sales,
                    "AcceleratorRate": accelerator_rate,
                    "BaseCommission": None,
                    "AcceleratorBonus": None,
                    "FinalPayout": None,
                    "AuditFlag": "MISSING_REP_RULE",
                }
            )
            continue

        base_rate = rep_rule["BaseRate"]
        base_commission = round(amount * base_rate, 2)
        if tier is None:
            expected_rows.append(
                {
                    **order,
                    "RepName": rep_rule["RepName"],
                    "BaseRate": base_rate,
                    "QuarterRepSales": quarter_sales,
                    "AcceleratorRate": None,
                    "BaseCommission": base_commission,
                    "AcceleratorBonus": None,
                    "FinalPayout": None,
                    "AuditFlag": "MISSING_ACCELERATOR_RULE",
                }
            )
            continue

        accelerator_bonus = round(amount * accelerator_rate, 2)
        final_payout = round(base_commission + accelerator_bonus, 2)

        expected_rows.append(
            {
                **order,
                "RepName": rep_rule["RepName"],
                "BaseRate": base_rate,
                "QuarterRepSales": quarter_sales,
                "AcceleratorRate": accelerator_rate,
                "BaseCommission": base_commission,
                "AcceleratorBonus": accelerator_bonus,
                "FinalPayout": final_payout,
                "AuditFlag": "OK",
            }
        )

    return expected_rows


def test_output_file_exists():
    assert OUTPUT_FILE.exists(), f"Missing output file: {OUTPUT_FILE}"


def test_workbook_structure_and_headers():
    wb = load_workbook(OUTPUT_FILE, data_only=False)
    assert wb.sheetnames == ["Orders", "RepRules", "AcceleratorTiers", "Payouts"]

    payouts_ws = wb["Payouts"]
    headers = [payouts_ws.cell(row=1, column=col).value for col in range(1, len(EXPECTED_HEADERS) + 1)]
    assert headers == EXPECTED_HEADERS

    orders_ws = wb["Orders"]
    assert payouts_ws.max_row == orders_ws.max_row


def test_payout_sheet_uses_formulas_for_all_data_columns():
    wb = load_workbook(OUTPUT_FILE, data_only=False)
    payouts_ws = wb["Payouts"]

    for row_idx in range(2, payouts_ws.max_row + 1):
        for col in "ABCDEFGHIJKLMN":
            value = payouts_ws[f"{col}{row_idx}"].value
            assert isinstance(value, str) and value.startswith("="), f"{col}{row_idx} should contain a formula"


def test_formulas_reference_required_source_sheets():
    wb = load_workbook(OUTPUT_FILE, data_only=False)
    payouts_ws = wb["Payouts"]

    for cell in ("A2", "B2", "C2", "D2", "F2", "G2"):
        assert "Orders!" in payouts_ws[cell].value
    for cell in ("E2", "H2"):
        assert "RepRules!" in payouts_ws[cell].value
    assert "AcceleratorTiers!" in payouts_ws["J2"].value
    assert "AcceleratorTiers!" in payouts_ws["N2"].value


def test_cached_values_match_expected_semantics():
    wb_values = load_workbook(OUTPUT_FILE, data_only=True)
    payouts_ws = wb_values["Payouts"]
    expected_rows = build_expected_rows(wb_values)

    for row_idx, expected in enumerate(expected_rows, start=2):
        assert payouts_ws[f"A{row_idx}"].value == expected["OrderID"]
        assert payouts_ws[f"B{row_idx}"].value == expected["OrderDate"]
        assert payouts_ws[f"C{row_idx}"].value == expected["Quarter"]
        assert payouts_ws[f"D{row_idx}"].value == expected["RepID"]
        assert normalize_blank(payouts_ws[f"E{row_idx}"].value) == expected["RepName"]
        assert payouts_ws[f"F{row_idx}"].value == expected["Segment"]
        assert round(float(payouts_ws[f"G{row_idx}"].value), 2) == expected["Amount"]
        assert normalize_blank(payouts_ws[f"H{row_idx}"].value) == expected["BaseRate"]
        assert round(float(payouts_ws[f"I{row_idx}"].value), 2) == expected["QuarterRepSales"]
        assert normalize_blank(payouts_ws[f"J{row_idx}"].value) == expected["AcceleratorRate"]
        assert normalize_blank(payouts_ws[f"K{row_idx}"].value) == expected["BaseCommission"]
        assert normalize_blank(payouts_ws[f"L{row_idx}"].value) == expected["AcceleratorBonus"]
        assert normalize_blank(payouts_ws[f"M{row_idx}"].value) == expected["FinalPayout"]
        assert payouts_ws[f"N{row_idx}"].value == expected["AuditFlag"]
