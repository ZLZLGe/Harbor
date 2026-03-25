#!/usr/bin/env python3
import math

from openpyxl import load_workbook

OUTPUT_FILE = "/root/retail_margin_pack.xlsx"

EXPECTED_PIVOTS = {
    "Orders by Territory": {"row": "Territory", "data": "Orders", "subtotal": "sum", "col": None},
    "Net Revenue by Format": {"row": "StoreFormat", "data": "NetRevenue", "subtotal": "sum", "col": None},
    "Promo Spend by Territory": {"row": "Territory", "data": "PromoSpend", "subtotal": "sum", "col": None},
    "Net Revenue by Territory Band": {"row": "Territory", "data": "NetRevenue", "subtotal": "sum", "col": "ReturnBand"},
}


def pivot_field_names(pivot):
    return [field.name for field in pivot.cache.cacheFields]


def field_name_from_axis(pivot, axis_fields):
    names = pivot_field_names(pivot)
    if axis_fields and len(axis_fields) > 0:
        idx = axis_fields[0].x
        if idx is not None and 0 <= idx < len(names):
            return names[idx]
    return None


def data_field_name(pivot):
    names = pivot_field_names(pivot)
    fld = pivot.dataFields[0].fld
    if fld is not None and 0 <= fld < len(names):
        return names[fld]
    return None


def main():
    workbook = load_workbook(OUTPUT_FILE)
    assert workbook.sheetnames == [
        "SourceData",
        "Orders by Territory",
        "Net Revenue by Format",
        "Promo Spend by Territory",
        "Net Revenue by Territory Band",
    ]

    ws = workbook["SourceData"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(cell) if cell is not None else "" for cell in rows[0]]
    data = [dict(zip(headers, row)) for row in rows[1:] if any(cell is not None for cell in row)]
    assert headers == [
        "Territory",
        "StoreFormat",
        "Orders",
        "Units",
        "AvgTicket",
        "ReturnRate",
        "PromoSpend",
        "GrossRevenue",
        "ReturnBand",
        "NetRevenue",
    ]

    bands = {row["ReturnBand"] for row in data if row["ReturnBand"] is not None}
    assert bands == {"Q1", "Q2", "Q3", "Q4"}

    for row in data[:50]:
        gross = float(row["Orders"]) * float(row["AvgTicket"])
        net = gross * (1 - float(row["ReturnRate"])) - float(row["PromoSpend"])
        assert math.isclose(float(row["GrossRevenue"]), gross, rel_tol=0, abs_tol=1e-6)
        assert math.isclose(float(row["NetRevenue"]), net, rel_tol=0, abs_tol=1e-6)

    for sheet_name, config in EXPECTED_PIVOTS.items():
        pivot = workbook[sheet_name]._pivots[0]
        assert field_name_from_axis(pivot, pivot.rowFields) == config["row"]
        assert data_field_name(pivot) == config["data"]
        assert pivot.dataFields[0].subtotal == config["subtotal"]
        if config["col"] is None:
            assert len(pivot.colFields) == 0
        else:
            assert field_name_from_axis(pivot, pivot.colFields) == config["col"]


if __name__ == "__main__":
    main()
