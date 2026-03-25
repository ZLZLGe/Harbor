#!/usr/bin/env python3
import math

from openpyxl import load_workbook

OUTPUT_FILE = "/root/training_grant_dashboard.xlsx"

EXPECTED_PIVOTS = {
    "Participants by Region": {"row": "ProgramRegion", "data": "Participants", "subtotal": "sum", "col": None},
    "Completed by Track": {"row": "CourseTrack", "data": "CompletedParticipants", "subtotal": "sum", "col": None},
    "Grant Spend by Region": {"row": "ProgramRegion", "data": "GrantSpend", "subtotal": "sum", "col": None},
    "Grant Spend by Region Band": {"row": "ProgramRegion", "data": "GrantSpend", "subtotal": "sum", "col": "CompletionBand"},
}


def pivot_field_names(pivot):
    return [field.name for field in pivot.cache.cacheFields]


def field_name_from_axis(pivot, axis_fields):
    names = pivot_field_names(pivot)
    if axis_fields and len(axis_fields) > 0:
        idx = axis_fields[0].x
        if idx is not None and 0 <= idx < len(names):
            return names[idx]
    return None


def data_field_name(pivot):
    names = pivot_field_names(pivot)
    fld = pivot.dataFields[0].fld
    if fld is not None and 0 <= fld < len(names):
        return names[fld]
    return None


def main():
    workbook = load_workbook(OUTPUT_FILE)
    assert workbook.sheetnames == [
        "SourceData",
        "Participants by Region",
        "Completed by Track",
        "Grant Spend by Region",
        "Grant Spend by Region Band",
    ]

    ws = workbook["SourceData"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(cell) if cell is not None else "" for cell in rows[0]]
    data = [dict(zip(headers, row)) for row in rows[1:] if any(cell is not None for cell in row)]
    assert headers == [
        "ProgramRegion",
        "CourseTrack",
        "Participants",
        "CompletionRate",
        "GrantPerParticipant",
        "MentorHours",
        "CompletedParticipants",
        "CompletionBand",
        "GrantSpend",
    ]

    bands = {row["CompletionBand"] for row in data if row["CompletionBand"] is not None}
    assert bands == {"Q1", "Q2", "Q3", "Q4"}

    for row in data[:50]:
        expected_completed = round(float(row["Participants"]) * float(row["CompletionRate"]))
        expected_spend = float(row["Participants"]) * float(row["GrantPerParticipant"])
        assert int(row["CompletedParticipants"]) == expected_completed
        assert math.isclose(float(row["GrantSpend"]), expected_spend, rel_tol=0, abs_tol=1e-6)

    for sheet_name, config in EXPECTED_PIVOTS.items():
        pivot = workbook[sheet_name]._pivots[0]
        assert field_name_from_axis(pivot, pivot.rowFields) == config["row"]
        assert data_field_name(pivot) == config["data"]
        assert pivot.dataFields[0].subtotal == config["subtotal"]
        if config["col"] is None:
            assert len(pivot.colFields) == 0
        else:
            assert field_name_from_axis(pivot, pivot.colFields) == config["col"]


if __name__ == "__main__":
    main()
