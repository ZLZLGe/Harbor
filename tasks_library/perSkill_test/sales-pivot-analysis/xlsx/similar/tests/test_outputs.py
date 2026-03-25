#!/usr/bin/env python3
import math

import pandas as pd
from openpyxl import load_workbook

INPUT_FILE = "/root/income.xlsx"
OUTPUT_FILE = "/root/income_band_review.xlsx"

EXPECTED_PIVOTS = {
    "Earners by Band": {"row": "IncomeBand", "data": "EARNERS", "subtotal": "sum", "col": None},
    "Median Income by Prefix": {"row": "SA2_PREFIX", "data": "MEDIAN_INCOME", "subtotal": "average", "col": None},
    "Mean Income by Band": {"row": "IncomeBand", "data": "MEAN_INCOME", "subtotal": "average", "col": None},
    "Payroll by Prefix Band": {"row": "SA2_PREFIX", "data": "EstimatedPayroll", "subtotal": "sum", "col": "IncomeBand"},
}


def pivot_field_names(pivot):
    return [field.name for field in pivot.cache.cacheFields]


def field_name_from_axis(pivot, axis_fields):
    names = pivot_field_names(pivot)
    if axis_fields and len(axis_fields) > 0:
        idx = axis_fields[0].x
        if idx is not None and 0 <= idx < len(names):
            return names[idx]
    return None


def data_field_name(pivot):
    names = pivot_field_names(pivot)
    fld = pivot.dataFields[0].fld
    if fld is not None and 0 <= fld < len(names):
        return names[fld]
    return None


def main():
    workbook = load_workbook(OUTPUT_FILE)
    assert workbook.sheetnames == [
        "SourceData",
        "Earners by Band",
        "Median Income by Prefix",
        "Mean Income by Band",
        "Payroll by Prefix Band",
    ]

    ws = workbook["SourceData"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(cell) if cell is not None else "" for cell in rows[0]]
    data = [dict(zip(headers, row)) for row in rows[1:] if any(cell is not None for cell in row)]
    assert headers == [
        "SA2_CODE",
        "SA2_NAME",
        "EARNERS",
        "MEDIAN_INCOME",
        "MEAN_INCOME",
        "SA2_PREFIX",
        "IncomeBand",
        "EstimatedPayroll",
    ]

    input_df = pd.read_excel(INPUT_FILE)
    assert len(data) == len(input_df)

    bands = {row["IncomeBand"] for row in data if row["IncomeBand"] is not None}
    assert bands == {"Q1", "Q2", "Q3", "Q4"}

    for row in data[:50]:
        expected_payroll = float(row["EARNERS"]) * float(row["MEDIAN_INCOME"])
        assert math.isclose(float(row["EstimatedPayroll"]), expected_payroll, rel_tol=0, abs_tol=1e-6)
        assert row["SA2_PREFIX"] == str(row["SA2_CODE"])[:3]

    for sheet_name, config in EXPECTED_PIVOTS.items():
        pivot = workbook[sheet_name]._pivots[0]
        assert field_name_from_axis(pivot, pivot.rowFields) == config["row"]
        assert data_field_name(pivot) == config["data"]
        assert pivot.dataFields[0].subtotal == config["subtotal"]
        if config["col"] is None:
            assert len(pivot.colFields) == 0
        else:
            assert field_name_from_axis(pivot, pivot.colFields) == config["col"]


if __name__ == "__main__":
    main()
