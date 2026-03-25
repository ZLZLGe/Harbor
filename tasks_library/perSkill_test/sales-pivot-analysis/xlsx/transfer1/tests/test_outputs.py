#!/usr/bin/env python3
import math

from openpyxl import load_workbook

OUTPUT_FILE = "/root/clinic_demand_review.xlsx"

EXPECTED_PIVOTS = {
    "Visits by Zone": {"row": "ServiceZone", "data": "BookedVisits", "subtotal": "sum", "col": None},
    "Capacity by Zone": {"row": "ServiceZone", "data": "CapacityVisits", "subtotal": "sum", "col": None},
    "Revenue Gap by Band": {"row": "PressureBand", "data": "RevenueGap", "subtotal": "sum", "col": None},
    "Gap by Zone and Band": {"row": "ServiceZone", "data": "RevenueGap", "subtotal": "sum", "col": "PressureBand"},
}


def pivot_field_names(pivot):
    return [field.name for field in pivot.cache.cacheFields]


def field_name_from_axis(pivot, axis_fields):
    names = pivot_field_names(pivot)
    if axis_fields and len(axis_fields) > 0:
        idx = axis_fields[0].x
        if idx is not None and 0 <= idx < len(names):
            return names[idx]
    return None


def data_field_name(pivot):
    names = pivot_field_names(pivot)
    fld = pivot.dataFields[0].fld
    if fld is not None and 0 <= fld < len(names):
        return names[fld]
    return None


def main():
    workbook = load_workbook(OUTPUT_FILE)
    assert workbook.sheetnames == [
        "SourceData",
        "Visits by Zone",
        "Capacity by Zone",
        "Revenue Gap by Band",
        "Gap by Zone and Band",
    ]

    ws = workbook["SourceData"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(cell) if cell is not None else "" for cell in rows[0]]
    data = [dict(zip(headers, row)) for row in rows[1:] if any(cell is not None for cell in row)]
    assert headers == [
        "ServiceZone",
        "ClinicName",
        "BookedVisits",
        "CapacityVisits",
        "AvgVisitRevenue",
        "NurseHours",
        "PressureRatio",
        "PressureBand",
        "RevenueGap",
    ]

    bands = {row["PressureBand"] for row in data if row["PressureBand"] is not None}
    assert bands == {"Q1", "Q2", "Q3", "Q4"}

    for row in data[:50]:
        expected_ratio = float(row["BookedVisits"]) / float(row["CapacityVisits"])
        expected_gap = max(float(row["BookedVisits"]) - float(row["CapacityVisits"]), 0.0) * float(row["AvgVisitRevenue"])
        assert math.isclose(float(row["PressureRatio"]), expected_ratio, rel_tol=0, abs_tol=1e-9)
        assert math.isclose(float(row["RevenueGap"]), expected_gap, rel_tol=0, abs_tol=1e-6)

    for sheet_name, config in EXPECTED_PIVOTS.items():
        pivot = workbook[sheet_name]._pivots[0]
        assert field_name_from_axis(pivot, pivot.rowFields) == config["row"]
        assert data_field_name(pivot) == config["data"]
        assert pivot.dataFields[0].subtotal == config["subtotal"]
        if config["col"] is None:
            assert len(pivot.colFields) == 0
        else:
            assert field_name_from_axis(pivot, pivot.colFields) == config["col"]


if __name__ == "__main__":
    main()
