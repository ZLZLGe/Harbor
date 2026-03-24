#!/usr/bin/env python3

from __future__ import annotations

from decimal import Decimal, ROUND_HALF_UP

import pandas as pd
import pytest
from openpyxl import load_workbook

OUTPUT_FILE = "/root/warehouse_restock_analysis.xlsx"
INVENTORY_FILE = "/root/warehouse_inventory.xlsx"
RULES_FILE = "/root/restock_rules.xlsx"

EXPECTED_SHEETS = [
    "InventoryDetail",
    "Warehouse Value",
    "Category Aging",
    "Priority Gap",
    "Warehouse Priority Matrix",
]

DETAIL_HEADERS = [
    "BatchID",
    "Warehouse",
    "SKU",
    "ItemName",
    "Category",
    "ReceivedDate",
    "AsOfDate",
    "AgeDays",
    "AgingBucket",
    "OnHandUnits",
    "UnitCost",
    "InventoryValue",
    "WeeklyDemand",
    "WeeksCover",
    "TargetWeeksCover",
    "SafetyWeeks",
    "GapToTargetUnits",
    "TurnoverRisk",
    "RestockPriority",
]

PIVOT_SPECS = [
    ("Warehouse Value", "Warehouse", "InventoryValue", "sum", None),
    ("Category Aging", "Category", "OnHandUnits", "sum", "AgingBucket"),
    ("Priority Gap", "RestockPriority", "GapToTargetUnits", "sum", None),
    ("Warehouse Priority Matrix", "Warehouse", "InventoryValue", "sum", "RestockPriority"),
]

PRIORITY_ORDER = {"Urgent": 0, "Replenish": 1, "Normal": 2, "Hold": 3, "Monitor": 4}


def rounded(value: object) -> float:
    return float(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def normalize_warehouse(series: pd.Series) -> pd.Series:
    return series.astype(str).str.strip().str.title()


def normalize_sku(series: pd.Series) -> pd.Series:
    return series.astype(str).str.strip().str.upper()


def aging_bucket(age_days: int) -> str:
    if age_days <= 30:
        return "0-30"
    if age_days <= 60:
        return "31-60"
    if age_days <= 90:
        return "61-90"
    if age_days <= 180:
        return "91-180"
    return "181+"


def turnover_risk(row: pd.Series) -> str:
    if row["WeeklyDemand"] == 0:
        return "Dormant"
    if row["AgeDays"] > row["CriticalAgingDays"] and row["WeeksCover"] > row["TargetWeeksCover"]:
        return "Critical Aging"
    if row["WeeksCover"] < row["SafetyWeeks"]:
        return "Low Cover"
    if row["WeeksCover"] > row["TargetWeeksCover"] * 1.5:
        return "Excess Cover"
    return "Healthy"


def restock_priority(row: pd.Series) -> str:
    if row["WeeklyDemand"] == 0:
        return "Monitor"
    if row["WeeksCover"] < row["SafetyWeeks"]:
        return "Urgent"
    if row["GapToTargetUnits"] > 0:
        return "Replenish"
    if row["TurnoverRisk"] in {"Critical Aging", "Excess Cover"}:
        return "Hold"
    return "Normal"


def compute_expected_frame() -> pd.DataFrame:
    ledger = pd.read_excel(INVENTORY_FILE, sheet_name="Ledger")
    products = pd.read_excel(INVENTORY_FILE, sheet_name="Products")
    demand = pd.read_excel(RULES_FILE, sheet_name="DemandPlan")
    policy = pd.read_excel(RULES_FILE, sheet_name="Policy")
    parameters = pd.read_excel(RULES_FILE, sheet_name="Parameters")

    for frame in (ledger, demand):
        frame["Warehouse"] = normalize_warehouse(frame["Warehouse"])
        frame["SKU"] = normalize_sku(frame["SKU"])

    products["SKU"] = normalize_sku(products["SKU"])
    products["Category"] = products["Category"].astype(str).str.strip()
    products["ItemName"] = products["ItemName"].astype(str).str.strip()
    policy["Category"] = policy["Category"].astype(str).str.strip()

    ledger["ReceivedDate"] = pd.to_datetime(ledger["ReceivedDate"])
    as_of_date = pd.to_datetime(
        parameters.loc[parameters["Parameter"].astype(str).str.strip() == "AsOfDate", "Value"].iloc[0]
    )

    detail = ledger.merge(products, on="SKU", how="left").merge(
        demand,
        on=["Warehouse", "SKU"],
        how="left",
    ).merge(
        policy,
        on="Category",
        how="left",
    )

    detail["AgeDays"] = (as_of_date - detail["ReceivedDate"]).dt.days.astype(int)
    detail["AgingBucket"] = detail["AgeDays"].apply(aging_bucket)
    detail["InventoryValue"] = (detail["OnHandUnits"] * detail["UnitCost"]).round(2)
    detail["WeeksCover"] = detail.apply(
        lambda row: round(row["OnHandUnits"] / row["WeeklyDemand"], 2) if row["WeeklyDemand"] else 0.0,
        axis=1,
    )
    detail["GapToTargetUnits"] = (
        (detail["TargetWeeksCover"] * detail["WeeklyDemand"] - detail["OnHandUnits"])
        .clip(lower=0)
        .astype(int)
    )
    detail["TurnoverRisk"] = detail.apply(turnover_risk, axis=1)
    detail["RestockPriority"] = detail.apply(restock_priority, axis=1)
    detail["ReceivedDate"] = detail["ReceivedDate"].dt.strftime("%Y-%m-%d")
    detail["AsOfDate"] = as_of_date.strftime("%Y-%m-%d")
    detail["priority_order"] = detail["RestockPriority"].map(PRIORITY_ORDER)
    detail = detail.sort_values(["Warehouse", "priority_order", "AgeDays", "BatchID"], ascending=[True, True, False, True])
    detail = detail.drop(columns=["priority_order"]).reset_index(drop=True)
    return detail[DETAIL_HEADERS]


def worksheet_to_frame(workbook) -> pd.DataFrame:
    sheet = workbook["InventoryDetail"]
    rows = list(sheet.iter_rows(values_only=True))
    frame = pd.DataFrame(rows[1:], columns=rows[0])
    for column in ["AgeDays", "OnHandUnits", "WeeklyDemand", "TargetWeeksCover", "SafetyWeeks", "GapToTargetUnits"]:
        frame[column] = frame[column].astype(int)
    for column in ["UnitCost", "InventoryValue", "WeeksCover"]:
        frame[column] = frame[column].apply(rounded)
    frame["ReceivedDate"] = frame["ReceivedDate"].astype(str).str.strip()
    frame["AsOfDate"] = frame["AsOfDate"].astype(str).str.strip()
    return frame[DETAIL_HEADERS]


def pivot_field_names(pivot) -> list[str]:
    return [field.name for field in pivot.cache.cacheFields]


def field_name_from_ref(pivot, refs) -> str | None:
    if not refs:
        return None
    index = refs[0].x
    if index is None:
        return None
    return pivot_field_names(pivot)[index]


def data_field_source_name(pivot) -> str | None:
    if not pivot.dataFields:
        return None
    index = pivot.dataFields[0].fld
    if index is None:
        return None
    return pivot_field_names(pivot)[index]


@pytest.fixture(scope="module")
def workbook():
    return load_workbook(OUTPUT_FILE)


@pytest.fixture(scope="module")
def expected_frame() -> pd.DataFrame:
    frame = compute_expected_frame().copy()
    for column in ["UnitCost", "InventoryValue", "WeeksCover"]:
        frame[column] = frame[column].apply(rounded)
    return frame


@pytest.fixture(scope="module")
def actual_frame(workbook) -> pd.DataFrame:
    return worksheet_to_frame(workbook)


def test_workbook_has_required_sheets(workbook):
    assert workbook.sheetnames == EXPECTED_SHEETS


def test_inventory_detail_headers_are_exact(actual_frame):
    assert list(actual_frame.columns) == DETAIL_HEADERS


def test_inventory_detail_matches_expected(expected_frame, actual_frame):
    pd.testing.assert_frame_equal(
        actual_frame.reset_index(drop=True),
        expected_frame.reset_index(drop=True),
        check_dtype=False,
    )


def test_priority_distribution(actual_frame):
    assert actual_frame["RestockPriority"].value_counts().to_dict() == {
        "Hold": 3,
        "Monitor": 1,
        "Normal": 1,
        "Replenish": 2,
        "Urgent": 2,
    }


def test_dormant_batch_is_marked_monitor(actual_frame):
    row = actual_frame.loc[actual_frame["BatchID"] == "B006"].iloc[0]
    assert row["WeeklyDemand"] == 0
    assert row["WeeksCover"] == 0.0
    assert row["TurnoverRisk"] == "Dormant"
    assert row["RestockPriority"] == "Monitor"


def test_gap_totals_by_priority_match_expected(expected_frame, actual_frame):
    expected = expected_frame.groupby("RestockPriority", as_index=False)["GapToTargetUnits"].sum().sort_values("RestockPriority").reset_index(drop=True)
    actual = actual_frame.groupby("RestockPriority", as_index=False)["GapToTargetUnits"].sum().sort_values("RestockPriority").reset_index(drop=True)
    pd.testing.assert_frame_equal(actual, expected, check_dtype=False)


def test_inventory_value_by_warehouse_matches_expected(expected_frame, actual_frame):
    expected = expected_frame.groupby("Warehouse", as_index=False)["InventoryValue"].sum().sort_values("Warehouse").reset_index(drop=True)
    actual = actual_frame.groupby("Warehouse", as_index=False)["InventoryValue"].sum().sort_values("Warehouse").reset_index(drop=True)
    pd.testing.assert_frame_equal(actual, expected, check_dtype=False)


@pytest.mark.parametrize("sheet_name,row_field,data_field,subtotal,col_field", PIVOT_SPECS)
def test_pivot_configuration(workbook, expected_frame, sheet_name, row_field, data_field, subtotal, col_field):
    sheet = workbook[sheet_name]
    assert len(sheet._pivots) == 1

    pivot = sheet._pivots[0]
    assert field_name_from_ref(pivot, pivot.rowFields) == row_field
    assert data_field_source_name(pivot) == data_field
    assert pivot.dataFields[0].subtotal == subtotal

    if col_field is None:
        assert len(pivot.colFields) == 0
    else:
        assert field_name_from_ref(pivot, pivot.colFields) == col_field

    source = pivot.cache.cacheSource.worksheetSource
    assert source.sheet == "InventoryDetail"
    assert source.ref == f"A1:S{len(expected_frame) + 1}"
