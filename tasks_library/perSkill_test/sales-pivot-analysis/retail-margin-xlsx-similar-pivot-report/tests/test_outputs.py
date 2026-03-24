#!/usr/bin/env python3

from __future__ import annotations

from decimal import Decimal, ROUND_HALF_UP

import pandas as pd
import pytest
from openpyxl import load_workbook

OUTPUT_FILE = "/root/retail_margin_report.xlsx"
SALES_FILE = "/root/sales_transactions.xlsx"
PRODUCT_FILE = "/root/product_master.xlsx"

EXPECTED_SHEETS = [
    "SourceData",
    "Margin by Region",
    "Margin by Category",
    "Net Sales by Channel",
    "Discount Band Profit",
]

EXPECTED_HEADERS = [
    "OrderID",
    "OrderDate",
    "Region",
    "Channel",
    "SKU",
    "Category",
    "Units",
    "UnitPrice",
    "DiscountPct",
    "UnitCost",
    "GrossSales",
    "NetSales",
    "TotalCost",
    "GrossProfit",
    "DiscountBand",
]

PIVOT_SPECS = [
    ("Margin by Region", "Region", "GrossProfit", "sum", None),
    ("Margin by Category", "Category", "GrossProfit", "sum", None),
    ("Net Sales by Channel", "Channel", "NetSales", "sum", None),
    ("Discount Band Profit", "DiscountBand", "GrossProfit", "sum", "Channel"),
]


def rounded(value: object) -> float:
    return float(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def compute_expected_frame() -> pd.DataFrame:
    sales = pd.read_excel(SALES_FILE)
    products = pd.read_excel(PRODUCT_FILE)

    sales["SKU"] = sales["SKU"].astype(str).str.strip().str.upper()
    sales["Region"] = sales["Region"].astype(str).str.strip().str.title()
    sales["Channel"] = sales["Channel"].astype(str).str.strip().str.title()
    sales["OrderDate"] = sales["OrderDate"].astype(str).str.strip()

    for column in ["Units", "UnitPrice", "DiscountPct"]:
        sales[column] = pd.to_numeric(sales[column], errors="raise")

    products["SKU"] = products["SKU"].astype(str).str.strip().str.upper()
    products["Category"] = products["Category"].astype(str).str.strip()
    products["UnitCost"] = pd.to_numeric(products["UnitCost"], errors="raise")

    merged = sales.merge(products[["SKU", "Category", "UnitCost"]], on="SKU", how="inner")
    merged["GrossSales"] = (merged["Units"] * merged["UnitPrice"]).round(2)
    merged["NetSales"] = (merged["GrossSales"] * (1 - merged["DiscountPct"])).round(2)
    merged["TotalCost"] = (merged["Units"] * merged["UnitCost"]).round(2)
    merged["GrossProfit"] = (merged["NetSales"] - merged["TotalCost"]).round(2)

    def band(value: float) -> str:
        if value == 0:
            return "No Discount"
        if value < 0.10:
            return "1-9%"
        if value < 0.20:
            return "10-19%"
        return "20%+"

    merged["DiscountBand"] = merged["DiscountPct"].apply(band)
    return merged[EXPECTED_HEADERS]


def worksheet_to_frame(workbook) -> pd.DataFrame:
    sheet = workbook["SourceData"]
    rows = list(sheet.iter_rows(values_only=True))
    frame = pd.DataFrame(rows[1:], columns=rows[0])
    for column in ["Units", "UnitPrice", "DiscountPct", "UnitCost", "GrossSales", "NetSales", "TotalCost", "GrossProfit"]:
        frame[column] = frame[column].apply(rounded)
    frame["OrderDate"] = frame["OrderDate"].astype(str).str.strip()
    return frame[EXPECTED_HEADERS]


def pivot_field_names(pivot) -> list[str]:
    return [field.name for field in pivot.cache.cacheFields]


def field_name_from_ref(pivot, refs) -> str | None:
    fields = pivot_field_names(pivot)
    if not refs:
        return None
    index = refs[0].x
    if index is None:
        return None
    return fields[index]


def data_field_source_name(pivot) -> str | None:
    fields = pivot_field_names(pivot)
    if not pivot.dataFields:
        return None
    index = pivot.dataFields[0].fld
    if index is None:
        return None
    return fields[index]


@pytest.fixture(scope="module")
def workbook():
    return load_workbook(OUTPUT_FILE)


@pytest.fixture(scope="module")
def expected_frame() -> pd.DataFrame:
    frame = compute_expected_frame().copy()
    for column in ["Units", "UnitPrice", "DiscountPct", "UnitCost", "GrossSales", "NetSales", "TotalCost", "GrossProfit"]:
        frame[column] = frame[column].apply(rounded)
    return frame


@pytest.fixture(scope="module")
def actual_frame(workbook) -> pd.DataFrame:
    return worksheet_to_frame(workbook)


def test_workbook_has_required_sheets(workbook):
    assert workbook.sheetnames == EXPECTED_SHEETS


def test_source_headers_are_exact(actual_frame):
    assert list(actual_frame.columns) == EXPECTED_HEADERS


def test_source_data_matches_expected(expected_frame, actual_frame):
    pd.testing.assert_frame_equal(
        actual_frame.reset_index(drop=True),
        expected_frame.reset_index(drop=True),
        check_dtype=False,
    )


def test_discount_band_values_are_expected(actual_frame):
    assert set(actual_frame["DiscountBand"]) == {"No Discount", "1-9%", "10-19%", "20%+"}


def test_region_profit_totals_match_expected(expected_frame, actual_frame):
    expected = expected_frame.groupby("Region", as_index=False)["GrossProfit"].sum().sort_values("Region").reset_index(drop=True)
    actual = actual_frame.groupby("Region", as_index=False)["GrossProfit"].sum().sort_values("Region").reset_index(drop=True)
    pd.testing.assert_frame_equal(actual, expected, check_dtype=False)


def test_channel_net_sales_totals_match_expected(expected_frame, actual_frame):
    expected = expected_frame.groupby("Channel", as_index=False)["NetSales"].sum().sort_values("Channel").reset_index(drop=True)
    actual = actual_frame.groupby("Channel", as_index=False)["NetSales"].sum().sort_values("Channel").reset_index(drop=True)
    pd.testing.assert_frame_equal(actual, expected, check_dtype=False)


@pytest.mark.parametrize("sheet_name,row_field,data_field,subtotal,col_field", PIVOT_SPECS)
def test_pivot_configuration(workbook, expected_frame, sheet_name, row_field, data_field, subtotal, col_field):
    sheet = workbook[sheet_name]
    assert len(sheet._pivots) == 1

    pivot = sheet._pivots[0]
    assert field_name_from_ref(pivot, pivot.rowFields) == row_field
    assert data_field_source_name(pivot) == data_field
    assert pivot.dataFields[0].subtotal == subtotal

    if col_field is None:
        assert len(pivot.colFields) == 0
    else:
        assert field_name_from_ref(pivot, pivot.colFields) == col_field

    source = pivot.cache.cacheSource.worksheetSource
    assert source.sheet == "SourceData"
    assert source.ref == f"A1:O{len(expected_frame) + 1}"
