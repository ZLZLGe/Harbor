#!/usr/bin/env python3

import os
import statistics

import pytest
from openpyxl import load_workbook

EXCEL_FILE = "/root/metabolite_treatment_shift_scorecard.xlsx"


def build_raw_lookup(data_ws):
    sample_by_col = {}
    for col in range(4, data_ws.max_column + 1):
        sample = data_ws.cell(row=1, column=col).value
        if sample:
            sample_by_col[sample] = col

    metabolite_rows = {}
    for row in range(2, data_ws.max_row + 1):
        metabolite_id = data_ws.cell(row=row, column=1).value
        if metabolite_id:
            metabolite_rows[metabolite_id] = row

    return metabolite_rows, sample_by_col


def expected_metrics(value_wb):
    task = value_wb["Task"]
    data = value_wb["Data"]
    metabolite_rows, sample_by_col = build_raw_lookup(data)

    samples = [task.cell(row=10, column=col).value for col in range(3, 13)]
    groups = [task.cell(row=9, column=col).value for col in range(3, 13)]
    responder_samples = [sample for sample, group in zip(samples, groups) if group == "Responder"]
    nonresponder_samples = [sample for sample, group in zip(samples, groups) if group == "Nonresponder"]

    metrics = []
    for task_row in range(11, 19):
        metabolite_id = task.cell(row=task_row, column=1).value
        metabolite_name = task.cell(row=task_row, column=2).value
        data_row = metabolite_rows[metabolite_id]

        lookup_values = {}
        for sample in samples:
            lookup_values[sample] = data.cell(row=data_row, column=sample_by_col[sample]).value

        responder_values = [lookup_values[sample] for sample in responder_samples]
        nonresponder_values = [lookup_values[sample] for sample in nonresponder_samples]

        responder_mean = statistics.mean(responder_values)
        responder_std = statistics.stdev(responder_values)
        nonresponder_mean = statistics.mean(nonresponder_values)
        nonresponder_std = statistics.stdev(nonresponder_values)
        log2_shift = responder_mean - nonresponder_mean
        fold_change = 2 ** log2_shift
        abs_shift = abs(log2_shift)
        direction = "Higher in Responder" if log2_shift > 0 else "Higher in Nonresponder"

        metrics.append(
            {
                "task_row": task_row,
                "summary_row": task_row + 21,
                "metabolite_id": metabolite_id,
                "metabolite_name": metabolite_name,
                "lookup_values": lookup_values,
                "responder_mean": responder_mean,
                "responder_std": responder_std,
                "nonresponder_mean": nonresponder_mean,
                "nonresponder_std": nonresponder_std,
                "log2_shift": log2_shift,
                "fold_change": fold_change,
                "abs_shift": abs_shift,
                "direction": direction,
            }
        )

    return metrics


@pytest.fixture(scope="module")
def formula_wb():
    assert os.path.exists(EXCEL_FILE), f"Missing workbook: {EXCEL_FILE}"
    return load_workbook(EXCEL_FILE, data_only=False)


@pytest.fixture(scope="module")
def value_wb():
    assert os.path.exists(EXCEL_FILE), f"Missing workbook: {EXCEL_FILE}"
    return load_workbook(EXCEL_FILE, data_only=True)


def test_file_and_sheets_exist(value_wb):
    assert os.path.exists(EXCEL_FILE)
    assert value_wb.sheetnames == ["Task", "Data"]


def test_lookup_formulas_and_values(formula_wb, value_wb):
    task_formula = formula_wb["Task"]
    task_values = value_wb["Task"]
    metrics = expected_metrics(value_wb)

    for metric in metrics:
        task_row = metric["task_row"]
        for col in range(3, 13):
            sample = task_values.cell(row=10, column=col).value
            formula = task_formula.cell(row=task_row, column=col).value
            value = task_values.cell(row=task_row, column=col).value
            assert isinstance(formula, str) and formula.startswith("=")
            assert "Data!" in formula
            assert value == pytest.approx(metric["lookup_values"][sample], abs=1e-9)


def test_group_statistics_are_formula_driven_and_correct(formula_wb, value_wb):
    task_formula = formula_wb["Task"]
    task_values = value_wb["Task"]
    metrics = expected_metrics(value_wb)

    stat_rows = {
        24: "responder_mean",
        25: "responder_std",
        26: "nonresponder_mean",
        27: "nonresponder_std",
    }

    for offset, metric in enumerate(metrics, start=2):
        for row, field in stat_rows.items():
            formula = task_formula.cell(row=row, column=offset).value
            value = task_values.cell(row=row, column=offset).value
            assert isinstance(formula, str) and formula.startswith("=")
            assert value == pytest.approx(metric[field], abs=1e-6)


def test_shift_summary_matches_expected(formula_wb, value_wb):
    task_formula = formula_wb["Task"]
    task_values = value_wb["Task"]
    metrics = expected_metrics(value_wb)

    for metric in metrics:
        row = metric["summary_row"]
        assert task_values[f"A{row}"].value == metric["metabolite_id"]
        assert task_values[f"B{row}"].value == metric["metabolite_name"]

        for col in ["C", "D", "E"]:
            formula = task_formula[f"{col}{row}"].value
            assert isinstance(formula, str) and formula.startswith("=")

        assert task_values[f"C{row}"].value == pytest.approx(metric["log2_shift"], abs=1e-6)
        assert task_values[f"D{row}"].value == pytest.approx(metric["fold_change"], abs=1e-6)
        assert task_values[f"E{row}"].value == pytest.approx(metric["abs_shift"], abs=1e-6)


def test_top_four_ranking_is_correct(formula_wb, value_wb):
    task_formula = formula_wb["Task"]
    task_values = value_wb["Task"]
    ranked = sorted(expected_metrics(value_wb), key=lambda item: item["abs_shift"], reverse=True)[:4]

    for idx, metric in enumerate(ranked, start=32):
        for col in ["I", "J", "K", "L"]:
            formula = task_formula[f"{col}{idx}"].value
            assert isinstance(formula, str) and formula.startswith("=")

        assert task_values[f"H{idx}"].value == idx - 31
        assert task_values[f"I{idx}"].value == metric["metabolite_id"]
        assert task_values[f"J{idx}"].value == metric["metabolite_name"]
        assert task_values[f"K{idx}"].value == pytest.approx(metric["abs_shift"], abs=1e-6)
        assert task_values[f"L{idx}"].value == metric["direction"]


def test_no_excel_error_strings_in_output(value_wb):
    task = value_wb["Task"]
    for row in range(11, 40):
        for col in range(1, 13):
            value = task.cell(row=row, column=col).value
            if isinstance(value, str):
                assert not value.startswith("#"), f"Excel error at {row},{col}: {value}"
