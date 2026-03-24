#!/usr/bin/env python3

import os

import pytest
from openpyxl import load_workbook

EXCEL_FILE = "/root/campus_energy_tariff_variance.xlsx"


def load_meter_data(value_wb):
    ws = value_wb["MeterData"]
    rows = {}
    by_building = {}

    for row in range(2, ws.max_row + 1):
        record = {
            "month": ws[f"A{row}"].value,
            "building": ws[f"B{row}"].value,
            "meter_id": ws[f"C{row}"].value,
            "plan": ws[f"D{row}"].value,
            "band": ws[f"E{row}"].value,
            "kwh": ws[f"F{row}"].value,
        }
        rows[(record["month"], record["building"], record["meter_id"], record["plan"], record["band"])] = record
        by_building.setdefault(record["building"], {})[record["band"]] = record

    return rows, by_building


def load_rates(value_wb):
    ws = value_wb["Rates"]
    rates = {}
    for row in range(2, ws.max_row + 1):
        rates[(ws[f"A{row}"].value, ws[f"B{row}"].value)] = ws[f"C{row}"].value
    return rates


def load_budget(value_wb):
    ws = value_wb["Budget"]
    budget = {}
    for row in range(2, ws.max_row + 1):
        budget[ws[f"A{row}"].value] = {
            "budget": ws[f"B{row}"].value,
            "shift_share": ws[f"C{row}"].value,
        }
    return budget


def expected_metrics(value_wb):
    review = value_wb["Tariff Review"]
    lookup_rows, by_building = load_meter_data(value_wb)
    rates = load_rates(value_wb)
    budget = load_budget(value_wb)

    detail = []
    for row in range(8, 18):
        key = (
            review[f"A{row}"].value,
            review[f"B{row}"].value,
            review[f"C{row}"].value,
            review[f"D{row}"].value,
            review[f"E{row}"].value,
        )
        record = lookup_rows[key]
        rate = rates[(record["plan"], record["band"])]
        detail.append(
            {
                "row": row,
                "kwh": record["kwh"],
                "rate": rate,
                "cost": record["kwh"] * rate,
            }
        )

    summary = []
    for row in range(23, 28):
        building = review[f"A{row}"].value
        peak = by_building[building]["Peak"]
        valley = by_building[building]["Valley"]
        peak_rate = rates[(peak["plan"], "Peak")]
        valley_rate = rates[(valley["plan"], "Valley")]
        peak_cost = peak["kwh"] * peak_rate
        valley_cost = valley["kwh"] * valley_rate
        total_cost = peak_cost + valley_cost
        building_budget = budget[building]["budget"]
        shift_share = budget[building]["shift_share"]
        variance = total_cost - building_budget
        saveable = peak["kwh"] * shift_share * (peak_rate - valley_rate)
        action = "Act Now" if variance > 0 and saveable >= 300 else "Plan Shift" if saveable >= 200 else "Track"
        summary.append(
            {
                "row": row,
                "building": building,
                "peak_cost": peak_cost,
                "valley_cost": valley_cost,
                "total_cost": total_cost,
                "budget": building_budget,
                "variance": variance,
                "shift_share": shift_share,
                "saveable": saveable,
                "action": action,
            }
        )

    ranking = sorted(summary, key=lambda item: item["saveable"], reverse=True)[:4]
    return detail, summary, ranking


@pytest.fixture(scope="module")
def formula_wb():
    assert os.path.exists(EXCEL_FILE), f"Missing workbook: {EXCEL_FILE}"
    return load_workbook(EXCEL_FILE, data_only=False)


@pytest.fixture(scope="module")
def value_wb():
    assert os.path.exists(EXCEL_FILE), f"Missing workbook: {EXCEL_FILE}"
    return load_workbook(EXCEL_FILE, data_only=True)


def test_file_and_sheet_order(value_wb):
    assert os.path.exists(EXCEL_FILE)
    assert value_wb.sheetnames == ["Tariff Review", "MeterData", "Rates", "Budget"]


def test_detail_pull_area_is_formula_driven_and_correct(formula_wb, value_wb):
    review_formula = formula_wb["Tariff Review"]
    review_values = value_wb["Tariff Review"]
    detail, _, _ = expected_metrics(value_wb)

    for item in detail:
        row = item["row"]
        assert isinstance(review_formula[f"F{row}"].value, str) and review_formula[f"F{row}"].value.startswith("=")
        assert isinstance(review_formula[f"G{row}"].value, str) and review_formula[f"G{row}"].value.startswith("=")
        assert isinstance(review_formula[f"H{row}"].value, str) and review_formula[f"H{row}"].value.startswith("=")
        assert "MeterData!" in review_formula[f"F{row}"].value
        assert "Rates!" in review_formula[f"G{row}"].value

        assert review_values[f"F{row}"].value == pytest.approx(item["kwh"], abs=1e-9)
        assert review_values[f"G{row}"].value == pytest.approx(item["rate"], abs=1e-9)
        assert review_values[f"H{row}"].value == pytest.approx(item["cost"], abs=1e-6)


def test_building_summary_matches_expected(formula_wb, value_wb):
    review_formula = formula_wb["Tariff Review"]
    review_values = value_wb["Tariff Review"]
    _, summary, _ = expected_metrics(value_wb)

    for item in summary:
        row = item["row"]
        for col in "BCDEFGHI":
            assert isinstance(review_formula[f"{col}{row}"].value, str) and review_formula[f"{col}{row}"].value.startswith("=")

        assert "Budget!" in review_formula[f"E{row}"].value
        assert "Budget!" in review_formula[f"G{row}"].value

        assert review_values[f"B{row}"].value == pytest.approx(item["peak_cost"], abs=1e-6)
        assert review_values[f"C{row}"].value == pytest.approx(item["valley_cost"], abs=1e-6)
        assert review_values[f"D{row}"].value == pytest.approx(item["total_cost"], abs=1e-6)
        assert review_values[f"E{row}"].value == pytest.approx(item["budget"], abs=1e-9)
        assert review_values[f"F{row}"].value == pytest.approx(item["variance"], abs=1e-6)
        assert review_values[f"G{row}"].value == pytest.approx(item["shift_share"], abs=1e-9)
        assert review_values[f"H{row}"].value == pytest.approx(item["saveable"], abs=1e-6)
        assert review_values[f"I{row}"].value == item["action"]


def test_priority_table_is_sorted_by_saveable_amount(formula_wb, value_wb):
    review_formula = formula_wb["Tariff Review"]
    review_values = value_wb["Tariff Review"]
    _, _, ranking = expected_metrics(value_wb)

    for rank_row, item in zip(range(32, 36), ranking):
        for col in "JKLMN":
            assert isinstance(review_formula[f"{col}{rank_row}"].value, str)
            assert review_formula[f"{col}{rank_row}"].value.startswith("=")

        assert review_values[f"I{rank_row}"].value == rank_row - 31
        assert review_values[f"J{rank_row}"].value == item["building"]
        assert review_values[f"K{rank_row}"].value == pytest.approx(item["total_cost"], abs=1e-6)
        assert review_values[f"L{rank_row}"].value == pytest.approx(item["variance"], abs=1e-6)
        assert review_values[f"M{rank_row}"].value == pytest.approx(item["saveable"], abs=1e-6)
        assert review_values[f"N{rank_row}"].value == item["action"]


def test_no_excel_error_strings_in_output(value_wb):
    review = value_wb["Tariff Review"]
    for row in list(range(8, 18)) + list(range(23, 28)) + list(range(31, 36)):
        for col in range(1, 15):
            value = review.cell(row=row, column=col).value
            if isinstance(value, str):
                assert not value.startswith("#"), f"Excel error at {row},{col}: {value}"
