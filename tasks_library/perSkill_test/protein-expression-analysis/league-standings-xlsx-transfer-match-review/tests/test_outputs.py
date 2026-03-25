import os

import pytest
from openpyxl import load_workbook

EXCEL_FILE = "/root/league_standings_review.xlsx"


def _build_expected(wb_values):
    review = wb_values["Review"]
    results = wb_values["Results"]
    teams = wb_values["Teams"]

    name_map = {}
    for row in range(2, teams.max_row + 1):
        name_map[teams[f"A{row}"].value] = teams[f"B{row}"].value

    stats = {
        review[f"A{row}"].value: {
            "team_name": name_map[review[f"A{row}"].value],
            "wins": 0,
            "draws": 0,
            "losses": 0,
            "gf": 0,
            "ga": 0,
            "gd": 0,
            "points": 0,
        }
        for row in range(8, 16)
    }

    final_matches = 0
    total_goals = 0

    for row in range(2, results.max_row + 1):
        if results[f"C{row}"].value != "Final":
            continue

        final_matches += 1
        home = results[f"D{row}"].value
        away = results[f"E{row}"].value
        home_goals = results[f"F{row}"].value
        away_goals = results[f"G{row}"].value
        total_goals += home_goals + away_goals

        stats[home]["gf"] += home_goals
        stats[home]["ga"] += away_goals
        stats[away]["gf"] += away_goals
        stats[away]["ga"] += home_goals

        if home_goals > away_goals:
            stats[home]["wins"] += 1
            stats[home]["points"] += 3
            stats[away]["losses"] += 1
        elif away_goals > home_goals:
            stats[away]["wins"] += 1
            stats[away]["points"] += 3
            stats[home]["losses"] += 1
        else:
            stats[home]["draws"] += 1
            stats[away]["draws"] += 1
            stats[home]["points"] += 1
            stats[away]["points"] += 1

    for item in stats.values():
        item["gd"] = item["gf"] - item["ga"]

    ranked = sorted(
        stats.items(),
        key=lambda kv: (-kv[1]["points"], -kv[1]["gd"], -kv[1]["gf"], kv[0]),
    )

    for rank, (team_code, values) in enumerate(ranked, start=1):
        values["rank"] = rank

    return stats, ranked, final_matches, total_goals


@pytest.fixture(scope="module")
def wb_formula():
    assert os.path.exists(EXCEL_FILE), f"missing workbook: {EXCEL_FILE}"
    return load_workbook(EXCEL_FILE, data_only=False)


@pytest.fixture(scope="module")
def wb_values():
    assert os.path.exists(EXCEL_FILE), f"missing workbook: {EXCEL_FILE}"
    return load_workbook(EXCEL_FILE, data_only=True)


def test_required_sheets_exist(wb_formula):
    assert wb_formula.sheetnames == ["Review", "Results", "Teams"]


def test_target_regions_use_formulas(wb_formula):
    review = wb_formula["Review"]

    for row in range(8, 16):
        for col in range(2, 11):
            value = review.cell(row, col).value
            assert isinstance(value, str) and value.startswith("="), f"Review!{review.cell(row, col).coordinate} should be a formula"

    for cell_ref in ["L3", "M3", "N3"]:
        value = review[cell_ref].value
        assert isinstance(value, str) and value.startswith("="), f"Review!{cell_ref} should be a formula"

    for row in range(8, 16):
        for col in range(12, 19):
            value = review.cell(row, col).value
            assert isinstance(value, str) and value.startswith("="), f"Review!{review.cell(row, col).coordinate} should be a formula"


def test_team_review_rows_are_correct(wb_values):
    review = wb_values["Review"]
    expected, _, _, _ = _build_expected(wb_values)

    for row in range(8, 16):
        team_code = review[f"A{row}"].value
        values = expected[team_code]
        assert review[f"B{row}"].value == values["team_name"]
        assert review[f"C{row}"].value == pytest.approx(values["wins"], rel=0, abs=1e-9)
        assert review[f"D{row}"].value == pytest.approx(values["draws"], rel=0, abs=1e-9)
        assert review[f"E{row}"].value == pytest.approx(values["losses"], rel=0, abs=1e-9)
        assert review[f"F{row}"].value == pytest.approx(values["gf"], rel=0, abs=1e-9)
        assert review[f"G{row}"].value == pytest.approx(values["ga"], rel=0, abs=1e-9)
        assert review[f"H{row}"].value == pytest.approx(values["gd"], rel=0, abs=1e-9)
        assert review[f"I{row}"].value == pytest.approx(values["points"], rel=0, abs=1e-9)
        assert review[f"J{row}"].value == pytest.approx(values["rank"], rel=0, abs=1e-9)


def test_overview_cells_are_correct(wb_values):
    review = wb_values["Review"]
    _, ranked, final_matches, total_goals = _build_expected(wb_values)

    assert review["L3"].value == pytest.approx(final_matches, rel=0, abs=1e-9)
    assert review["M3"].value == pytest.approx(total_goals, rel=0, abs=1e-9)
    assert review["N3"].value == ranked[0][0]


def test_sorted_annex_matches_tie_break_order(wb_values):
    review = wb_values["Review"]
    _, ranked, _, _ = _build_expected(wb_values)

    for idx, row in enumerate(range(8, 16), start=1):
        team_code, values = ranked[idx - 1]
        expected_zone = (
            "PROMOTION" if idx <= 2 else "PLAYOFF" if idx <= 4 else "RELEGATION" if idx >= 7 else "SAFE"
        )

        assert review[f"L{row}"].value == idx
        assert review[f"M{row}"].value == team_code
        assert review[f"N{row}"].value == values["team_name"]
        assert review[f"O{row}"].value == pytest.approx(values["points"], rel=0, abs=1e-9)
        assert review[f"P{row}"].value == pytest.approx(values["gd"], rel=0, abs=1e-9)
        assert review[f"Q{row}"].value == pytest.approx(values["gf"], rel=0, abs=1e-9)
        assert review[f"R{row}"].value == expected_zone


def test_postponed_matches_are_excluded_and_alpha_tie_break_applies(wb_values):
    review = wb_values["Review"]
    results = wb_values["Results"]
    expected, ranked, final_matches, _ = _build_expected(wb_values)

    postponed_rows = [row for row in range(2, results.max_row + 1) if results[f"C{row}"].value == "Postponed"]
    assert len(postponed_rows) == 2
    assert review["L3"].value == final_matches

    exact_tie_pair = [team for team, values in ranked if values["points"] == 9 and values["gd"] == 4 and values["gf"] == 8]
    assert exact_tie_pair == ["ECL", "HZN"]
    assert expected["ECL"]["rank"] < expected["HZN"]["rank"]


def test_no_excel_error_strings_after_recalc(wb_values):
    for sheet in wb_values.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    assert not cell.value.startswith("#"), f"Excel error left at {sheet.title}!{cell.coordinate}: {cell.value}"
