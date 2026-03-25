import os

import pytest
from openpyxl import load_workbook

EXCEL_FILE = "/root/energy_budget_audit.xlsx"


def _expected_rows(wb_values):
    audit = wb_values["BudgetAudit"]
    readings = wb_values["MeterReadings"]
    tariffs = wb_values["TariffTable"]
    buildings = wb_values["BuildingInfo"]

    month = audit["B2"].value

    building_map = {}
    for row in range(2, buildings.max_row + 1):
        building_map[buildings[f"A{row}"].value] = {
            "name": buildings[f"B{row}"].value,
            "area": buildings[f"C{row}"].value,
            "budget": buildings[f"D{row}"].value,
        }

    tariff_map = {}
    for row in range(2, tariffs.max_row + 1):
        if tariffs[f"A{row}"].value == month:
            tariff_map[tariffs[f"B{row}"].value] = tariffs[f"C{row}"].value

    usage_map = {}
    for row in range(2, readings.max_row + 1):
        if readings[f"C{row}"].value != month:
            continue
        key = (readings[f"B{row}"].value, readings[f"D{row}"].value)
        usage_map[key] = usage_map.get(key, 0) + readings[f"E{row}"].value

    expected = {}
    for row in range(7, 13):
        building_id = audit[f"A{row}"].value
        area = building_map[building_id]["area"]
        budget = building_map[building_id]["budget"]
        peak = usage_map[(building_id, "Peak")]
        flat = usage_map[(building_id, "Flat")]
        valley = usage_map[(building_id, "Valley")]
        peak_price = tariff_map["Peak"]
        flat_price = tariff_map["Flat"]
        valley_price = tariff_map["Valley"]
        total_cost = peak * peak_price + flat * flat_price + valley * valley_price
        cost_per_sqm = total_cost / area
        variance = total_cost - budget
        variance_pct = variance / budget
        status = "OVER" if variance > 0 else "WITHIN"

        expected[row] = {
            "area": area,
            "budget": budget,
            "peak": peak,
            "flat": flat,
            "valley": valley,
            "peak_price": peak_price,
            "flat_price": flat_price,
            "valley_price": valley_price,
            "total_cost": total_cost,
            "cost_per_sqm": cost_per_sqm,
            "variance": variance,
            "variance_pct": variance_pct,
            "status": status,
            "name": building_map[building_id]["name"],
        }

    return expected


@pytest.fixture(scope="module")
def wb_formula():
    assert os.path.exists(EXCEL_FILE), f"missing workbook: {EXCEL_FILE}"
    return load_workbook(EXCEL_FILE, data_only=False)


@pytest.fixture(scope="module")
def wb_values():
    assert os.path.exists(EXCEL_FILE), f"missing workbook: {EXCEL_FILE}"
    return load_workbook(EXCEL_FILE, data_only=True)


def test_required_sheets_exist(wb_formula):
    assert wb_formula.sheetnames == ["BudgetAudit", "MeterReadings", "TariffTable", "BuildingInfo"]


def test_detail_and_ranking_cells_use_formulas(wb_formula):
    audit = wb_formula["BudgetAudit"]

    for row in range(7, 13):
        for col in range(3, 16):
            value = audit.cell(row, col).value
            assert isinstance(value, str) and value.startswith("="), f"BudgetAudit!{audit.cell(row, col).coordinate} should be a formula"

    for cell_ref in ["C3", "E3", "G3"]:
        value = audit[cell_ref].value
        assert isinstance(value, str) and value.startswith("="), f"BudgetAudit!{cell_ref} should be a formula"

    for row in range(18, 21):
        for col in range(1, 8):
            value = audit.cell(row, col).value
            assert isinstance(value, str) and value.startswith("="), f"BudgetAudit!{audit.cell(row, col).coordinate} should be a formula"


def test_budget_audit_values_are_correct(wb_values):
    audit = wb_values["BudgetAudit"]
    expected = _expected_rows(wb_values)

    for row, values in expected.items():
        assert audit[f"C{row}"].value == pytest.approx(values["area"], rel=0, abs=1e-9)
        assert audit[f"D{row}"].value == pytest.approx(values["budget"], rel=0, abs=1e-9)
        assert audit[f"E{row}"].value == pytest.approx(values["peak"], rel=0, abs=1e-9)
        assert audit[f"F{row}"].value == pytest.approx(values["flat"], rel=0, abs=1e-9)
        assert audit[f"G{row}"].value == pytest.approx(values["valley"], rel=0, abs=1e-9)
        assert audit[f"H{row}"].value == pytest.approx(values["peak_price"], rel=0, abs=1e-9)
        assert audit[f"I{row}"].value == pytest.approx(values["flat_price"], rel=0, abs=1e-9)
        assert audit[f"J{row}"].value == pytest.approx(values["valley_price"], rel=0, abs=1e-9)
        assert audit[f"K{row}"].value == pytest.approx(values["total_cost"], rel=0, abs=1e-9)
        assert audit[f"L{row}"].value == pytest.approx(values["cost_per_sqm"], rel=0, abs=1e-9)
        assert audit[f"M{row}"].value == pytest.approx(values["variance"], rel=0, abs=1e-9)
        assert audit[f"N{row}"].value == pytest.approx(values["variance_pct"], rel=0, abs=1e-9)
        assert audit[f"O{row}"].value == values["status"]


def test_kpi_summary_is_correct(wb_values):
    audit = wb_values["BudgetAudit"]
    expected = _expected_rows(wb_values)

    total_cost = sum(item["total_cost"] for item in expected.values())
    over_count = sum(1 for item in expected.values() if item["status"] == "OVER")
    largest_variance = max(item["variance"] for item in expected.values())

    assert audit["C3"].value == pytest.approx(total_cost, rel=0, abs=1e-9)
    assert audit["E3"].value == pytest.approx(over_count, rel=0, abs=1e-9)
    assert audit["G3"].value == pytest.approx(largest_variance, rel=0, abs=1e-9)


def test_over_budget_ranking_is_correct(wb_values):
    audit = wb_values["BudgetAudit"]
    expected = _expected_rows(wb_values)

    ranked = []
    for row, values in expected.items():
        if values["variance"] > 0:
            ranked.append(
                {
                    "building_id": audit[f"A{row}"].value,
                    "name": values["name"],
                    "variance": values["variance"],
                    "variance_pct": values["variance_pct"],
                    "total_cost": values["total_cost"],
                    "cost_per_sqm": values["cost_per_sqm"],
                }
            )

    ranked.sort(key=lambda item: item["variance"], reverse=True)
    top3 = ranked[:3]

    for idx, row in enumerate(range(18, 21), start=1):
        item = top3[idx - 1]
        assert audit[f"A{row}"].value == idx
        assert audit[f"B{row}"].value == item["building_id"]
        assert audit[f"C{row}"].value == item["name"]
        assert audit[f"D{row}"].value == pytest.approx(item["variance"], rel=0, abs=1e-9)
        assert audit[f"E{row}"].value == pytest.approx(item["variance_pct"], rel=0, abs=1e-9)
        assert audit[f"F{row}"].value == pytest.approx(item["total_cost"], rel=0, abs=1e-9)
        assert audit[f"G{row}"].value == pytest.approx(item["cost_per_sqm"], rel=0, abs=1e-9)


def test_no_excel_error_strings_after_recalc(wb_values):
    for sheet in wb_values.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    assert not cell.value.startswith("#"), f"Excel error left at {sheet.title}!{cell.coordinate}: {cell.value}"
